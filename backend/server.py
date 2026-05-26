from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
import json
import re
from contextvars import ContextVar
from datetime import datetime, timezone, timedelta

from auth import (
    User,
    UserPublic,
    RegisterRequest,
    LoginRequest,
    AuthResponse,
    hash_password,
    verify_password,
    create_token,
    decode_token,
)
from email_sender import send_password_reset_code, send_feedback_email

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
real_db = client[os.environ['DB_NAME']]

EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')

# ---------- Per-request user context ----------
current_user_id_var: ContextVar[Optional[str]] = ContextVar("current_user_id", default=None)

# When the current user is FREE-tier with > 15 tools, this is the set of
# tool ids they're allowed to see (the 15 oldest). For PRO / lifetime users
# OR free users with <= 15 tools this stays None and the proxy applies no
# extra filter. Computed once per request in the middleware after auth.
free_visible_tool_ids_var: ContextVar[Optional[set]] = ContextVar(
    "free_visible_tool_ids", default=None,
)

# Collections that store rows referencing a tool via `tool_id`. When the
# free-tier filter is active, queries against these collections are also
# narrowed so a free user never sees a claim / maintenance entry / etc.
# tied to a hidden tool. Anything not in this set falls back to the
# normal owner_id-only scope.
TOOL_REF_COLLECTIONS = {
    "claims",
    "claim_events",
    "claim_payments",
    "maintenance",
    "maintenance_logs",
    "checkouts",
    "checkout_history",
    "warranty_claims",
    "photos",
    "documents",
    "tool_history",
    "tool_changes",
}


class _ScopedCollection:
    """Wraps a Motor collection so all queries/inserts are auto-filtered by owner_id."""

    def __init__(self, base, user_id: str, name: str = ""):
        self._base = base
        self._uid = user_id
        self._name = name

    def _scope(self, q=None):
        q = dict(q or {})
        q["owner_id"] = self._uid
        # Free-tier per-request lockdown — only applied to the `tools`
        # collection itself (id-based) and to tables that reference a
        # tool via `tool_id`. PRO users always have None here so nothing
        # changes for them.
        visible = free_visible_tool_ids_var.get()
        if visible is not None:
            if self._name == "tools":
                # Merge with any existing `id` filter the caller supplied.
                if "id" in q:
                    existing = q["id"]
                    if isinstance(existing, str):
                        if existing not in visible:
                            # Force the query to return nothing — caller asked
                            # for a specific tool that's hidden.
                            q["id"] = {"$in": []}
                    elif isinstance(existing, dict) and "$in" in existing:
                        q["id"] = {"$in": [i for i in existing["$in"] if i in visible]}
                    # else: leave more complex operators alone
                else:
                    q["id"] = {"$in": list(visible)}
            elif self._name in TOOL_REF_COLLECTIONS:
                # Same merging logic but on `tool_id` field.
                if "tool_id" in q:
                    existing = q["tool_id"]
                    if isinstance(existing, str):
                        if existing not in visible:
                            q["tool_id"] = {"$in": []}
                    elif isinstance(existing, dict) and "$in" in existing:
                        q["tool_id"] = {"$in": [i for i in existing["$in"] if i in visible]}
                else:
                    q["tool_id"] = {"$in": list(visible)}
        return q

    def find(self, q=None, *args, **kw):
        return self._base.find(self._scope(q), *args, **kw)

    async def find_one(self, q, *args, **kw):
        return await self._base.find_one(self._scope(q), *args, **kw)

    async def insert_one(self, doc):
        d = dict(doc)
        d["owner_id"] = self._uid
        return await self._base.insert_one(d)

    async def insert_many(self, docs):
        docs = [{**d, "owner_id": self._uid} for d in docs]
        return await self._base.insert_many(docs)

    async def update_one(self, q, *args, **kw):
        return await self._base.update_one(self._scope(q), *args, **kw)

    async def update_many(self, q, *args, **kw):
        return await self._base.update_many(self._scope(q), *args, **kw)

    async def delete_one(self, q):
        return await self._base.delete_one(self._scope(q))

    async def delete_many(self, q):
        return await self._base.delete_many(self._scope(q))

    async def count_documents(self, q):
        return await self._base.count_documents(self._scope(q))

    def aggregate(self, pipeline):
        match = {"owner_id": self._uid}
        visible = free_visible_tool_ids_var.get()
        if visible is not None:
            if self._name == "tools":
                match["id"] = {"$in": list(visible)}
            elif self._name in TOOL_REF_COLLECTIONS:
                match["tool_id"] = {"$in": list(visible)}
        scoped = [{"$match": match}, *list(pipeline)]
        return self._base.aggregate(scoped)


class _DBProxy:
    """Drop-in replacement for `db` that auto-scopes by current user."""

    def __getattr__(self, name):
        coll = real_db[name]
        uid = current_user_id_var.get()
        if uid:
            return _ScopedCollection(coll, uid, name=name)
        return coll

    def __getitem__(self, name):
        return self.__getattr__(name)


db = _DBProxy()


# ---------- Auth dependency / helpers ----------
async def get_current_user(request: Request) -> User:
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(401, "Not authenticated")
    token = auth[7:].strip()
    uid = decode_token(token)
    udoc = await real_db.users.find_one({"id": uid}, {"_id": 0})
    if not udoc:
        raise HTTPException(401, "User not found")
    return User(**udoc)


def to_public(u: User) -> UserPublic:
    return UserPublic(
        id=u.id,
        email=u.email,
        name=u.name or "",
        created_at=u.created_at,
    )


PUBLIC_PATHS = ("/api/auth/", "/api/health", "/api/", "/api/feedback")


# ---------------------------------------------------------------------------
# Generic in-memory rate limiter (used by auth, AI, PDF and feedback endpoints)
# ---------------------------------------------------------------------------
# We deliberately keep this in-process and dependency-free. The data is a dict
# of dicts: { "bucket_name": { "key": [timestamp, ...] } }. "key" is whichever
# identifier we want to limit on (user_id, IP, etc).
#
# Rationale for not using slowapi/redis:
#   - Single worker today; if we scale to N workers each one will allow ~N×
#     the limit. That's acceptable for our threat model (brute force, AI cost
#     protection) and avoids the operational overhead of running Redis.
#   - Trivially testable.
_rate_limit_buckets: Dict[str, Dict[str, List[float]]] = {}


def _client_ip(request: Request) -> str:
    """Best-effort client IP, honouring X-Forwarded-For from K8s ingress."""
    xff = (request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
    return xff or (request.client.host if request.client else "unknown")


def _rate_limit(
    bucket: str,
    key: str,
    *,
    max_count: int,
    window_seconds: int,
) -> bool:
    """Return True if the request is *under* the limit (allowed),
    False if rate-limited. Caller should raise HTTPException(429) on False.
    """
    import time as _time
    now = _time.time()
    cutoff = now - window_seconds
    by_key = _rate_limit_buckets.setdefault(bucket, {})
    recent = [t for t in by_key.get(key, []) if t > cutoff]
    if len(recent) >= max_count:
        by_key[key] = recent
        return False
    recent.append(now)
    by_key[key] = recent
    return True


def _enforce_rate_limit(
    bucket: str,
    key: str,
    *,
    max_count: int,
    window_seconds: int,
    message: str,
) -> None:
    """Convenience wrapper — raises HTTPException(429) if over the limit."""
    if not _rate_limit(
        bucket,
        key,
        max_count=max_count,
        window_seconds=window_seconds,
    ):
        raise HTTPException(429, message)


app = FastAPI()


@app.middleware("http")
async def attach_user_to_context(request: Request, call_next):
    """Read JWT from Authorization header and set the current user id in context.
    Also enforces auth on all /api/* routes except /api/auth/* and /api/health.
    """
    path = request.url.path
    if not path.startswith("/api/"):
        return await call_next(request)
    # Public endpoints — no auth required
    if (
        path.startswith("/api/auth/")
        or path == "/api/"
        or path == "/api/health"
        or path == "/api/feedback"
        or path == "/api/guides"
        or path == "/api/revenuecat/webhook"
        or path.startswith("/api/migration/")
    ):
        return await call_next(request)
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return JSONResponse({"detail": "Not authenticated"}, status_code=401)
    token = auth[7:].strip()
    try:
        uid = decode_token(token)
    except HTTPException as e:
        return JSONResponse({"detail": e.detail}, status_code=e.status_code)
    except Exception:
        return JSONResponse({"detail": "Invalid token"}, status_code=401)
    token_var = current_user_id_var.set(uid)
    visible_var = None
    try:
        # Free-tier visibility cap — must be computed BEFORE handlers run
        # so the _ScopedCollection proxy can read it. Skipped for fast
        # endpoints that don't touch tools to keep auth-only requests
        # snappy. Errors here are swallowed so they never break a
        # request — pessimistic mode = no filter (PRO behaviour).
        try:
            from subscriptions import is_pro as _is_pro
            if not await _is_pro(real_db, uid):
                # Count first; if <= 15, no filter needed.
                cnt = await real_db.tools.count_documents({"owner_id": uid})
                if cnt > 15:
                    cursor = real_db.tools.find(
                        {"owner_id": uid},
                        {"id": 1, "_id": 0},
                    ).sort("created_at", 1).limit(15)
                    ids = {doc["id"] async for doc in cursor if "id" in doc}
                    visible_var = free_visible_tool_ids_var.set(ids)
        except Exception:
            pass
        return await call_next(request)
    finally:
        current_user_id_var.reset(token_var)
        if visible_var is not None:
            free_visible_tool_ids_var.reset(visible_var)


api_router = APIRouter(prefix="/api")


def now_iso():
    return datetime.now(timezone.utc).isoformat()


# ---------- Models ----------
class Location(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = ""
    parent_id: Optional[str] = None  # for nested locations
    parent_layout_id: Optional[str] = None  # if this location is a drawer in a toolbox
    drawer_index: Optional[int] = None
    created_at: str = Field(default_factory=now_iso)


class LocationCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    parent_id: Optional[str] = None
    parent_layout_id: Optional[str] = None
    drawer_index: Optional[int] = None


class LocationUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    parent_id: Optional[str] = None


class Tag(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    color: Optional[str] = "#FFB300"
    created_at: str = Field(default_factory=now_iso)


class TagCreate(BaseModel):
    name: str
    color: Optional[str] = "#FFB300"


class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    created_at: str = Field(default_factory=now_iso)


class CategoryCreate(BaseModel):
    name: str


class Borrower(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class BorrowerCreate(BaseModel):
    name: str
    contact: Optional[str] = ""


# Dealer & Agents
class Agent(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: Optional[str] = ""
    email: Optional[str] = ""
    location: Optional[str] = ""  # e.g. "North Houston", "Route 12", etc.
    notes: Optional[str] = ""
    started_at: str = Field(default_factory=now_iso)
    ended_at: Optional[str] = None  # when this agent stopped being current


class AgentCreate(BaseModel):
    name: str
    phone: Optional[str] = ""
    email: Optional[str] = ""
    location: Optional[str] = ""
    notes: Optional[str] = ""


class Dealer(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: Optional[str] = ""
    website: Optional[str] = ""
    address: Optional[str] = ""
    notes: Optional[str] = ""
    # Additional company contact channels (free-form: phone, email, URL, or note)
    warranty_contact: Optional[str] = ""
    tech_support_contact: Optional[str] = ""
    customer_support_contact: Optional[str] = ""
    route_frequency: Optional[str] = "N/A"  # Weekly | Bi-weekly | Monthly | N/A
    route_day_of_week: Optional[str] = ""  # Mon | Tue | Wed | Thu | Fri | Sat | Sun (when frequency is Weekly/Bi-weekly)
    route_anchor_date: Optional[str] = ""  # YYYY-MM-DD anchor used to compute next visit (mainly for Bi-weekly/Monthly)
    agents: List[Agent] = []
    current_agent_id: Optional[str] = None
    credit_balance: float = 0.0
    personal_balance: float = 0.0
    transactions: List[Dict[str, Any]] = []  # list of BalanceTransaction
    created_at: str = Field(default_factory=now_iso)


class BalanceTransaction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    account: str  # "credit" or "personal"
    type: str  # "payment" (decreases balance) or "charge" (increases balance)
    amount: float
    note: Optional[str] = ""
    date: str = Field(default_factory=lambda: datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    created_at: str = Field(default_factory=now_iso)


class TransactionCreate(BaseModel):
    account: str  # "credit" | "personal"
    type: str  # "payment" | "charge"
    amount: float
    note: Optional[str] = ""
    date: Optional[str] = None


class DealerCreate(BaseModel):
    name: str
    phone: Optional[str] = ""
    website: Optional[str] = ""
    address: Optional[str] = ""
    notes: Optional[str] = ""
    warranty_contact: Optional[str] = ""
    tech_support_contact: Optional[str] = ""
    customer_support_contact: Optional[str] = ""
    route_frequency: Optional[str] = "N/A"
    route_day_of_week: Optional[str] = ""
    route_anchor_date: Optional[str] = ""


class DealerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    warranty_contact: Optional[str] = None
    tech_support_contact: Optional[str] = None
    customer_support_contact: Optional[str] = None
    route_frequency: Optional[str] = None
    route_day_of_week: Optional[str] = None
    route_anchor_date: Optional[str] = None


# Documents
class Document(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    data: str  # base64
    mime_type: Optional[str] = "application/octet-stream"
    size: Optional[int] = 0  # bytes
    uploaded_at: str = Field(default_factory=now_iso)


# Maintenance / Calibration tracking
class ServiceEvent(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str  # YYYY-MM-DD
    cost: Optional[float] = 0.0
    technician: Optional[str] = ""
    notes: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)


class MaintenanceSchedule(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: str = "Service"  # "Calibration" / "Service" / "Inspection" / custom
    interval_months: int = 12
    last_done_date: Optional[str] = None  # YYYY-MM-DD
    next_due_date: Optional[str] = None  # auto-calculated
    notes: Optional[str] = ""
    history: List[ServiceEvent] = []
    created_at: str = Field(default_factory=now_iso)


class MaintenanceScheduleCreate(BaseModel):
    type: str = "Service"
    interval_months: int = 12
    last_done_date: Optional[str] = None
    notes: Optional[str] = ""


class MaintenanceScheduleUpdate(BaseModel):
    type: Optional[str] = None
    interval_months: Optional[int] = None
    last_done_date: Optional[str] = None
    notes: Optional[str] = None


class ServiceEventCreate(BaseModel):
    date: Optional[str] = None  # default today
    cost: Optional[float] = 0.0
    technician: Optional[str] = ""
    notes: Optional[str] = ""


# Theft / Loss reporting
class LostStatus(BaseModel):
    is_lost: bool = False
    type: str = "lost"  # "lost" or "stolen"
    reported_date: Optional[str] = None  # YYYY-MM-DD
    police_report_number: Optional[str] = ""
    insurance_company: Optional[str] = ""
    insurance_claim_number: Optional[str] = ""
    notes: Optional[str] = ""
    reported_by: Optional[str] = ""
    recovered_at: Optional[str] = None  # if recovered, set to ISO


class ReportLostRequest(BaseModel):
    type: str = "lost"  # lost or stolen
    reported_date: Optional[str] = None
    police_report_number: Optional[str] = ""
    insurance_company: Optional[str] = ""
    insurance_claim_number: Optional[str] = ""
    notes: Optional[str] = ""
    reported_by: Optional[str] = ""


# Warranty
class Warranty(BaseModel):
    has_warranty: bool = False
    provider: Optional[str] = ""
    contact: Optional[str] = ""
    terms: Optional[str] = ""
    length_months: Optional[int] = 0
    coverage_type: Optional[str] = "months"  # "months" | "limited" | "lifetime"
    start_date: Optional[str] = ""  # YYYY-MM-DD
    expiry_date: Optional[str] = ""  # YYYY-MM-DD
    document: Optional[Document] = None


# Repair info
class RepairInfo(BaseModel):
    company_notified: Optional[str] = ""  # legacy — auto-derived from dealer now
    notified_at: Optional[str] = ""  # date or ISO
    expected_completion: Optional[str] = ""  # date
    repair_status: Optional[str] = "Not Reported"  # Not Reported / Reported / In Repair / Awaiting Parts / Repaired
    contact: Optional[str] = ""  # legacy — auto-derived from agent now
    notes: Optional[str] = ""
    broken_photo: Optional[str] = ""  # extra photo, only shown on broken-item view
    # Repair / replacement cost — defaults to 0; user enters dollar amount if
    # they paid for the repair / replacement out of pocket. Feeds into the
    # Repair Cost Report and Year End Report.
    repair_cost: Optional[float] = 0.0


# Warranty claim — long-lived record that survives "Mark Repaired"
CLAIM_STATUSES = ["broken", "awaiting_approval", "waiting_replacement", "completed", "rejected"]


class WarrantyClaim(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    tool_id: str
    tool_name: str = ""
    tool_photo: Optional[str] = None
    broken_photo: Optional[str] = ""
    dealer_id: Optional[str] = None
    dealer_name: str = ""
    repair_company: Optional[str] = ""
    contact: Optional[str] = ""
    notified_at: Optional[str] = ""
    expected_completion: Optional[str] = ""
    claim_status: str = "broken"
    notes: Optional[str] = ""
    # Cost the user paid out of pocket for this repair / replacement.
    # 0 = free (e.g., under warranty). Persisted on the claim so it survives
    # after the tool is marked Repaired and repair_info is cleared.
    repair_cost: Optional[float] = 0.0
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)
    completed_at: Optional[str] = None


class WarrantyClaimUpdate(BaseModel):
    claim_status: Optional[str] = None
    repair_company: Optional[str] = None
    contact: Optional[str] = None
    notified_at: Optional[str] = None
    expected_completion: Optional[str] = None
    notes: Optional[str] = None
    repair_cost: Optional[float] = None


# Wish list — tools the user wants to buy
class WishlistItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    url: Optional[str] = ""
    description: Optional[str] = ""
    price: Optional[float] = None
    dealer_id: Optional[str] = None
    dealer_name: Optional[str] = ""
    priority: Optional[str] = "normal"  # low / normal / high
    notes: Optional[str] = ""
    # Optional product details — used when converting to a Tool so the user
    # doesn't have to re-enter info they already saved on the wish.
    model_number: Optional[str] = ""
    photos: List[str] = []  # base64 data-URI strings (same format as Tool.photos)
    purchased: bool = False
    purchased_at: Optional[str] = None
    converted_tool_id: Optional[str] = None
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)


class WishlistItemCreate(BaseModel):
    name: str
    url: Optional[str] = ""
    description: Optional[str] = ""
    price: Optional[float] = None
    dealer_id: Optional[str] = None
    dealer_name: Optional[str] = ""
    priority: Optional[str] = "normal"
    notes: Optional[str] = ""
    model_number: Optional[str] = ""
    photos: List[str] = []


class WishlistItemUpdate(BaseModel):
    name: Optional[str] = None
    url: Optional[str] = None
    description: Optional[str] = None
    price: Optional[float] = None
    dealer_id: Optional[str] = None
    dealer_name: Optional[str] = None
    priority: Optional[str] = None
    notes: Optional[str] = None
    model_number: Optional[str] = None
    photos: Optional[List[str]] = None
    purchased: Optional[bool] = None


# Consumable
class ConsumableInfo(BaseModel):
    store_name: Optional[str] = ""
    website: Optional[str] = ""
    sku: Optional[str] = ""
    notes: Optional[str] = ""


# Checkout
class CheckoutRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    borrower_name: str
    borrower_id: Optional[str] = None
    checked_out_at: str = Field(default_factory=now_iso)
    checked_in_at: Optional[str] = None
    notes: Optional[str] = ""


class Tool(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = ""
    brand: Optional[str] = ""
    # Legacy single-value fields (kept for backward compat with older app
    # builds and CSV import). New code reads/writes model_numbers / serial_numbers
    # arrays. See _resolve_model_serial_arrays() for the migration glue.
    model: Optional[str] = ""
    serial_number: Optional[str] = ""
    # New multi-value fields — users can stack any number of model #s and
    # serial #s per tool (e.g., for kits, replacement parts, etc).
    model_numbers: List[str] = []
    serial_numbers: List[str] = []
    is_set: bool = False
    set_serials: List[str] = []
    cost: Optional[float] = 0.0
    # Manufacturer's Suggested Retail Price — informational, used as an
    # optional column on Insurance / Inventory / Lost-Stolen / Year End
    # reports (alongside or instead of the user's actual purchase cost).
    # 0 = not set; user fills this in if they want MSRP totals in reports.
    msrp_price: Optional[float] = 0.0
    quantity: Optional[int] = 1
    purchase_date: Optional[str] = ""
    condition: Optional[str] = "Good"
    location_id: Optional[str] = None
    location_name: Optional[str] = ""
    category_id: Optional[str] = None
    category_name: Optional[str] = ""
    tag_ids: List[str] = []
    tag_names: List[str] = []
    photos: List[str] = []
    documents: List[Document] = []
    receipts: List[str] = []  # base64 receipt photos auto-saved by AI scanner

    is_consumable: bool = False
    consumable_info: Optional[ConsumableInfo] = None
    needs_repair: bool = False
    repair_info: Optional[RepairInfo] = None
    warranty: Optional[Warranty] = None
    dealer_id: Optional[str] = None
    dealer_name: Optional[str] = ""
    purchased_from_agent_id: Optional[str] = None  # snapshot at purchase
    purchased_from_agent_name: Optional[str] = ""
    is_checked_out: bool = False
    current_checkout: Optional[CheckoutRecord] = None
    checkout_history: List[CheckoutRecord] = []
    maintenance: List[MaintenanceSchedule] = []
    lost_status: Optional[LostStatus] = None
    # Sale tracking
    for_sale: bool = False
    sale_price: Optional[float] = 0.0
    sale_listed_at: Optional[str] = ""
    sale_notes: Optional[str] = ""
    is_sold: bool = False
    sold_at: Optional[str] = ""
    sold_price: Optional[float] = 0.0
    sold_to: Optional[str] = ""
    sold_notes: Optional[str] = ""
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)


class ToolCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    brand: Optional[str] = ""
    model: Optional[str] = ""
    serial_number: Optional[str] = ""
    model_numbers: Optional[List[str]] = None
    serial_numbers: Optional[List[str]] = None
    is_set: bool = False
    set_serials: List[str] = []
    cost: Optional[float] = 0.0
    msrp_price: Optional[float] = 0.0
    quantity: Optional[int] = 1
    purchase_date: Optional[str] = ""
    condition: Optional[str] = "Good"
    location_id: Optional[str] = None
    location_name: Optional[str] = ""
    category_id: Optional[str] = None
    category_name: Optional[str] = ""
    tag_ids: List[str] = []
    tag_names: List[str] = []
    photos: List[str] = []
    receipts: List[str] = []

    documents: List[Document] = []
    is_consumable: bool = False
    consumable_info: Optional[ConsumableInfo] = None
    needs_repair: bool = False
    repair_info: Optional[RepairInfo] = None
    warranty: Optional[Warranty] = None
    dealer_id: Optional[str] = None
    dealer_name: Optional[str] = ""
    purchased_from_agent_id: Optional[str] = None
    purchased_from_agent_name: Optional[str] = ""


class ToolUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    brand: Optional[str] = None
    model: Optional[str] = None
    serial_number: Optional[str] = None
    model_numbers: Optional[List[str]] = None
    serial_numbers: Optional[List[str]] = None
    is_set: Optional[bool] = None
    set_serials: Optional[List[str]] = None
    cost: Optional[float] = None
    msrp_price: Optional[float] = None
    quantity: Optional[int] = None
    purchase_date: Optional[str] = None
    condition: Optional[str] = None
    location_id: Optional[str] = None
    location_name: Optional[str] = None
    category_id: Optional[str] = None
    category_name: Optional[str] = None
    tag_ids: Optional[List[str]] = None
    tag_names: Optional[List[str]] = None
    photos: Optional[List[str]] = None
    documents: Optional[List[Document]] = None
    receipts: Optional[List[str]] = None
    is_consumable: Optional[bool] = None
    consumable_info: Optional[ConsumableInfo] = None
    needs_repair: Optional[bool] = None
    repair_info: Optional[RepairInfo] = None
    warranty: Optional[Warranty] = None
    dealer_id: Optional[str] = None
    dealer_name: Optional[str] = None
    purchased_from_agent_id: Optional[str] = None
    purchased_from_agent_name: Optional[str] = None
    # Sale tracking — included so PUT /tools/{id} can edit these fields
    for_sale: Optional[bool] = None
    sale_price: Optional[float] = None
    sale_listed_at: Optional[str] = None
    sale_notes: Optional[str] = None
    is_sold: Optional[bool] = None
    sold_at: Optional[str] = None
    sold_price: Optional[float] = None
    sold_to: Optional[str] = None
    sold_notes: Optional[str] = None


class CheckoutRequest(BaseModel):
    borrower_name: str
    borrower_id: Optional[str] = None
    notes: Optional[str] = ""


# ---------- Helpers ----------
def build_tool_query(
    search: Optional[str] = None,
    location_id: Optional[str] = None,
    tag_id: Optional[str] = None,
    category_id: Optional[str] = None,
    dealer_id: Optional[str] = None,
    checked_out: Optional[bool] = None,
    is_consumable: Optional[bool] = None,
    needs_repair: Optional[bool] = None,
    for_sale: Optional[bool] = None,
    is_sold: Optional[bool] = None,
):
    query: Dict[str, Any] = {}
    if search:
        rx = {"$regex": re.escape(search), "$options": "i"}
        query["$or"] = [
            {"name": rx},
            {"description": rx},
            {"brand": rx},
            {"model": rx},
            {"serial_number": rx},
            {"set_serials": rx},
            {"model_numbers": rx},
            {"serial_numbers": rx},
            {"tag_names": rx},
            {"location_name": rx},
            {"category_name": rx},
            {"dealer_name": rx},
            {"purchased_from_agent_name": rx},
            {"sold_to": rx},
        ]
    if location_id:
        query["location_id"] = location_id
    if tag_id:
        query["tag_ids"] = tag_id
    if category_id:
        query["category_id"] = category_id
    if dealer_id:
        query["dealer_id"] = dealer_id
    if checked_out is not None:
        query["is_checked_out"] = checked_out
    if is_consumable is not None:
        query["is_consumable"] = is_consumable
    if needs_repair is not None:
        query["needs_repair"] = needs_repair
    if for_sale is not None:
        query["for_sale"] = for_sale
    if is_sold is not None:
        query["is_sold"] = is_sold
    # By default, exclude sold items from regular tool listings unless
    # explicitly asked for them. They live in the "sold" archive instead.
    if is_sold is None:
        query["is_sold"] = {"$ne": True}
    return query


# ---------- Root ----------
@api_router.get("/")
async def root():
    return {"message": "Toolbox Vault API"}


@api_router.get("/health")
async def health():
    """Lightweight health probe for monitoring & uptime checks. Public."""
    return {"status": "ok", "service": "toolbox-vault-api"}


# ---------- Locations ----------
@api_router.post("/locations", response_model=Location)
async def create_location(payload: LocationCreate):
    loc = Location(**payload.dict())
    await db.locations.insert_one(loc.dict())
    return loc


@api_router.get("/locations", response_model=List[Location])
async def list_locations():
    items = await db.locations.find({}, {"_id": 0}).to_list(2000)
    return [Location(**i) for i in items]


@api_router.put("/locations/{loc_id}", response_model=Location)
async def update_location(loc_id: str, payload: LocationUpdate):
    doc = await db.locations.find_one({"id": loc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Location not found")
    # Build updates preserving explicit null for parent_id (to move to root)
    raw = payload.dict(exclude_unset=True)
    updates = {}
    for k, v in raw.items():
        if k == "parent_id":
            updates[k] = v  # keep None for root move
        elif v is not None:
            updates[k] = v
    # Prevent cycles: ensure new parent isn't a descendant of this location
    if "parent_id" in updates and updates["parent_id"]:
        if updates["parent_id"] == loc_id:
            raise HTTPException(400, "Location cannot be its own parent")
        # Walk up the chain to ensure no cycle
        cur = updates["parent_id"]
        depth = 0
        while cur and depth < 50:
            p = await db.locations.find_one({"id": cur}, {"_id": 0, "parent_id": 1, "id": 1})
            if not p:
                break
            if p.get("id") == loc_id:
                raise HTTPException(400, "Cannot create a cycle in locations")
            cur = p.get("parent_id")
            depth += 1
    await db.locations.update_one({"id": loc_id}, {"$set": updates})
    new_doc = await db.locations.find_one({"id": loc_id}, {"_id": 0})

    # If the NAME changed, propagate the new name to every tool that has
    # this location_id cached. Otherwise tools keep showing the old name
    # forever (bug reported 2026-05-23: rename "test" → "test2", tool
    # description card still shows "test").
    if "name" in updates and updates["name"] != doc.get("name"):
        await db.tools.update_many(
            {"location_id": loc_id},
            {"$set": {"location_name": updates["name"]}},
        )

    return Location(**new_doc)


@api_router.delete("/locations/{loc_id}")
async def delete_location(loc_id: str, cascade: bool = False):
    if cascade:
        # collect this id and all descendants iteratively
        all_ids = [loc_id]
        frontier = [loc_id]
        while frontier:
            children = await db.locations.find(
                {"parent_id": {"$in": frontier}}, {"_id": 0, "id": 1}
            ).to_list(5000)
            ids = [c["id"] for c in children]
            if not ids:
                break
            all_ids.extend(ids)
            frontier = ids
        await db.locations.delete_many({"id": {"$in": all_ids}})
        return {"ok": True, "deleted": len(all_ids)}
    # default: only delete if no children, else reparent children to this loc's parent
    doc = await db.locations.find_one({"id": loc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Location not found")
    parent_of_deleted = doc.get("parent_id")
    await db.locations.update_many(
        {"parent_id": loc_id}, {"$set": {"parent_id": parent_of_deleted}}
    )
    await db.locations.delete_one({"id": loc_id})
    return {"ok": True}


# ---------- Tags ----------
@api_router.post("/tags", response_model=Tag)
async def create_tag(payload: TagCreate):
    name = payload.name.strip()
    existing = await db.tags.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}, {"_id": 0})
    if existing:
        return Tag(**existing)
    t = Tag(name=name, color=payload.color or "#FFB300")
    await db.tags.insert_one(t.dict())
    return t


@api_router.get("/tags", response_model=List[Tag])
async def list_tags():
    items = await db.tags.find({}, {"_id": 0}).sort("name", 1).to_list(2000)
    return [Tag(**i) for i in items]


@api_router.put("/tags/{tag_id}", response_model=Tag)
async def update_tag(tag_id: str, payload: TagCreate):
    doc = await db.tags.find_one({"id": tag_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Tag not found")
    new_name = (payload.name or "").strip()
    if not new_name:
        raise HTTPException(400, "Name required")
    old_name = doc.get("name") or ""
    update = {"name": new_name}
    if payload.color:
        update["color"] = payload.color
    await db.tags.update_one({"id": tag_id}, {"$set": update})
    if old_name and old_name != new_name:
        # Rename references on tools (tag_names is a list of strings)
        await db.tools.update_many(
            {"tag_names": old_name},
            {"$set": {"tag_names.$[el]": new_name}},
            array_filters=[{"el": old_name}],
        )
    new = await db.tags.find_one({"id": tag_id}, {"_id": 0})
    return Tag(**new)


@api_router.delete("/tags/{tag_id}")
async def delete_tag(tag_id: str):
    res = await db.tags.delete_one({"id": tag_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Tag not found")
    return {"ok": True}


# ---------- Categories ----------
@api_router.post("/categories", response_model=Category)
async def create_category(payload: CategoryCreate):
    name = payload.name.strip()
    existing = await db.categories.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}, {"_id": 0})
    if existing:
        return Category(**existing)
    c = Category(name=name)
    await db.categories.insert_one(c.dict())
    return c


@api_router.get("/categories", response_model=List[Category])
async def list_categories():
    items = await db.categories.find({}, {"_id": 0}).sort("name", 1).to_list(2000)
    return [Category(**i) for i in items]


@api_router.put("/categories/{cat_id}", response_model=Category)
async def update_category(cat_id: str, payload: CategoryCreate):
    doc = await db.categories.find_one({"id": cat_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Category not found")
    new_name = (payload.name or "").strip()
    if not new_name:
        raise HTTPException(400, "Name required")
    old_name = doc.get("name") or ""
    await db.categories.update_one({"id": cat_id}, {"$set": {"name": new_name}})
    if old_name and old_name != new_name:
        # Rename references on tools. Some tools store the category as a
        # legacy string field (`category`), others store the modern
        # `category_id` + cached `category_name` pair. Cover both.
        await db.tools.update_many({"category": old_name}, {"$set": {"category": new_name}})
        await db.tools.update_many(
            {"category_id": cat_id},
            {"$set": {"category_name": new_name}},
        )
    new = await db.categories.find_one({"id": cat_id}, {"_id": 0})
    return Category(**new)


@api_router.delete("/categories/{cat_id}")
async def delete_category(cat_id: str):
    res = await db.categories.delete_one({"id": cat_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Category not found")
    return {"ok": True}


# ---------- Borrowers ----------
@api_router.post("/borrowers", response_model=Borrower)
async def create_borrower(payload: BorrowerCreate):
    b = Borrower(**payload.dict())
    await db.borrowers.insert_one(b.dict())
    return b


@api_router.get("/borrowers", response_model=List[Borrower])
async def list_borrowers():
    items = await db.borrowers.find({}, {"_id": 0}).to_list(2000)
    return [Borrower(**i) for i in items]


@api_router.put("/borrowers/{borrower_id}", response_model=Borrower)
async def update_borrower(borrower_id: str, payload: BorrowerCreate):
    existing = await db.borrowers.find_one({"id": borrower_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Borrower not found")
    new_name = payload.name.strip()
    new_contact = (payload.contact or "").strip()
    old_name = existing.get("name", "")
    update_doc = {"name": new_name, "contact": new_contact}
    await db.borrowers.update_one({"id": borrower_id}, {"$set": update_doc})

    # Propagate name change across tools' checkout history & current_checkout
    if new_name and new_name != old_name:
        # Update by borrower_id (preferred) — covers any record referencing this borrower
        await db.tools.update_many(
            {"current_checkout.borrower_id": borrower_id},
            {"$set": {"current_checkout.borrower_name": new_name}},
        )
        await db.tools.update_many(
            {"checkout_history.borrower_id": borrower_id},
            {"$set": {"checkout_history.$[el].borrower_name": new_name}},
            array_filters=[{"el.borrower_id": borrower_id}],
        )
        # Also update legacy records that match by name (case-insensitive) but had no id
        rx = {"$regex": f"^{re.escape(old_name)}$", "$options": "i"}
        await db.tools.update_many(
            {"current_checkout.borrower_name": rx, "current_checkout.borrower_id": {"$in": [None, ""]}},
            {"$set": {"current_checkout.borrower_name": new_name}},
        )
        await db.tools.update_many(
            {"checkout_history.borrower_name": rx},
            {"$set": {"checkout_history.$[el].borrower_name": new_name}},
            array_filters=[{"el.borrower_name": rx, "$or": [{"el.borrower_id": {"$in": [None, ""]}}, {"el.borrower_id": borrower_id}]}],
        )

    updated = await db.borrowers.find_one({"id": borrower_id}, {"_id": 0})
    return Borrower(**updated)


@api_router.delete("/borrowers/{borrower_id}")
async def delete_borrower(borrower_id: str):
    res = await db.borrowers.delete_one({"id": borrower_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Borrower not found")
    return {"ok": True}


@api_router.get("/borrowers/{borrower_id}/history")
async def borrower_history(borrower_id: str):
    b = await db.borrowers.find_one({"id": borrower_id}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Borrower not found")
    name = b.get("name", "")
    name_rx = {"$regex": f"^{re.escape(name)}$", "$options": "i"}

    # All tools that have ever been checked out by this borrower id OR name
    tools = await db.tools.find(
        {
            "$or": [
                {"checkout_history.borrower_id": borrower_id},
                {"checkout_history.borrower_name": name_rx},
                {"current_checkout.borrower_id": borrower_id},
                {"current_checkout.borrower_name": name_rx},
            ]
        },
        {"_id": 0},
    ).to_list(5000)

    per_tool: List[Dict[str, Any]] = []
    total_checkouts = 0
    currently_held: List[Dict[str, Any]] = []
    all_records: List[Dict[str, Any]] = []

    for t in tools:
        # Collect all checkouts (history + current) attributed to this borrower
        records = []
        for r in (t.get("checkout_history") or []):
            if r.get("borrower_id") == borrower_id or (
                r.get("borrower_name", "").lower() == name.lower()
            ):
                records.append(r)
        cur = t.get("current_checkout") or {}
        is_active = bool(t.get("is_checked_out")) and (
            cur.get("borrower_id") == borrower_id
            or (cur.get("borrower_name", "").lower() == name.lower())
        )
        if is_active:
            records.append(cur)
            currently_held.append(
                {
                    "tool_id": t.get("id"),
                    "tool_name": t.get("name"),
                    "checked_out_at": cur.get("checked_out_at"),
                    "notes": cur.get("notes", ""),
                }
            )
        if not records:
            continue
        last = max(records, key=lambda r: r.get("checked_out_at", ""))
        per_tool.append(
            {
                "tool_id": t.get("id"),
                "tool_name": t.get("name"),
                "photo": (t.get("photos") or [None])[0],
                "checkout_count": len(records),
                "last_checked_out_at": last.get("checked_out_at"),
                "currently_out": is_active,
            }
        )
        total_checkouts += len(records)
        for r in records:
            all_records.append(
                {
                    "tool_id": t.get("id"),
                    "tool_name": t.get("name"),
                    "checked_out_at": r.get("checked_out_at"),
                    "checked_in_at": r.get("checked_in_at"),
                    "notes": r.get("notes", ""),
                }
            )

    per_tool.sort(key=lambda x: x["checkout_count"], reverse=True)
    all_records.sort(key=lambda r: r.get("checked_out_at") or "", reverse=True)
    return {
        "borrower": Borrower(**b).dict(),
        "total_checkouts": total_checkouts,
        "unique_tools": len(per_tool),
        "currently_held": currently_held,
        "per_tool": per_tool,
        "history": all_records[:200],
    }


# ---------- Dealers ----------
@api_router.post("/dealers", response_model=Dealer)
async def create_dealer(payload: DealerCreate, user: User = Depends(get_current_user)):
    d = Dealer(**payload.dict())
    await db.dealers.insert_one(d.dict())
    return d


@api_router.get("/dealers", response_model=List[Dealer])
async def list_dealers():
    items = await db.dealers.find({}, {"_id": 0}).sort("name", 1).to_list(2000)
    return [Dealer(**i) for i in items]


@api_router.get("/dealers/{dealer_id}", response_model=Dealer)
async def get_dealer(dealer_id: str):
    d = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Dealer not found")
    return Dealer(**d)


@api_router.put("/dealers/{dealer_id}", response_model=Dealer)
async def update_dealer(dealer_id: str, payload: DealerUpdate):
    d = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Dealer not found")
    updates = {k: v for k, v in payload.dict().items() if v is not None}
    await db.dealers.update_one({"id": dealer_id}, {"$set": updates})
    new = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})

    # Propagate name change to every tool with this dealer_id cached.
    if "name" in updates and updates["name"] != d.get("name"):
        await db.tools.update_many(
            {"dealer_id": dealer_id},
            {"$set": {"dealer_name": updates["name"]}},
        )

    return Dealer(**new)


@api_router.delete("/dealers/{dealer_id}")
async def delete_dealer(dealer_id: str):
    res = await db.dealers.delete_one({"id": dealer_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Dealer not found")
    return {"ok": True}


@api_router.post("/dealers/{dealer_id}/agents", response_model=Dealer)
async def add_agent(dealer_id: str, payload: AgentCreate, user: User = Depends(get_current_user)):
    d = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Dealer not found")
    agent = Agent(**payload.dict())
    agents = d.get("agents") or []
    agents.append(agent.dict())
    update = {"agents": agents}
    if not d.get("current_agent_id"):
        update["current_agent_id"] = agent.id
    await db.dealers.update_one({"id": dealer_id}, {"$set": update})
    new = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    return Dealer(**new)


@api_router.put("/dealers/{dealer_id}/agents/{agent_id}", response_model=Dealer)
async def update_agent(dealer_id: str, agent_id: str, payload: AgentCreate):
    d = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Dealer not found")
    agents = d.get("agents") or []
    found = False
    for a in agents:
        if a.get("id") == agent_id:
            a["name"] = payload.name
            a["phone"] = payload.phone or ""
            a["email"] = payload.email or ""
            a["location"] = payload.location or ""
            a["notes"] = payload.notes or ""
            found = True
            break
    if not found:
        raise HTTPException(404, "Agent not found")
    await db.dealers.update_one({"id": dealer_id}, {"$set": {"agents": agents}})
    new = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    return Dealer(**new)


@api_router.delete("/dealers/{dealer_id}/agents/{agent_id}", response_model=Dealer)
async def remove_agent(dealer_id: str, agent_id: str):
    d = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Dealer not found")
    agents = [a for a in (d.get("agents") or []) if a.get("id") != agent_id]
    update = {"agents": agents}
    if d.get("current_agent_id") == agent_id:
        update["current_agent_id"] = agents[0]["id"] if agents else None
    await db.dealers.update_one({"id": dealer_id}, {"$set": update})
    new = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    return Dealer(**new)


@api_router.post("/dealers/{dealer_id}/current-agent/{agent_id}", response_model=Dealer)
async def set_current_agent(dealer_id: str, agent_id: str):
    d = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Dealer not found")
    agents = d.get("agents") or []
    if not any(a.get("id") == agent_id for a in agents):
        raise HTTPException(400, "Agent not found in dealer")
    # Mark previous current's ended_at, mark this one's started_at if needed
    now = now_iso()
    for a in agents:
        if a.get("id") == d.get("current_agent_id") and a.get("id") != agent_id:
            if not a.get("ended_at"):
                a["ended_at"] = now
        if a.get("id") == agent_id:
            a["ended_at"] = None
            if not a.get("started_at"):
                a["started_at"] = now
    await db.dealers.update_one(
        {"id": dealer_id},
        {"$set": {"current_agent_id": agent_id, "agents": agents}},
    )
    new = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    return Dealer(**new)


# ---------- Dealer Balance Transactions ----------
@api_router.post("/dealers/{dealer_id}/transactions", response_model=Dealer)
async def add_dealer_transaction(dealer_id: str, payload: TransactionCreate):
    d = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Dealer not found")
    if payload.account not in ("credit", "personal"):
        raise HTTPException(400, "account must be 'credit' or 'personal'")
    if payload.type not in ("payment", "charge"):
        raise HTTPException(400, "type must be 'payment' or 'charge'")
    amount = float(payload.amount or 0)
    if amount <= 0:
        raise HTTPException(400, "amount must be > 0")
    tx = BalanceTransaction(
        account=payload.account,
        type=payload.type,
        amount=amount,
        note=payload.note or "",
        date=payload.date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
    ).model_dump()
    txs = list(d.get("transactions") or [])
    txs.append(tx)
    # Update balance: payment decreases, charge increases
    delta = -amount if payload.type == "payment" else amount
    field = "credit_balance" if payload.account == "credit" else "personal_balance"
    new_balance = float(d.get(field) or 0.0) + delta
    await db.dealers.update_one(
        {"id": dealer_id},
        {"$set": {field: new_balance, "transactions": txs}},
    )
    new = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    return Dealer(**new)


@api_router.delete("/dealers/{dealer_id}/transactions/{tx_id}", response_model=Dealer)
async def delete_dealer_transaction(dealer_id: str, tx_id: str):
    d = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Dealer not found")
    txs = list(d.get("transactions") or [])
    target = next((t for t in txs if t.get("id") == tx_id), None)
    if not target:
        raise HTTPException(404, "Transaction not found")
    # Reverse the balance impact
    amount = float(target.get("amount") or 0)
    delta = amount if target.get("type") == "payment" else -amount
    field = "credit_balance" if target.get("account") == "credit" else "personal_balance"
    new_balance = float(d.get(field) or 0.0) + delta
    txs = [t for t in txs if t.get("id") != tx_id]
    await db.dealers.update_one(
        {"id": dealer_id},
        {"$set": {field: new_balance, "transactions": txs}},
    )
    new = await db.dealers.find_one({"id": dealer_id}, {"_id": 0})
    return Dealer(**new)


# ---------- Tools ----------

# Backward-compat shim. The frontend writes model_numbers[] / serial_numbers[]
# for new builds, but older app builds still send legacy {model, serial_number,
# set_serials, is_set} fields. This helper folds either shape into a normalized
# dict that we can persist consistently. Always writes BOTH the new arrays and
# the legacy single-value mirrors so older app installs keep rendering data.
def _resolve_model_serial_arrays(
    updates: Dict[str, Any],
    existing: Optional[Dict[str, Any]] = None,
) -> None:
    """Mutates `updates` in place to keep model_numbers/serial_numbers in sync
    with the legacy model / serial_number / set_serials fields."""
    existing = existing or {}

    def _clean(arr: Any) -> List[str]:
        if not isinstance(arr, list):
            return []
        return [str(x).strip() for x in arr if x is not None and str(x).strip()]

    has_new_models = "model_numbers" in updates and updates["model_numbers"] is not None
    has_new_serials = "serial_numbers" in updates and updates["serial_numbers"] is not None
    has_legacy_set = "set_serials" in updates and updates["set_serials"] is not None
    has_legacy_serial = "serial_number" in updates
    has_legacy_model = "model" in updates

    # ----- Resolve model_numbers -----
    if has_new_models:
        mns = _clean(updates["model_numbers"])
    elif has_legacy_set or has_legacy_serial or has_legacy_model:
        # Older app sent legacy fields → derive
        candidates: List[str] = []
        if has_legacy_set:
            candidates.extend(_clean(updates.get("set_serials")))
        if has_legacy_serial and updates.get("serial_number"):
            candidates.append(str(updates["serial_number"]).strip())
        if has_legacy_model and updates.get("model"):
            candidates.append(str(updates["model"]).strip())
        # Dedupe preserving order
        seen = set()
        mns = []
        for v in candidates:
            if v and v not in seen:
                seen.add(v)
                mns.append(v)
    else:
        mns = None  # not provided in this update at all

    # ----- Resolve serial_numbers -----
    if has_new_serials:
        sns = _clean(updates["serial_numbers"])
    else:
        sns = None  # not provided in this update at all

    # Persist resolved arrays + legacy mirrors so old app builds still render
    if mns is not None:
        updates["model_numbers"] = mns
        # Legacy mirrors derived from model_numbers
        updates["serial_number"] = mns[0] if mns else ""
        updates["set_serials"] = mns
        updates["is_set"] = len(mns) > 1
        # Clear legacy `model` since we no longer accept it separately
        if "model" not in updates:
            updates["model"] = ""

    if sns is not None:
        updates["serial_numbers"] = sns


@api_router.post("/tools", response_model=Tool)
async def create_tool(payload: ToolCreate, user: User = Depends(get_current_user)):
    _validate_photo_payload(payload.photos)
    # Free-tier 15-item limit. Pro / lifetime users always pass.
    from subscriptions import enforce_tool_limit  # local import to avoid cycles
    await enforce_tool_limit(real_db, user.id)
    payload_dict = payload.dict()
    # Normalize legacy/new model & serial fields so both shapes survive.
    _resolve_model_serial_arrays(payload_dict)
    # Strip None values so Tool() applies its defaults instead of crashing
    payload_dict = {k: v for k, v in payload_dict.items() if v is not None}
    tool = Tool(**payload_dict)

    # Denormalize the *_name fields from their *_id counterparts so the
    # tool description card has the right names from the very first
    # render. Without this, freshly-created tools showed empty
    # location/dealer/category names until the next edit (bug noticed
    # while fixing the rename-cascade issue, 2026-05-23).
    if tool.location_id and not tool.location_name:
        loc = await db.locations.find_one(
            {"id": tool.location_id}, {"_id": 0, "name": 1}
        )
        tool.location_name = (loc or {}).get("name", "") or ""
    if tool.dealer_id and not tool.dealer_name:
        dl = await db.dealers.find_one(
            {"id": tool.dealer_id}, {"_id": 0, "name": 1}
        )
        tool.dealer_name = (dl or {}).get("name", "") or ""
    if tool.category_id and not tool.category_name:
        cat = await db.categories.find_one(
            {"id": tool.category_id}, {"_id": 0, "name": 1}
        )
        tool.category_name = (cat or {}).get("name", "") or ""

    await db.tools.insert_one(tool.dict())
    # If created already broken, also create a warranty claim mirror with broken_photo
    if tool.needs_repair:
        ri = (tool.repair_info or RepairInfo()).dict() if hasattr(tool.repair_info, "dict") else (tool.repair_info or {})
        if isinstance(ri, dict):
            claim = WarrantyClaim(
                tool_id=tool.id,
                tool_name=tool.name,
                tool_photo=(tool.photos or [None])[0] if tool.photos else None,
                broken_photo=ri.get("broken_photo") or "",
                dealer_id=tool.dealer_id,
                dealer_name=tool.dealer_name or "",
                repair_company=ri.get("company_notified") or "",
                contact=ri.get("contact") or "",
                notified_at=ri.get("notified_at") or "",
                expected_completion=ri.get("expected_completion") or "",
                claim_status="broken",
                notes=ri.get("notes") or "",
                repair_cost=float(ri.get("repair_cost") or 0),
            )
            await db.warranty_claims.insert_one(claim.dict())
    return tool


# ---------------------------------------------------------------------------
# CSV Import / Export
# ---------------------------------------------------------------------------

# Logical field set the import wizard maps to. Keep this list authoritative.
_IMPORT_FIELDS = [
    {"id": "name", "label": "Name *", "required": True},
    {"id": "brand", "label": "Brand"},
    {"id": "model", "label": "Model"},
    {"id": "serial_number", "label": "Model number"},
    {"id": "quantity", "label": "Quantity"},
    {"id": "cost", "label": "Cost (per unit)"},
    {"id": "msrp_price", "label": "MSRP (per unit, optional)"},
    {"id": "description", "label": "Description / Notes"},
    {"id": "category", "label": "Category (by name)"},
    {"id": "location", "label": "Location (by name)"},
    {"id": "dealer", "label": "Dealer (by name)"},
    {"id": "condition", "label": "Condition"},
    {"id": "purchase_date", "label": "Purchase date (YYYY-MM-DD)"},
    {"id": "warranty_expiry", "label": "Warranty expiry (YYYY-MM-DD)"},
    {"id": "tags", "label": "Tags (comma-separated)"},
]


@api_router.get("/tools/import-fields")
async def tools_import_fields(user: User = Depends(get_current_user)):
    return {"fields": _IMPORT_FIELDS}


# Export field registry — id, CSV header label, and how to read the value
# from a tool dict given pre-resolved {cats, locs, dlrs, tags} lookup maps.
_EXPORT_FIELDS: List[Dict[str, Any]] = [
    {"id": "name", "label": "Name"},
    {"id": "brand", "label": "Brand"},
    {"id": "model", "label": "Model"},
    {"id": "serial_number", "label": "Model number"},
    {"id": "serial_numbers", "label": "Serial number(s)"},
    {"id": "quantity", "label": "Quantity"},
    {"id": "cost", "label": "Cost"},
    {"id": "msrp_price", "label": "MSRP"},
    {"id": "category", "label": "Category"},
    {"id": "location", "label": "Location"},
    {"id": "dealer", "label": "Dealer"},
    {"id": "tags", "label": "Tags"},
    {"id": "condition", "label": "Condition"},
    {"id": "purchase_date", "label": "Purchase date"},
    {"id": "warranty_expiry", "label": "Warranty expiry"},
    {"id": "description", "label": "Description"},
    {"id": "is_consumable", "label": "Is consumable"},
    {"id": "is_set", "label": "Is set"},
    {"id": "set_serials", "label": "Set serials"},
]
_EXPORT_FIELD_IDS = [f["id"] for f in _EXPORT_FIELDS]


def _build_export_value(field_id: str, t: Dict[str, Any], lookups: Dict[str, Dict[str, str]]) -> Any:
    if field_id == "name":
        return t.get("name") or ""
    if field_id == "brand":
        return t.get("brand") or ""
    if field_id == "model":
        return t.get("model") or ""
    if field_id == "serial_number":
        # Legacy "Model number" export. Prefer the new model_numbers[] array
        # (semicolon-joined) so users see all model numbers in one cell;
        # fall back to the single legacy field for tools not yet migrated.
        mns = t.get("model_numbers") or []
        if mns:
            return "; ".join([str(x) for x in mns if x])
        return t.get("serial_number") or ""
    if field_id == "serial_numbers":
        return "; ".join([str(x) for x in (t.get("serial_numbers") or []) if x])
    if field_id == "quantity":
        return t.get("quantity") or 1
    if field_id == "cost":
        return t.get("cost") or 0
    if field_id == "msrp_price":
        return t.get("msrp_price") or 0
    if field_id == "category":
        return lookups["cats"].get(t.get("category_id") or "", "")
    if field_id == "location":
        return lookups["locs"].get(t.get("location_id") or "", "")
    if field_id == "dealer":
        return lookups["dlrs"].get(t.get("dealer_id") or "", "")
    if field_id == "tags":
        return ", ".join(
            sorted([
                lookups["tags"].get(tid, "")
                for tid in (t.get("tag_ids") or [])
                if lookups["tags"].get(tid)
            ])
        )
    if field_id == "condition":
        return t.get("condition") or ""
    if field_id == "purchase_date":
        return t.get("purchase_date") or ""
    if field_id == "warranty_expiry":
        return t.get("warranty_expiry") or ""
    if field_id == "description":
        return t.get("description") or ""
    if field_id == "is_consumable":
        return "yes" if t.get("is_consumable") else ""
    if field_id == "is_set":
        return "yes" if t.get("is_set") else ""
    if field_id == "set_serials":
        return "; ".join(t.get("set_serials") or [])
    return ""


@api_router.get("/tools/export-fields")
async def tools_export_fields(user: User = Depends(get_current_user)):
    """List of fields the user can choose from when exporting."""
    return {"fields": _EXPORT_FIELDS}


class ExportPayload(BaseModel):
    fields: Optional[List[str]] = None  # subset of _EXPORT_FIELD_IDS; None/empty = all
    format: Optional[str] = "csv"        # "csv" | "xlsx"


@api_router.post("/tools/export-csv")
async def tools_export_csv_post(
    payload: ExportPayload,
    user: User = Depends(get_current_user),
):
    """Field-customisable export. POST body:
       {fields: ["name","brand",...], format: "csv" | "xlsx"}
    Falls through to all fields when `fields` is empty/missing."""
    requested = payload.fields or []
    if requested:
        chosen = [fid for fid in requested if fid in _EXPORT_FIELD_IDS]
    else:
        chosen = list(_EXPORT_FIELD_IDS)
    if not chosen:
        chosen = list(_EXPORT_FIELD_IDS)
    fmt = (payload.format or "csv").lower().strip()
    if fmt not in ("csv", "xlsx"):
        fmt = "csv"

    return await _do_export(chosen, fmt)


@api_router.get("/tools/export-csv")
async def tools_export_csv(user: User = Depends(get_current_user)):
    """Backwards-compat GET — exports all fields as CSV."""
    return await _do_export(list(_EXPORT_FIELD_IDS), "csv")


async def _do_export(chosen_field_ids: List[str], fmt: str) -> Dict[str, Any]:
    """Shared implementation used by POST and legacy GET. `fmt` in ("csv","xlsx")."""
    tools = await db.tools.find({}, {"_id": 0}).sort("name", 1).to_list(20000)
    cat_ids = {t.get("category_id") for t in tools if t.get("category_id")}
    cats = {
        c["id"]: c.get("name") or ""
        for c in await db.categories.find(
            {"id": {"$in": list(cat_ids)}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(5000)
    } if cat_ids else {}
    loc_ids = {t.get("location_id") for t in tools if t.get("location_id")}
    locs = {
        l["id"]: l.get("name") or ""
        for l in await db.locations.find(
            {"id": {"$in": list(loc_ids)}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(5000)
    } if loc_ids else {}
    dlr_ids = {t.get("dealer_id") for t in tools if t.get("dealer_id")}
    dlrs = {
        d["id"]: d.get("name") or ""
        for d in await db.dealers.find(
            {"id": {"$in": list(dlr_ids)}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(5000)
    } if dlr_ids else {}
    all_tag_ids: List[str] = []
    for t in tools:
        all_tag_ids.extend(t.get("tag_ids") or [])
    uniq_tag_ids = list(set(all_tag_ids))
    tags = {
        tg["id"]: tg.get("name") or ""
        for tg in await db.tags.find(
            {"id": {"$in": uniq_tag_ids}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(5000)
    } if uniq_tag_ids else {}

    lookups = {"cats": cats, "locs": locs, "dlrs": dlrs, "tags": tags}
    label_for = {f["id"]: f["label"] for f in _EXPORT_FIELDS}
    headers = [label_for[fid] for fid in chosen_field_ids]
    rows_data = [
        [_build_export_value(fid, t, lookups) for fid in chosen_field_ids]
        for t in tools
    ]

    today = datetime.utcnow().strftime("%Y-%m-%d")
    import base64 as _b64

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except Exception as e:
            raise HTTPException(500, f"openpyxl not available: {e}")
        import io as _io
        wb = Workbook()
        ws = wb.active
        ws.title = "Tools"
        ws.append(headers)
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill("solid", fgColor="FFB300")
        for col_i in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_i)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(vertical="center")
        for row in rows_data:
            ws.append(row)
        # Freeze header row + auto-size columns based on content
        ws.freeze_panes = "A2"
        for col_i, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in rows_data:
                v = row[col_i - 1]
                vlen = len(str(v)) if v is not None else 0
                if vlen > max_len:
                    max_len = vlen
            # Cap width so sheet stays readable
            ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = min(
                max(12, max_len + 2), 50
            )
        buf = _io.BytesIO()
        wb.save(buf)
        raw = buf.getvalue()
        return {
            "filename": f"toolbox-vault-export-{today}.xlsx",
            "base64": _b64.b64encode(raw).decode("ascii"),
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "rows": len(tools),
            "fields": chosen_field_ids,
            "format": "xlsx",
        }

    # Default: CSV
    import csv as _csv
    import io as _io
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(headers)
    for row in rows_data:
        w.writerow(row)
    raw = buf.getvalue().encode("utf-8")
    return {
        "filename": f"toolbox-vault-export-{today}.csv",
        "base64": _b64.b64encode(raw).decode("ascii"),
        "mime": "text/csv",
        "rows": len(tools),
        "fields": chosen_field_ids,
        "format": "csv",
    }


class ImportRow(BaseModel):
    name: Optional[str] = ""
    brand: Optional[str] = ""
    model: Optional[str] = ""
    serial_number: Optional[str] = ""
    # NOTE: quantity / cost accept Any so we can tolerate values from
    # third-party CSVs like "13,500.00", "$1,200", "1.0", "1 ea", "" etc.
    # The raw value is sanitised inside tools_import() via _to_int / _to_float.
    quantity: Optional[Any] = 1
    cost: Optional[Any] = 0.0
    msrp_price: Optional[Any] = 0.0
    description: Optional[str] = ""
    category: Optional[str] = ""        # name (case-insensitive lookup; auto-create if missing)
    location: Optional[str] = ""        # name match (existing only)
    dealer: Optional[str] = ""          # name match (existing only)
    condition: Optional[str] = ""
    purchase_date: Optional[str] = ""
    warranty_expiry: Optional[str] = ""
    tags: Optional[str] = ""            # comma-separated names; auto-create if missing


class ImportPayload(BaseModel):
    rows: List[ImportRow]
    create_missing_categories: bool = True
    create_missing_tags: bool = True
    create_missing_locations: bool = True
    create_missing_dealers: bool = True


def _norm(s: Optional[str]) -> str:
    return (s or "").strip() if isinstance(s, str) else (str(s).strip() if s is not None else "")


def _to_float(v: Any) -> float:
    """Tolerant float parser — handles strings with currency symbols,
    thousand separators, percent signs, blanks, etc.
    Returns 0.0 if the value cannot be coerced."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return 0.0
    s = str(v).strip()
    if not s:
        return 0.0
    # Strip everything except digits, dot, minus, comma — then drop commas.
    # Handles "$13,500.00", "13.500,00 €" (best-effort), "1,234", "1.0"
    cleaned = "".join(ch for ch in s if ch.isdigit() or ch in ".-,")
    # If both ',' and '.' present, assume ',' is thousand-sep (US/UK format)
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned and "." not in cleaned:
        # Could be European decimal (e.g. "13,50") OR a thousands-sep ("13,500").
        # Heuristic: if there are exactly 3 digits after the comma and no other
        # comma, treat as thousands (e.g. 13,500 → 13500). Otherwise treat
        # the comma as a decimal point.
        parts = cleaned.split(",")
        if len(parts) == 2 and len(parts[1]) == 3:
            cleaned = cleaned.replace(",", "")
        else:
            cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned) if cleaned not in ("", "-", ".") else 0.0
    except Exception:
        return 0.0


def _to_int(v: Any, default: int = 1) -> int:
    """Tolerant int parser. Falls back to `default` on failure (and clamps to >=1)."""
    f = _to_float(v)
    try:
        n = int(f)
    except Exception:
        n = default
    return max(1, n) if default >= 1 else n


def _norm_lower(s: Optional[str]) -> str:
    return (s or "").strip().lower()


@api_router.post("/tools/import")
async def tools_import(payload: ImportPayload, user: User = Depends(get_current_user)):
    """Bulk-create tools from a normalised list of rows. The frontend is
    responsible for column mapping; this endpoint just creates tools and
    resolves FK names → ids (with optional auto-create for categories/tags).
    """
    # Free-tier 15-item limit applied once for the whole batch.
    from subscriptions import enforce_tool_limit  # local import to avoid cycles
    rows_count = len(payload.rows or [])
    if rows_count > 0:
        await enforce_tool_limit(real_db, user.id, additional=rows_count)
    # Pre-load all the lookup collections once.
    cats = await db.categories.find({}, {"_id": 0}).to_list(5000)
    cats_by_name = {(_norm_lower(c.get("name"))): c for c in cats if c.get("name")}
    tags = await db.tags.find({}, {"_id": 0}).to_list(5000)
    tags_by_name = {(_norm_lower(t.get("name"))): t for t in tags if t.get("name")}
    locs = await db.locations.find({}, {"_id": 0}).to_list(5000)
    locs_by_name = {(_norm_lower(l.get("name"))): l for l in locs if l.get("name")}
    dlrs = await db.dealers.find({}, {"_id": 0}).to_list(5000)
    dlrs_by_name = {(_norm_lower(d.get("name"))): d for d in dlrs if d.get("name")}

    created: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    auto_created = {
        "categories": [],  # list of {id, name}
        "tags": [],
        "locations": [],
        "dealers": [],
    }

    for idx, raw in enumerate(payload.rows):
        try:
            name = _norm(raw.name)
            if not name:
                raise ValueError("Name is required")

            # Category — auto-create if missing
            category_id = None
            category_name = ""
            cname = _norm(raw.category)
            if cname:
                key = cname.lower()
                if key in cats_by_name:
                    c = cats_by_name[key]
                    category_id = c.get("id")
                    category_name = c.get("name") or cname
                elif payload.create_missing_categories:
                    new_cat = Category(name=cname)
                    await db.categories.insert_one(new_cat.dict())
                    cats_by_name[key] = new_cat.dict()
                    category_id = new_cat.id
                    category_name = new_cat.name
                    auto_created["categories"].append({"id": new_cat.id, "name": new_cat.name})

            # Location — match existing (case-insensitive) or auto-create
            location_id = None
            location_name = ""
            lname = _norm(raw.location)
            if lname:
                key = lname.lower()
                if key in locs_by_name:
                    l = locs_by_name[key]
                    location_id = l.get("id")
                    location_name = l.get("name") or lname
                elif payload.create_missing_locations:
                    new_loc = Location(name=lname)
                    await db.locations.insert_one(new_loc.dict())
                    locs_by_name[key] = new_loc.dict()
                    location_id = new_loc.id
                    location_name = new_loc.name
                    auto_created["locations"].append({"id": new_loc.id, "name": new_loc.name})

            # Dealer — match existing (case-insensitive) or auto-create
            dealer_id = None
            dealer_name = ""
            dname = _norm(raw.dealer)
            if dname:
                key = dname.lower()
                if key in dlrs_by_name:
                    d = dlrs_by_name[key]
                    dealer_id = d.get("id")
                    dealer_name = d.get("name") or dname
                elif payload.create_missing_dealers:
                    new_dlr = Dealer(name=dname)
                    await db.dealers.insert_one(new_dlr.dict())
                    dlrs_by_name[key] = new_dlr.dict()
                    dealer_id = new_dlr.id
                    dealer_name = new_dlr.name
                    auto_created["dealers"].append({"id": new_dlr.id, "name": new_dlr.name})

            # Tags — comma-separated; auto-create if missing
            tag_ids: List[str] = []
            tag_names: List[str] = []
            if raw.tags:
                for piece in (raw.tags or "").split(","):
                    tname = _norm(piece)
                    if not tname:
                        continue
                    key = tname.lower()
                    if key in tags_by_name:
                        tg = tags_by_name[key]
                        tag_ids.append(tg.get("id"))
                        tag_names.append(tg.get("name") or tname)
                    elif payload.create_missing_tags:
                        new_tag = Tag(name=tname)
                        await db.tags.insert_one(new_tag.dict())
                        tags_by_name[key] = new_tag.dict()
                        tag_ids.append(new_tag.id)
                        tag_names.append(new_tag.name)
                        auto_created["tags"].append({"id": new_tag.id, "name": new_tag.name})

            qty = _to_int(raw.quantity, default=1)
            cost = _to_float(raw.cost)
            msrp_price = _to_float(raw.msrp_price)

            # Build model_numbers[] from legacy import fields (deduped).
            _mn_cands = []
            if _norm(raw.serial_number):
                _mn_cands.append(_norm(raw.serial_number))
            if _norm(raw.model):
                _mn_cands.append(_norm(raw.model))
            _mn_seen = set()
            _model_numbers: List[str] = []
            for _v in _mn_cands:
                if _v and _v not in _mn_seen:
                    _mn_seen.add(_v)
                    _model_numbers.append(_v)

            tool = Tool(
                name=name,
                brand=_norm(raw.brand),
                model=_norm(raw.model),
                serial_number=_norm(raw.serial_number),
                model_numbers=_model_numbers,
                serial_numbers=[],
                quantity=qty,
                cost=cost,
                msrp_price=msrp_price,
                description=_norm(raw.description),
                category_id=category_id,
                category_name=category_name,
                location_id=location_id,
                location_name=location_name,
                dealer_id=dealer_id,
                dealer_name=dealer_name,
                tag_ids=tag_ids,
                tag_names=tag_names,
                condition=_norm(raw.condition),
                purchase_date=_norm(raw.purchase_date) or None,
                warranty_expiry=_norm(raw.warranty_expiry) or None,
            )
            await db.tools.insert_one(tool.dict())
            created.append({"id": tool.id, "name": tool.name})
        except Exception as e:
            errors.append({"row": idx + 1, "name": _norm(raw.name), "error": str(e)})

    return {
        "created": len(created),
        "errors": errors,
        "ids": [c["id"] for c in created],
        "auto_created": auto_created,
    }


# ---------------------------------------------------------------------------
# Photo size cap — prevent oversized base64 payloads from bloating the DB.
# Each photo gets its own ~5MB cap (well above what camera+compress yields)
# and the total per-tool photo payload is capped at 25MB so a single tool
# can have multiple photos but never balloon the document.
# ---------------------------------------------------------------------------
MAX_PHOTO_BYTES = 5 * 1024 * 1024            # ~5MB per photo (raw decoded ≈ 3.7MB)
MAX_TOTAL_PHOTO_BYTES = 25 * 1024 * 1024     # ~25MB total per tool


def _validate_photo_payload(photos: Optional[List[str]]) -> None:
    if not photos:
        return
    total = 0
    for i, p in enumerate(photos):
        if not p:
            continue
        size = len(p)  # length in bytes of the base64 string itself
        if size > MAX_PHOTO_BYTES:
            raise HTTPException(
                413,
                f"Photo #{i + 1} is too large ({size // 1024} KB). "
                f"Maximum allowed is {MAX_PHOTO_BYTES // (1024 * 1024)} MB per photo. "
                "Please re-take or resize the photo before saving.",
            )
        total += size
    if total > MAX_TOTAL_PHOTO_BYTES:
        raise HTTPException(
            413,
            f"Total photo payload is too large ({total // (1024 * 1024)} MB). "
            f"Maximum allowed is {MAX_TOTAL_PHOTO_BYTES // (1024 * 1024)} MB across all photos for one tool.",
        )


@api_router.get("/tools", response_model=List[Tool])
async def list_tools(
    search: Optional[str] = None,
    location_id: Optional[str] = None,
    tag_id: Optional[str] = None,
    category_id: Optional[str] = None,
    dealer_id: Optional[str] = None,
    checked_out: Optional[bool] = None,
    is_consumable: Optional[bool] = None,
    needs_repair: Optional[bool] = None,
    for_sale: Optional[bool] = None,
    is_sold: Optional[bool] = None,
):
    """List tools — returns a slim payload for fast list rendering.

    To keep the inventory list snappy on phones, this endpoint:
      - Returns ONLY the first photo (the cover), not all photos
      - Strips `documents` (heavy base64 PDFs/images attached to the tool)
      - Strips `receipts` (heavy base64 receipt images)
    The full set of photos / documents / receipts is only loaded when the
    user opens a tool's detail page (GET /tools/{id}).
    """
    query = build_tool_query(
        search, location_id, tag_id, category_id, dealer_id,
        checked_out, is_consumable, needs_repair, for_sale, is_sold,
    )
    items = await db.tools.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    out: List[Tool] = []
    for i in items:
        # Slim payload: keep only first photo, drop documents & receipts
        photos = i.get("photos") or []
        i["photos"] = photos[:1] if photos else []
        i["documents"] = []
        i["receipts"] = []
        out.append(Tool(**i))
    return out


@api_router.get("/tools/{tool_id}", response_model=Tool)
async def get_tool(tool_id: str):
    doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Tool not found")
    return Tool(**doc)


@api_router.put("/tools/{tool_id}", response_model=Tool)
async def update_tool(tool_id: str, payload: ToolUpdate):
    if payload.photos is not None:
        _validate_photo_payload(payload.photos)
    doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Tool not found")

    # Use exclude_unset so the client can EXPLICITLY clear a field by
    # sending null (e.g., `{"location_id": null}` to detach the tool from
    # its location). The previous "drop if v is None" rule made the
    # endpoint physically unable to clear a foreign-key field — values
    # could only be reassigned, never removed.
    updates = payload.dict(exclude_unset=True)
    updates["updated_at"] = now_iso()

    # Keep legacy model/serial fields and new model_numbers/serial_numbers
    # arrays in sync — whichever shape the client sent, both get persisted.
    _resolve_model_serial_arrays(updates, doc)

    # ---------------------------------------------------------------
    # Keep denormalized *_name fields in sync with their *_id fields.
    # The frontend only sends *_id when re-assigning location/dealer/
    # category, but the tool model also stores a cached *_name field
    # (used by list views to avoid an extra join). If we update the id
    # without refreshing the name, the description card keeps showing
    # the OLD name forever — exactly the "renamed to test2 but card
    # still shows test" bug reported 2026-05-23.
    # ---------------------------------------------------------------
    if "location_id" in updates:
        if updates["location_id"]:
            loc = await db.locations.find_one(
                {"id": updates["location_id"]}, {"_id": 0, "name": 1}
            )
            updates["location_name"] = (loc or {}).get("name", "") or ""
        else:
            updates["location_name"] = ""

    if "dealer_id" in updates:
        if updates["dealer_id"]:
            dl = await db.dealers.find_one(
                {"id": updates["dealer_id"]}, {"_id": 0, "name": 1}
            )
            updates["dealer_name"] = (dl or {}).get("name", "") or ""
        else:
            updates["dealer_name"] = ""

    if "category_id" in updates:
        if updates["category_id"]:
            cat = await db.categories.find_one(
                {"id": updates["category_id"]}, {"_id": 0, "name": 1}
            )
            updates["category_name"] = (cat or {}).get("name", "") or ""
        else:
            updates["category_name"] = ""

    # When the caller marks the tool as Repaired (needs_repair: false, repair_info: null),
    # Pydantic's None values get filtered out above. Restore the null so the tool's
    # repair_info (including broken_photo) is actually cleared — otherwise the next claim
    # would inherit the previous claim's photo and notes.
    if payload.needs_repair is False and doc.get("needs_repair"):
        updates["repair_info"] = None

    # Auto-checkin when a tool is being newly flagged as broken / needing repair
    becomes_broken = (
        updates.get("needs_repair") is True
        and not doc.get("needs_repair")
        and doc.get("is_checked_out")
    )
    if becomes_broken:
        record = doc.get("current_checkout") or {}
        if record:
            record = dict(record)
            record["checked_in_at"] = now_iso()
            note_extra = " [auto check-in: marked for repair]"
            record["notes"] = (record.get("notes") or "") + note_extra
            history = doc.get("checkout_history") or []
            history.append(record)
            updates["is_checked_out"] = False
            updates["current_checkout"] = None
            updates["checkout_history"] = history

    await db.tools.update_one({"id": tool_id}, {"$set": updates})
    new_doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})

    # Sync warranty claim record:
    # Newly broken → create a new open claim (if no open claim already exists for this tool)
    just_flagged = updates.get("needs_repair") is True and not doc.get("needs_repair")
    just_unflagged = updates.get("needs_repair") is False and doc.get("needs_repair")
    if just_flagged:
        existing_open = await db.warranty_claims.find_one(
            {"tool_id": tool_id, "claim_status": {"$nin": ["completed", "rejected"]}},
            {"_id": 0, "id": 1},
        )
        if not existing_open:
            ri = (new_doc.get("repair_info") or {})
            claim = WarrantyClaim(
                tool_id=tool_id,
                tool_name=new_doc.get("name", ""),
                tool_photo=(new_doc.get("photos") or [None])[0],
                broken_photo=ri.get("broken_photo") or "",
                dealer_id=new_doc.get("dealer_id"),
                dealer_name=new_doc.get("dealer_name") or "",
                repair_company=ri.get("company_notified") or "",
                contact=ri.get("contact") or "",
                notified_at=ri.get("notified_at") or "",
                expected_completion=ri.get("expected_completion") or "",
                claim_status="broken",
                notes=ri.get("notes") or "",
                repair_cost=float(ri.get("repair_cost") or 0),
            )
            await db.warranty_claims.insert_one(claim.dict())
    elif "repair_info" in updates and new_doc.get("needs_repair"):
        # Repair info edited while still broken → keep open claim in sync
        ri = updates.get("repair_info") or {}
        await db.warranty_claims.update_many(
            {"tool_id": tool_id, "claim_status": {"$nin": ["completed", "rejected"]}},
            {
                "$set": {
                    "repair_company": ri.get("company_notified") or "",
                    "contact": ri.get("contact") or "",
                    "notified_at": ri.get("notified_at") or "",
                    "expected_completion": ri.get("expected_completion") or "",
                    "notes": ri.get("notes") or "",
                    "broken_photo": ri.get("broken_photo") or "",
                    "repair_cost": float(ri.get("repair_cost") or 0),
                    "updated_at": now_iso(),
                }
            },
        )
    if just_unflagged:
        # User hit "Mark Repaired" — close any still-open claim as completed
        await db.warranty_claims.update_many(
            {"tool_id": tool_id, "claim_status": {"$nin": ["completed", "rejected"]}},
            {
                "$set": {
                    "claim_status": "completed",
                    "completed_at": now_iso(),
                    "updated_at": now_iso(),
                }
            },
        )

    return Tool(**new_doc)


@api_router.delete("/tools/{tool_id}")
async def delete_tool(tool_id: str):
    res = await db.tools.delete_one({"id": tool_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Tool not found")
    # Cascade: also remove any warranty claims that referenced this tool —
    # otherwise the dealer-claims summary keeps counting orphaned claims
    # but the detail screen can't resolve them back to a tool.
    await db.warranty_claims.delete_many({"tool_id": tool_id})
    return {"ok": True}


# ---------- Sale / Sold ----------
class MarkSoldRequest(BaseModel):
    sold_price: Optional[float] = 0.0
    sold_to: Optional[str] = ""
    sold_at: Optional[str] = ""  # YYYY-MM-DD; defaults to today if empty
    sold_notes: Optional[str] = ""
    sold_quantity: Optional[int] = None  # None or >= current qty → mark fully sold;
                                         # less than current qty → just decrement.


@api_router.post("/tools/{tool_id}/mark-sold", response_model=Tool)
async def mark_tool_sold(tool_id: str, payload: MarkSoldRequest):
    """Mark a tool as sold. If `sold_quantity` is supplied AND less than the
    current `quantity`, the tool's quantity is simply decremented and the
    tool stays in active inventory (partial sale). Otherwise the tool is
    fully marked sold (existing behavior)."""
    doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Tool not found")

    current_qty = max(1, int(doc.get("quantity") or 1))
    sold_qty_raw = payload.sold_quantity
    sold_qty = current_qty if sold_qty_raw in (None, 0) else max(1, int(sold_qty_raw))
    if sold_qty > current_qty:
        sold_qty = current_qty

    sold_at = (payload.sold_at or "").strip()
    if not sold_at:
        from datetime import datetime as _dt
        sold_at = _dt.utcnow().strftime("%Y-%m-%d")

    # Partial sale: decrement quantity only — don't mark sold.
    if sold_qty < current_qty:
        await db.tools.update_one(
            {"id": tool_id},
            {"$set": {
                "quantity": current_qty - sold_qty,
                "updated_at": now_iso(),
            }},
        )
        new_doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
        return Tool(**new_doc)

    # Full sale: mark the tool as sold.
    updates = {
        "is_sold": True,
        "sold_at": sold_at,
        "sold_price": float(payload.sold_price or 0.0),
        "sold_to": (payload.sold_to or "").strip(),
        "sold_notes": (payload.sold_notes or "").strip(),
        "for_sale": False,
        "updated_at": now_iso(),
    }
    # Auto check-in if currently checked out
    if doc.get("is_checked_out"):
        record = doc.get("current_checkout") or {}
        if record:
            record = dict(record)
            record["checked_in_at"] = now_iso()
            record["notes"] = (record.get("notes") or "") + " [auto check-in: marked sold]"
            history = doc.get("checkout_history") or []
            history.append(record)
            updates["is_checked_out"] = False
            updates["current_checkout"] = None
            updates["checkout_history"] = history

    await db.tools.update_one({"id": tool_id}, {"$set": updates})
    new_doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    return Tool(**new_doc)


@api_router.post("/tools/{tool_id}/unmark-sold", response_model=Tool)
async def unmark_tool_sold(tool_id: str):
    """Restore a sold tool back into regular inventory (clears sold fields)."""
    doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Tool not found")
    await db.tools.update_one(
        {"id": tool_id},
        {
            "$set": {
                "is_sold": False,
                "sold_at": "",
                "sold_price": 0.0,
                "sold_to": "",
                "sold_notes": "",
                "updated_at": now_iso(),
            }
        },
    )
    new_doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    return Tool(**new_doc)


@api_router.post("/tools/{tool_id}/checkout", response_model=Tool)
async def checkout_tool(tool_id: str, payload: CheckoutRequest):
    doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Tool not found")
    if doc.get("is_checked_out"):
        raise HTTPException(status_code=400, detail="Tool already checked out")
    record = CheckoutRecord(
        borrower_name=payload.borrower_name,
        borrower_id=payload.borrower_id,
        notes=payload.notes or "",
    )
    await db.tools.update_one(
        {"id": tool_id},
        {"$set": {"is_checked_out": True, "current_checkout": record.dict(), "updated_at": now_iso()}},
    )
    new_doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    return Tool(**new_doc)


@api_router.post("/tools/{tool_id}/checkin", response_model=Tool)
async def checkin_tool(tool_id: str):
    doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Tool not found")
    if not doc.get("is_checked_out"):
        raise HTTPException(status_code=400, detail="Tool is not checked out")
    record = doc.get("current_checkout") or {}
    record["checked_in_at"] = now_iso()
    history = doc.get("checkout_history") or []
    history.append(record)
    await db.tools.update_one(
        {"id": tool_id},
        {"$set": {"is_checked_out": False, "current_checkout": None, "checkout_history": history, "updated_at": now_iso()}},
    )
    new_doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    return Tool(**new_doc)


# ---------- Documents (per tool) ----------
@api_router.post("/tools/{tool_id}/documents", response_model=Tool)
async def add_tool_document(tool_id: str, payload: Document):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(404, "Tool not found")
    docs = tool.get("documents") or []
    new_doc = payload.model_dump()
    if not new_doc.get("id"):
        new_doc["id"] = str(uuid.uuid4())
    if not new_doc.get("uploaded_at"):
        new_doc["uploaded_at"] = now_iso()
    if not new_doc.get("size") and new_doc.get("data"):
        # Estimate size in bytes from base64 length
        new_doc["size"] = int(len(new_doc["data"]) * 3 / 4)
    docs.append(new_doc)
    await db.tools.update_one({"id": tool_id}, {"$set": {"documents": docs, "updated_at": now_iso()}})
    return Tool(**(await db.tools.find_one({"id": tool_id}, {"_id": 0})))


@api_router.delete("/tools/{tool_id}/documents/{doc_id}", response_model=Tool)
async def delete_tool_document(tool_id: str, doc_id: str):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(404, "Tool not found")
    docs = [d for d in (tool.get("documents") or []) if d.get("id") != doc_id]
    await db.tools.update_one({"id": tool_id}, {"$set": {"documents": docs, "updated_at": now_iso()}})
    return Tool(**(await db.tools.find_one({"id": tool_id}, {"_id": 0})))


# ---------- Maintenance Schedules ----------
def _calc_next_due(last: Optional[str], months: int) -> Optional[str]:
    if not last or not months:
        return None
    try:
        d = datetime.strptime(last, "%Y-%m-%d")
        # Approximate by adding months * 30.4 days; simpler than dateutil
        new_month = d.month + months
        new_year = d.year + (new_month - 1) // 12
        new_month = ((new_month - 1) % 12) + 1
        # Clamp day
        try:
            nd = d.replace(year=new_year, month=new_month)
        except ValueError:
            # Day overflow (e.g., Feb 30) — back off
            nd = d.replace(year=new_year, month=new_month, day=28)
        return nd.strftime("%Y-%m-%d")
    except Exception:
        return None


@api_router.post("/tools/{tool_id}/maintenance", response_model=Tool)
async def add_maintenance(tool_id: str, payload: MaintenanceScheduleCreate):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(404, "Tool not found")
    schedules = tool.get("maintenance") or []
    new_sch = MaintenanceSchedule(
        type=payload.type or "Service",
        interval_months=payload.interval_months or 12,
        last_done_date=payload.last_done_date,
        next_due_date=_calc_next_due(payload.last_done_date, payload.interval_months or 12),
        notes=payload.notes or "",
    ).model_dump()
    schedules.append(new_sch)
    await db.tools.update_one({"id": tool_id}, {"$set": {"maintenance": schedules, "updated_at": now_iso()}})
    return Tool(**(await db.tools.find_one({"id": tool_id}, {"_id": 0})))


@api_router.put("/tools/{tool_id}/maintenance/{sch_id}", response_model=Tool)
async def update_maintenance(tool_id: str, sch_id: str, payload: MaintenanceScheduleUpdate):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(404, "Tool not found")
    schedules = tool.get("maintenance") or []
    found = False
    for sch in schedules:
        if sch.get("id") == sch_id:
            found = True
            if payload.type is not None:
                sch["type"] = payload.type
            if payload.interval_months is not None:
                sch["interval_months"] = payload.interval_months
            if payload.last_done_date is not None:
                sch["last_done_date"] = payload.last_done_date
            if payload.notes is not None:
                sch["notes"] = payload.notes
            sch["next_due_date"] = _calc_next_due(sch.get("last_done_date"), sch.get("interval_months", 12))
            break
    if not found:
        raise HTTPException(404, "Schedule not found")
    await db.tools.update_one({"id": tool_id}, {"$set": {"maintenance": schedules, "updated_at": now_iso()}})
    return Tool(**(await db.tools.find_one({"id": tool_id}, {"_id": 0})))


@api_router.delete("/tools/{tool_id}/maintenance/{sch_id}", response_model=Tool)
async def delete_maintenance(tool_id: str, sch_id: str):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(404, "Tool not found")
    schedules = [s for s in (tool.get("maintenance") or []) if s.get("id") != sch_id]
    await db.tools.update_one({"id": tool_id}, {"$set": {"maintenance": schedules, "updated_at": now_iso()}})
    return Tool(**(await db.tools.find_one({"id": tool_id}, {"_id": 0})))


@api_router.post("/tools/{tool_id}/maintenance/{sch_id}/service", response_model=Tool)
async def log_service_event(tool_id: str, sch_id: str, payload: ServiceEventCreate):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(404, "Tool not found")
    schedules = tool.get("maintenance") or []
    found = False
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for sch in schedules:
        if sch.get("id") == sch_id:
            found = True
            event = ServiceEvent(
                date=payload.date or today,
                cost=payload.cost or 0.0,
                technician=payload.technician or "",
                notes=payload.notes or "",
            ).model_dump()
            sch.setdefault("history", []).append(event)
            sch["last_done_date"] = event["date"]
            sch["next_due_date"] = _calc_next_due(sch["last_done_date"], sch.get("interval_months", 12))
            break
    if not found:
        raise HTTPException(404, "Schedule not found")
    await db.tools.update_one({"id": tool_id}, {"$set": {"maintenance": schedules, "updated_at": now_iso()}})
    return Tool(**(await db.tools.find_one({"id": tool_id}, {"_id": 0})))


@api_router.get("/maintenance/upcoming")
async def upcoming_maintenance(days: int = 30):
    """Return all maintenance schedules with next_due in [today, today+days] OR overdue."""
    now = datetime.now(timezone.utc)
    horizon = (now + timedelta(days=days)).strftime("%Y-%m-%d")
    today = now.strftime("%Y-%m-%d")
    out: List[Dict[str, Any]] = []
    overdue_count = 0
    due_soon_count = 0
    async for tool in db.tools.find(
        {"maintenance": {"$exists": True, "$ne": []}},
        {"_id": 0, "id": 1, "name": 1, "photos": 1, "maintenance": 1},
    ):
        for sch in (tool.get("maintenance") or []):
            nd = sch.get("next_due_date") or ""
            if not nd:
                continue
            if nd <= horizon:  # overdue or within horizon
                is_overdue = nd < today
                if is_overdue:
                    overdue_count += 1
                else:
                    due_soon_count += 1
                out.append({
                    "tool_id": tool.get("id"),
                    "tool_name": tool.get("name"),
                    "tool_photo": (tool.get("photos") or [None])[0] if tool.get("photos") else None,
                    "schedule_id": sch.get("id"),
                    "type": sch.get("type"),
                    "interval_months": sch.get("interval_months"),
                    "last_done_date": sch.get("last_done_date"),
                    "next_due_date": nd,
                    "is_overdue": is_overdue,
                    "notes": sch.get("notes", ""),
                })
    out.sort(key=lambda x: x["next_due_date"] or "9999-12-31")
    return {
        "items": out,
        "total": len(out),
        "overdue": overdue_count,
        "due_soon": due_soon_count,
    }


# ---------- Theft / Loss Reporting ----------
@api_router.post("/tools/{tool_id}/report-lost", response_model=Tool)
async def report_lost(tool_id: str, payload: ReportLostRequest):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(404, "Tool not found")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    lost_status = LostStatus(
        is_lost=True,
        type=payload.type if payload.type in ("lost", "stolen") else "lost",
        reported_date=payload.reported_date or today,
        police_report_number=payload.police_report_number or "",
        insurance_company=payload.insurance_company or "",
        insurance_claim_number=payload.insurance_claim_number or "",
        notes=payload.notes or "",
        reported_by=payload.reported_by or "",
        recovered_at=None,
    ).model_dump()
    await db.tools.update_one({"id": tool_id}, {"$set": {"lost_status": lost_status, "updated_at": now_iso()}})
    return Tool(**(await db.tools.find_one({"id": tool_id}, {"_id": 0})))


@api_router.post("/tools/{tool_id}/recover", response_model=Tool)
async def mark_recovered(tool_id: str):
    tool = await db.tools.find_one({"id": tool_id}, {"_id": 0})
    if not tool:
        raise HTTPException(404, "Tool not found")
    lost = tool.get("lost_status") or {}
    lost["is_lost"] = False
    lost["recovered_at"] = now_iso()
    await db.tools.update_one({"id": tool_id}, {"$set": {"lost_status": lost, "updated_at": now_iso()}})
    return Tool(**(await db.tools.find_one({"id": tool_id}, {"_id": 0})))


# ---------- Bulk Operations ----------
class BulkRequest(BaseModel):
    tool_ids: List[str]
    action: str  # "delete" | "move_location" | "add_tag" | "remove_tag" | "set_category" | "report_lost"
    # Optional payload depending on action
    location_id: Optional[str] = None
    location_name: Optional[str] = None
    tag_id: Optional[str] = None
    tag_name: Optional[str] = None
    category_id: Optional[str] = None
    category_name: Optional[str] = None
    lost_payload: Optional[ReportLostRequest] = None


@api_router.post("/tools/bulk")
async def bulk_tools(payload: BulkRequest):
    if not payload.tool_ids:
        return {"ok": True, "affected": 0}
    affected = 0
    if payload.action == "delete":
        result = await db.tools.delete_many({"id": {"$in": payload.tool_ids}})
        affected = result.deleted_count
        # Cascade: drop any warranty claims that referenced the deleted tools.
        await db.warranty_claims.delete_many({"tool_id": {"$in": payload.tool_ids}})
    elif payload.action == "move_location":
        result = await db.tools.update_many(
            {"id": {"$in": payload.tool_ids}},
            {"$set": {
                "location_id": payload.location_id,
                "location_name": payload.location_name or "",
                "updated_at": now_iso(),
            }},
        )
        affected = result.modified_count
    elif payload.action == "set_category":
        result = await db.tools.update_many(
            {"id": {"$in": payload.tool_ids}},
            {"$set": {
                "category_id": payload.category_id,
                "category_name": payload.category_name or "",
                "updated_at": now_iso(),
            }},
        )
        affected = result.modified_count
    elif payload.action == "add_tag":
        if not payload.tag_id:
            raise HTTPException(400, "tag_id required")
        # Bulk add tag using $addToSet — single round-trip
        add_to_set: Dict[str, Any] = {"tag_ids": payload.tag_id}
        if payload.tag_name:
            add_to_set["tag_names"] = payload.tag_name
        result = await db.tools.update_many(
            {"id": {"$in": payload.tool_ids}, "tag_ids": {"$ne": payload.tag_id}},
            {"$addToSet": add_to_set, "$set": {"updated_at": now_iso()}},
        )
        affected = result.modified_count
    elif payload.action == "remove_tag":
        if not payload.tag_id:
            raise HTTPException(400, "tag_id required")
        # Bulk remove tag using $pull — single round-trip
        pull_doc: Dict[str, Any] = {"tag_ids": payload.tag_id}
        if payload.tag_name:
            pull_doc["tag_names"] = payload.tag_name
        result = await db.tools.update_many(
            {"id": {"$in": payload.tool_ids}, "tag_ids": payload.tag_id},
            {"$pull": pull_doc, "$set": {"updated_at": now_iso()}},
        )
        affected = result.modified_count
    elif payload.action == "report_lost":
        lp = payload.lost_payload or ReportLostRequest()
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        lost_status = LostStatus(
            is_lost=True,
            type=lp.type if lp.type in ("lost", "stolen") else "lost",
            reported_date=lp.reported_date or today,
            police_report_number=lp.police_report_number or "",
            insurance_company=lp.insurance_company or "",
            insurance_claim_number=lp.insurance_claim_number or "",
            notes=lp.notes or "",
            reported_by=lp.reported_by or "",
        ).model_dump()
        result = await db.tools.update_many(
            {"id": {"$in": payload.tool_ids}},
            {"$set": {"lost_status": lost_status, "updated_at": now_iso()}},
        )
        affected = result.modified_count
    else:
        raise HTTPException(400, f"Unknown action '{payload.action}'")
    return {"ok": True, "affected": affected}


# ---------- Aggregate / Stats ----------
@api_router.get("/aggregate")
async def aggregate(
    search: Optional[str] = None,
    location_id: Optional[str] = None,
    tag_id: Optional[str] = None,
    category_id: Optional[str] = None,
    dealer_id: Optional[str] = None,
    checked_out: Optional[bool] = None,
    is_consumable: Optional[bool] = None,
    needs_repair: Optional[bool] = None,
):
    query = build_tool_query(search, location_id, tag_id, category_id, dealer_id, checked_out, is_consumable, needs_repair)
    items = await db.tools.find(query, {"_id": 0}).to_list(5000)

    def _num(v: Any, default: float = 0.0) -> float:
        """Robust numeric coercion — handles ints, floats, strings like
        "1,234.56", "$300", "  5  ", as well as None / missing values."""
        if v is None:
            return default
        if isinstance(v, (int, float)):
            return float(v)
        try:
            s = str(v).strip().replace(",", "").replace("$", "")
            return float(s) if s else default
        except Exception:
            return default

    def _qty(v: Any) -> int:
        n = int(_num(v, 1) or 1)
        return n if n >= 1 else 1

    total_value = sum(_num(i.get("cost")) * _qty(i.get("quantity")) for i in items)
    checked_out_n = sum(1 for i in items if i.get("is_checked_out"))
    consumables_n = sum(1 for i in items if i.get("is_consumable"))
    needs_repair_n = sum(1 for i in items if i.get("needs_repair"))
    locations: Dict[str, int] = {}
    categories: Dict[str, int] = {}
    dealers: Dict[str, int] = {}
    tag_set: set = set()
    for i in items:
        ln = i.get("location_name") or "—"
        cn = i.get("category_name") or "—"
        dn = i.get("dealer_name") or "—"
        locations[ln] = locations.get(ln, 0) + 1
        categories[cn] = categories.get(cn, 0) + 1
        dealers[dn] = dealers.get(dn, 0) + 1
        for t in (i.get("tag_names") or []):
            tag_set.add(t)
    return {
        "count": len(items),
        "total_value": round(total_value, 2),
        "checked_out": checked_out_n,
        "available": len(items) - checked_out_n,
        "consumables": consumables_n,
        "needs_repair": needs_repair_n,
        "location_breakdown": locations,
        "category_breakdown": categories,
        "dealer_breakdown": dealers,
        "tag_count": len(tag_set),
        "unique_tags": sorted(tag_set),
    }


@api_router.get("/stats")
async def get_stats():
    total = await db.tools.count_documents({})
    checked_out = await db.tools.count_documents({"is_checked_out": True})
    consumables = await db.tools.count_documents({"is_consumable": True})
    needs_repair = await db.tools.count_documents({"needs_repair": True})
    locations = await db.locations.count_documents({})
    tags = await db.tags.count_documents({})
    categories = await db.categories.count_documents({})
    borrowers = await db.borrowers.count_documents({})
    dealers = await db.dealers.count_documents({})
    pipeline = [{"$group": {"_id": None, "total_value": {
        "$sum": {"$multiply": ["$cost", {"$ifNull": ["$quantity", 1]}]}
    }}}]
    agg = await db.tools.aggregate(pipeline).to_list(1)
    total_value = agg[0]["total_value"] if agg else 0
    # Warranty expiring within 60 days — exclude tools that are sold,
    # lost, or stolen since warranty alerts on those items are noise
    # (the user no longer has/cares about the item).
    soon = (datetime.now(timezone.utc) + timedelta(days=60)).date().isoformat()
    today = datetime.now(timezone.utc).date().isoformat()
    _warranty_active_filter = {
        "is_sold": {"$ne": True},
        "lost_status.is_lost": {"$ne": True},
    }
    expiring = await db.tools.count_documents({
        "warranty.has_warranty": True,
        "warranty.expiry_date": {"$gte": today, "$lte": soon},
        **_warranty_active_filter,
    })
    expired = await db.tools.count_documents({
        "warranty.has_warranty": True,
        "warranty.expiry_date": {"$lt": today, "$ne": ""},
        **_warranty_active_filter,
    })
    return {
        "total_tools": total,
        "checked_out": checked_out,
        "available": total - checked_out,
        "consumables": consumables,
        "needs_repair": needs_repair,
        "total_value": round(total_value, 2),
        "locations": locations,
        "tags": tags,
        "categories": categories,
        "borrowers": borrowers,
        "dealers": dealers,
        "warranty_expiring_soon": expiring,
        "warranty_expired": expired,
    }


@api_router.get("/warranty-alerts")
async def warranty_alerts(days: int = 60):
    today = datetime.now(timezone.utc).date()
    soon = (today + timedelta(days=days)).isoformat()
    today_iso = today.isoformat()
    # Pull only ACTIVE tools — exclude sold, lost, or stolen items so
    # they don't clutter the warranty alert list. Users don't want
    # warranty reminders on items they no longer own or have written
    # off as lost/stolen.
    items = await db.tools.find(
        {
            "warranty.has_warranty": True,
            "warranty.expiry_date": {"$ne": ""},
            "is_sold": {"$ne": True},
            "lost_status.is_lost": {"$ne": True},
        },
        {"_id": 0, "id": 1, "name": 1, "warranty": 1, "photos": 1},
    ).to_list(5000)
    expiring = []
    expired = []
    for i in items:
        ex = (i.get("warranty") or {}).get("expiry_date") or ""
        if not ex:
            continue
        if ex < today_iso:
            expired.append(i)
        elif ex <= soon:
            expiring.append(i)
    return {"expiring": expiring, "expired": expired}


# ---------- Warranty Claims ----------
async def _purge_orphan_claims() -> int:
    """Delete any warranty claim whose tool no longer exists. Heals stale
    state from before the cascade-on-tool-delete fix. Cheap (one find +
    one delete_many) and idempotent — safe to call before every list
    /summary read."""
    claim_rows = await db.warranty_claims.find(
        {}, {"_id": 0, "tool_id": 1}
    ).to_list(20000)
    tool_ids_with_claims = list({(r.get("tool_id") or "") for r in claim_rows if r.get("tool_id")})
    if not tool_ids_with_claims:
        return 0
    existing_rows = await db.tools.find(
        {"id": {"$in": tool_ids_with_claims}}, {"_id": 0, "id": 1}
    ).to_list(20000)
    existing_tool_ids = {r.get("id") for r in existing_rows}
    orphans = [tid for tid in tool_ids_with_claims if tid not in existing_tool_ids]
    if not orphans:
        return 0
    res = await db.warranty_claims.delete_many({"tool_id": {"$in": orphans}})
    return res.deleted_count


@api_router.get("/warranty-claims", response_model=List[WarrantyClaim])
async def list_warranty_claims(
    dealer_id: Optional[str] = None,
    tool_id: Optional[str] = None,
    status: Optional[str] = None,
    archived: Optional[bool] = None,  # true -> completed/rejected only; false -> active only
):
    await _purge_orphan_claims()
    q: Dict[str, Any] = {}
    if tool_id:
        q["tool_id"] = tool_id
    if dealer_id:
        # Special token "_none_" matches claims without a dealer
        if dealer_id == "_none_":
            q["$or"] = [{"dealer_id": None}, {"dealer_id": ""}]
        else:
            q["dealer_id"] = dealer_id
    if status:
        q["claim_status"] = status
    elif archived is True:
        q["claim_status"] = {"$in": ["completed", "rejected"]}
    elif archived is False:
        q["claim_status"] = {"$nin": ["completed", "rejected"]}
    items = await db.warranty_claims.find(q, {"_id": 0}).sort("updated_at", -1).to_list(5000)
    return [WarrantyClaim(**i) for i in items]


@api_router.get("/warranty-claims/summary")
async def warranty_claims_summary():
    await _purge_orphan_claims()
    items = await db.warranty_claims.find({}, {"_id": 0}).to_list(10000)
    dealers = await db.dealers.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(2000)
    dealer_name_by_id = {d["id"]: d["name"] for d in dealers}

    by_dealer: Dict[str, Dict[str, Any]] = {}
    totals = {
        "total": len(items),
        "open": 0,
        "completed": 0,
        "rejected": 0,
        "broken": 0,
        "awaiting_approval": 0,
        "waiting_replacement": 0,
    }
    for i in items:
        st = i.get("claim_status") or "broken"
        if st in totals:
            totals[st] = totals.get(st, 0) + 1
        if st not in ("completed", "rejected"):
            totals["open"] += 1
        did = i.get("dealer_id") or "_none_"
        if did not in by_dealer:
            by_dealer[did] = {
                "dealer_id": did if did != "_none_" else None,
                "dealer_name": dealer_name_by_id.get(did, i.get("dealer_name") or ("No Dealer" if did == "_none_" else "Unknown Dealer")),
                "open": 0,
                "completed": 0,
                "rejected": 0,
                "total": 0,
                "broken": 0,
                "awaiting_approval": 0,
                "waiting_replacement": 0,
            }
        bucket = by_dealer[did]
        bucket["total"] += 1
        if st in bucket:
            bucket[st] = bucket.get(st, 0) + 1
        if st not in ("completed", "rejected"):
            bucket["open"] += 1
    dealer_list = sorted(by_dealer.values(), key=lambda d: (-d["open"], d["dealer_name"].lower()))
    return {"totals": totals, "dealers": dealer_list}


@api_router.get("/warranty-claims/{claim_id}", response_model=WarrantyClaim)
async def get_warranty_claim(claim_id: str):
    doc = await db.warranty_claims.find_one({"id": claim_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Claim not found")
    return WarrantyClaim(**doc)


@api_router.put("/warranty-claims/{claim_id}", response_model=WarrantyClaim)
async def update_warranty_claim(claim_id: str, payload: WarrantyClaimUpdate):
    doc = await db.warranty_claims.find_one({"id": claim_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Claim not found")
    updates = {k: v for k, v in payload.dict().items() if v is not None}
    if "claim_status" in updates and updates["claim_status"] not in CLAIM_STATUSES:
        raise HTTPException(400, f"Invalid claim_status. Must be one of {CLAIM_STATUSES}")
    updates["updated_at"] = now_iso()
    new_status = updates.get("claim_status")
    archiving = new_status in ("completed", "rejected") and doc.get("claim_status") not in ("completed", "rejected")
    reopening = new_status not in (None, "completed", "rejected") and doc.get("claim_status") in ("completed", "rejected")
    if archiving:
        updates["completed_at"] = now_iso()
    if reopening:
        updates["completed_at"] = None
    await db.warranty_claims.update_one({"id": claim_id}, {"$set": updates})

    # Mirror to the underlying tool
    tool_id = doc.get("tool_id")
    if tool_id:
        if archiving:
            # Tool is no longer broken once claim is closed
            await db.tools.update_one(
                {"id": tool_id},
                {"$set": {"needs_repair": False, "repair_info": None, "updated_at": now_iso()}},
            )
        elif reopening:
            # Reopen — flag tool broken and rebuild repair_info from the claim
            ri = {
                "company_notified": doc.get("repair_company") or updates.get("repair_company") or "",
                "contact": doc.get("contact") or updates.get("contact") or "",
                "notified_at": doc.get("notified_at") or updates.get("notified_at") or "",
                "expected_completion": doc.get("expected_completion") or updates.get("expected_completion") or "",
                "repair_status": "Reported",
                "notes": doc.get("notes") or updates.get("notes") or "",
            }
            await db.tools.update_one(
                {"id": tool_id},
                {"$set": {"needs_repair": True, "repair_info": ri, "updated_at": now_iso()}},
            )
        elif new_status and new_status not in ("completed", "rejected"):
            # Plain status change while still active — keep tool's repair_info repair_status in sync
            label_map = {
                "broken": "Reported",
                "awaiting_approval": "Reported",
                "waiting_replacement": "Awaiting Parts",
            }
            tdoc = await db.tools.find_one({"id": tool_id}, {"_id": 0, "repair_info": 1, "needs_repair": 1})
            if tdoc and tdoc.get("needs_repair"):
                ri = dict(tdoc.get("repair_info") or {})
                ri["repair_status"] = label_map.get(new_status, ri.get("repair_status") or "Reported")
                await db.tools.update_one({"id": tool_id}, {"$set": {"repair_info": ri, "updated_at": now_iso()}})

        # Repair-cost mirror: if the user edited repair_cost on the claim
        # AND the tool is still flagged broken, sync the value back onto
        # tool.repair_info so the edit screen reflects it.
        if "repair_cost" in updates and not archiving:
            tdoc2 = await db.tools.find_one({"id": tool_id}, {"_id": 0, "repair_info": 1, "needs_repair": 1})
            if tdoc2 and tdoc2.get("needs_repair"):
                ri2 = dict(tdoc2.get("repair_info") or {})
                ri2["repair_cost"] = float(updates.get("repair_cost") or 0)
                await db.tools.update_one({"id": tool_id}, {"$set": {"repair_info": ri2, "updated_at": now_iso()}})

    new = await db.warranty_claims.find_one({"id": claim_id}, {"_id": 0})
    return WarrantyClaim(**new)


@api_router.delete("/warranty-claims/{claim_id}")
async def delete_warranty_claim(claim_id: str):
    res = await db.warranty_claims.delete_one({"id": claim_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Warranty claim not found")
    return {"ok": True}


# ---------- Wishlist ----------
@api_router.get("/wishlist", response_model=List[WishlistItem])
async def list_wishlist(purchased: Optional[bool] = None):
    q: Dict[str, Any] = {}
    if purchased is not None:
        q["purchased"] = purchased
    items = await db.wishlist_items.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [WishlistItem(**i) for i in items]


@api_router.post("/wishlist", response_model=WishlistItem)
async def create_wishlist(payload: WishlistItemCreate):
    item = WishlistItem(**payload.dict())
    if item.dealer_id and not item.dealer_name:
        d = await db.dealers.find_one({"id": item.dealer_id}, {"_id": 0, "name": 1})
        if d:
            item.dealer_name = d.get("name") or ""
    await db.wishlist_items.insert_one(item.dict())
    return item


@api_router.put("/wishlist/{item_id}", response_model=WishlistItem)
async def update_wishlist(item_id: str, payload: WishlistItemUpdate):
    doc = await db.wishlist_items.find_one({"id": item_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Wishlist item not found")
    updates = {k: v for k, v in payload.dict().items() if v is not None}
    updates["updated_at"] = now_iso()
    if "dealer_id" in updates and updates["dealer_id"]:
        d = await db.dealers.find_one({"id": updates["dealer_id"]}, {"_id": 0, "name": 1})
        if d:
            updates["dealer_name"] = d.get("name") or ""
    if updates.get("purchased") is True and not doc.get("purchased"):
        updates["purchased_at"] = now_iso()
    elif updates.get("purchased") is False:
        updates["purchased_at"] = None
        updates["converted_tool_id"] = None
    await db.wishlist_items.update_one({"id": item_id}, {"$set": updates})
    new = await db.wishlist_items.find_one({"id": item_id}, {"_id": 0})
    return WishlistItem(**new)


@api_router.delete("/wishlist/{item_id}")
async def delete_wishlist(item_id: str):
    res = await db.wishlist_items.delete_one({"id": item_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Wishlist item not found")
    return {"ok": True}


@api_router.post("/wishlist/{item_id}/convert", response_model=Tool)
async def convert_wishlist_to_tool(item_id: str, user: User = Depends(get_current_user)):
    """Convert a wishlist item into a real tool — marks as purchased."""
    item = await db.wishlist_items.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(404, "Wishlist item not found")
    # Free-tier 15-item limit. Pro / lifetime users always pass.
    from subscriptions import enforce_tool_limit  # local import to avoid cycles
    await enforce_tool_limit(real_db, user.id)
    tool = Tool(
        name=item.get("name", ""),
        description=item.get("description", "") or "",
        cost=item.get("price") or 0,
        dealer_id=item.get("dealer_id"),
        dealer_name=item.get("dealer_name") or "",
        # Carry through optional details captured on the wish so the
        # user doesn't have to re-enter them when finishing the tool.
        model=item.get("model_number", "") or "",
        photos=list(item.get("photos") or []),
        # Drop notes onto the description so we don't silently lose them.
        # If the user had both a description and notes, append the notes
        # on a new line. If only notes existed, that becomes the description.
        # (We do this in code rather than at the DB level so the wishlist
        # row still keeps its own notes/description fields intact.)
    )
    # Merge notes → description if both present (keep description first).
    extra_notes = (item.get("notes") or "").strip()
    if extra_notes:
        if tool.description:
            tool.description = f"{tool.description}\n\n{extra_notes}"
        else:
            tool.description = extra_notes
    await db.tools.insert_one(tool.dict())
    await db.wishlist_items.update_one(
        {"id": item_id},
        {"$set": {"purchased": True, "purchased_at": now_iso(), "converted_tool_id": tool.id, "updated_at": now_iso()}},
    )
    return tool


# ---------- AI Toolbox Analysis (REMOVED) ----------


# ---------- Personal Profile (singleton) ----------
class PersonalProfile(BaseModel):
    name: Optional[str] = ""
    address: Optional[str] = ""
    address2: Optional[str] = ""
    city: Optional[str] = ""
    state: Optional[str] = ""
    zip_code: Optional[str] = ""
    country: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    policy_number: Optional[str] = ""
    insurance_company: Optional[str] = ""
    notes: Optional[str] = ""
    is_company: Optional[bool] = False
    updated_at: str = Field(default_factory=now_iso)


@api_router.get("/personal-profile", response_model=PersonalProfile)
async def get_personal_profile():
    # Per-user singleton — owner_id auto-applied by the db proxy.
    doc = await db.personal_profile.find_one({"id": "self"}, {"_id": 0})
    if not doc:
        return PersonalProfile()
    return PersonalProfile(**doc)


@api_router.put("/personal-profile", response_model=PersonalProfile)
async def update_personal_profile(payload: PersonalProfile):
    data = payload.dict()
    data["id"] = "self"
    data["updated_at"] = now_iso()
    await db.personal_profile.update_one(
        {"id": "self"},
        {"$set": data},
        upsert=True,
    )
    return PersonalProfile(**data)


# ============================================================================
# Auth
# ============================================================================
auth_router = APIRouter(prefix="/api/auth")


async def _claim_orphan_data(user_id: str):
    """When the very first user registers, claim all legacy data (docs without
    an owner_id) for that user. Subsequent users get a clean slate.
    """
    user_count = await real_db.users.count_documents({})
    if user_count != 1:
        return
    collections = [
        "tools",
        "dealers",
        "borrowers",
        "locations",
        "tags",
        "categories",
        "wishlist_items",
        "warranty_claims",
        "personal_profile",
    ]
    for c in collections:
        await real_db[c].update_many(
            {"owner_id": {"$in": [None, ""]}},
            {"$set": {"owner_id": user_id}},
        )
    # personal_profile docs that used _id: "self" — make sure they have id field too
    await real_db.personal_profile.update_many(
        {"_id": "self", "owner_id": user_id, "id": {"$in": [None, ""]}},
        {"$set": {"id": "self"}},
    )


# --------------------------------------------------------------------------
# Default-content seeding for newly registered users
# --------------------------------------------------------------------------
# When a new user signs up, we seed their account with a small set of
# helpful defaults so they're not staring at an empty app on first launch:
#   • 5 well-known tool dealers, pre-populated with the publicly available
#     warranty / customer-service / tech-support contact channels
#   • A handful of tags + categories common to tool inventories
#
# The seed is fully idempotent — if the user already has a same-named
# record we skip rather than duplicate. This lets us safely re-run the
# seed against existing accounts (e.g. for the original owner) without
# creating doubles.

DEFAULT_DEALERS_SEED: List[Dict[str, Any]] = [
    {
        "name": "Snap-on Tools",
        "phone": "1-877-762-7664",
        "website": "www.snapon.com",
        "address": "Snap-on Incorporated, 2801 80th Street, Kenosha, WI 53143",
        "warranty_contact": "1-877-762-7664 · ncccsupport@snapon.com",
        "customer_support_contact": "1-877-762-7664",
        "tech_support_contact": "1-800-225-5786",
        "notes": "",
    },
    {
        "name": "Matco Tools",
        "phone": "866-289-8665",
        "website": "www.matcotools.com",
        "address": "Matco Tools Corporation, 4403 Allen Rd, Stow, OH 44224",
        "warranty_contact": "866-289-8665 (contact local distributor first)",
        "customer_support_contact": "866-289-8665 (Mon-Fri 8:00 AM - 6:30 PM EST)",
        "tech_support_contact": "866-289-8665",
        "notes": "",
    },
    {
        "name": "Mac Tools",
        "phone": "1-800-622-8665",
        "website": "www.mactools.com",
        "address": "Mac Tools, 5195 Blazer Parkway, Dublin, OH 43017",
        "warranty_contact": "1-800-MAC-TOOLS (622-8665)",
        "customer_support_contact": "1-800-MAC-TOOLS (622-8665)",
        "tech_support_contact": "1-800-MAC-TOOLS (622-8665) — select technical / product support option",
        "notes": "",
    },
    {
        "name": "Cornwell Tools",
        "phone": "1-800-321-8356",
        "website": "www.cornwelltools.com",
        "address": "Cornwell Quality Tools, 667 Seville Road, Wadsworth, OH 44281",
        "warranty_contact": "custserv@cornwelltools.com · 1-800-321-8356",
        "customer_support_contact": "1-800-321-8356 · 330-336-3506",
        "tech_support_contact": "1-800-321-8356",
        "notes": "",
    },
    {
        "name": "Harbor Freight",
        "phone": "1-800-444-3353",
        "website": "www.harborfreight.com",
        "address": "Harbor Freight Tools, 26677 Agoura Road, Calabasas, CA 91302",
        "warranty_contact": "1-888-838-3421 (many items have a lifetime in-store replacement warranty)",
        "customer_support_contact": "1-800-444-3353",
        "tech_support_contact": "1-888-838-3421 (gas-powered equipment / specialty)",
        "notes": "",
    },
]

DEFAULT_TAGS_SEED: List[str] = ["Hand tools", "Power tools", "Pneumatic tools"]

DEFAULT_CATEGORIES_SEED: List[str] = [
    "Timing",
    "Specialty Service",
    "Hand Tools",
    "Power Tools",
    "Pneumatic Tools",
]


async def seed_default_content_for_user(user_id: str) -> Dict[str, int]:
    """Insert the default dealers, tags and categories for *user_id*.

    Idempotent — same-named records (case-insensitive for tags / categories,
    exact match for dealer names) are skipped. Returns a small counters
    dict for logging.
    """
    counters = {"dealers": 0, "tags": 0, "categories": 0}

    # --- Dealers ---
    # NOTE: This function bypasses the `db` proxy (which auto-injects
    # `owner_id`) and writes to `real_db` directly. So we MUST set
    # `owner_id` ourselves — every other endpoint scopes by `owner_id`,
    # not `user_id`. Setting only `user_id` makes the seed records
    # invisible to the rest of the API, which is exactly the bug new
    # users hit ("0 dealers / tags / categories").
    for d in DEFAULT_DEALERS_SEED:
        # Skip if a dealer with this name already exists for the user.
        existing = await real_db.dealers.find_one(
            {"owner_id": user_id, "name": d["name"]}, {"_id": 0, "id": 1}
        )
        if existing:
            continue
        dealer = Dealer(name=d["name"])
        record = dealer.dict()
        record.update({
            "owner_id": user_id,
            "phone": d.get("phone", ""),
            "website": d.get("website", ""),
            "address": d.get("address", ""),
            "notes": d.get("notes", ""),
            "warranty_contact": d.get("warranty_contact", ""),
            "tech_support_contact": d.get("tech_support_contact", ""),
            "customer_support_contact": d.get("customer_support_contact", ""),
        })
        await real_db.dealers.insert_one(record)
        counters["dealers"] += 1

    # --- Tags ---
    for tag_name in DEFAULT_TAGS_SEED:
        existing = await real_db.tags.find_one(
            {"owner_id": user_id, "name": {"$regex": f"^{re.escape(tag_name)}$", "$options": "i"}},
            {"_id": 0, "id": 1},
        )
        if existing:
            continue
        tag = Tag(name=tag_name)
        rec = tag.dict()
        rec["owner_id"] = user_id
        await real_db.tags.insert_one(rec)
        counters["tags"] += 1

    # --- Categories ---
    for cat_name in DEFAULT_CATEGORIES_SEED:
        existing = await real_db.categories.find_one(
            {"owner_id": user_id, "name": {"$regex": f"^{re.escape(cat_name)}$", "$options": "i"}},
            {"_id": 0, "id": 1},
        )
        if existing:
            continue
        cat = Category(name=cat_name)
        rec = cat.dict()
        rec["owner_id"] = user_id
        await real_db.categories.insert_one(rec)
        counters["categories"] += 1

    return counters


# ---------------------------------------------------------------------------
# Admin: one-shot default-content backfill
# ---------------------------------------------------------------------------
# Allows an admin to retroactively seed dealers/tags/categories for any
# user (or every user) who registered before the seed function existed,
# or before the seed-field-name bug was fixed. Idempotent — names are
# de-duped, so it's safe to run multiple times. Admin-only.

class AdminSeedDefaultsRequest(BaseModel):
    # Optional. If omitted, the endpoint seeds EVERY user in the DB.
    # If provided, seeds only that one user.
    user_id: Optional[str] = None


from subscriptions import _require_admin as _require_admin_for_seed  # noqa: E402


@api_router.post("/admin/seed-defaults")
async def admin_seed_defaults(
    payload: AdminSeedDefaultsRequest,
    user: User = Depends(get_current_user),
):
    """Backfill the 5 default dealers / 3 tags / 5 categories for one user
    (when `user_id` is given) or every user (when body is `{}`).
    Idempotent — skips any item whose name already exists for that user."""
    _require_admin_for_seed(user)

    if payload.user_id:
        # Single-user mode
        target = await real_db.users.find_one(
            {"id": payload.user_id}, {"_id": 0, "id": 1, "email": 1}
        )
        if not target:
            raise HTTPException(404, "User not found")
        added = await seed_default_content_for_user(target["id"])
        return {
            "scope": "single",
            "user_id": target["id"],
            "email": target.get("email"),
            "added": added,
        }

    # All-users mode — iterate every user in the DB and seed each.
    summary: List[Dict[str, Any]] = []
    totals = {"dealers": 0, "tags": 0, "categories": 0, "users_touched": 0}
    cursor = real_db.users.find({}, {"_id": 0, "id": 1, "email": 1})
    async for u in cursor:
        added = await seed_default_content_for_user(u["id"])
        # Only record users who actually received NEW records (skips
        # users who already had all defaults — keeps the response small).
        if any(added.values()):
            summary.append({
                "user_id": u["id"],
                "email": u.get("email"),
                "added": added,
            })
            for k in ("dealers", "tags", "categories"):
                totals[k] += added[k]
        totals["users_touched"] += 1
    return {
        "scope": "all",
        "totals": totals,
        "newly_seeded_users": summary,
    }


@api_router.post("/admin/migrate-model-serial")
async def admin_migrate_model_serial(user: User = Depends(get_current_user)):
    """Backfill `model_numbers[]` for every tool that doesn't have one yet.
    Per the user's instructions:
      - Existing `model` and `serial_number` (legacy single-string model #s)
        and `set_serials` (legacy multi-model array) all merge into
        `model_numbers[]`, deduped, order preserved.
      - `serial_numbers[]` stays empty (no real serial numbers were ever
        captured).
      - Legacy fields are left intact so older app builds keep rendering.
    Idempotent — tools that already have model_numbers populated are skipped.
    Admin only."""
    _require_admin_for_seed(user)
    total = 0
    touched = 0
    cursor = real_db.tools.find({}, {"_id": 0})
    async for t in cursor:
        total += 1
        if "model_numbers" in t:
            continue  # already migrated (presence check — empty list also counts)
        candidates: List[str] = []
        for v in (t.get("set_serials") or []):
            if v:
                candidates.append(str(v).strip())
        sn = t.get("serial_number")
        if sn:
            candidates.append(str(sn).strip())
        md = t.get("model")
        if md:
            candidates.append(str(md).strip())
        seen = set()
        mns: List[str] = []
        for v in candidates:
            if v and v not in seen:
                seen.add(v)
                mns.append(v)
        update_doc: Dict[str, Any] = {
            "model_numbers": mns,
            "serial_numbers": t.get("serial_numbers") or [],
        }
        await real_db.tools.update_one({"id": t["id"]}, {"$set": update_doc})
        touched += 1
    return {"total_tools": total, "migrated": touched}



@auth_router.post("/register", response_model=AuthResponse)
async def register(payload: RegisterRequest, request: Request):
    # Rate limit: 3 new accounts per IP per hour (anti-spam).
    _enforce_rate_limit(
        "auth.register",
        _client_ip(request),
        max_count=3,
        window_seconds=3600,
        message="Too many sign-up attempts from this device. Please try again later.",
    )
    email = payload.email.strip().lower()
    existing = await real_db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        raise HTTPException(400, "Email already registered")
    user = User(
        email=email,
        password_hash=hash_password(payload.password),
        name=(payload.name or "").strip(),
    )
    await real_db.users.insert_one(user.dict())
    # First-user migration
    await _claim_orphan_data(user.id)
    # Seed helpful defaults (dealers, tags, categories) — idempotent.
    try:
        await seed_default_content_for_user(user.id)
    except Exception as e:
        # Never fail registration over seed errors — log and move on.
        logging.getLogger("server").warning(
            "Default-content seed failed for user %s: %s", user.id, e
        )
    token = create_token(user.id)
    return AuthResponse(token=token, user=to_public(user))


@auth_router.post("/login", response_model=AuthResponse)
async def login(request: Request):
    """Login. Always responds with a uniform 401 for any auth failure
    (bad email format, unknown email, wrong password) so that an attacker
    cannot enumerate which emails are registered.
    """
    # Rate limit: 5 login attempts per IP per minute (anti brute-force).
    _enforce_rate_limit(
        "auth.login",
        _client_ip(request),
        max_count=5,
        window_seconds=60,
        message="Too many login attempts. Please wait a minute and try again.",
    )
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(401, "Invalid email or password")
    email = (payload.get("email") or "").strip().lower() if isinstance(payload, dict) else ""
    password = payload.get("password") or "" if isinstance(payload, dict) else ""
    if not email or "@" not in email or "." not in email or not password:
        raise HTTPException(401, "Invalid email or password")
    udoc = await real_db.users.find_one({"email": email}, {"_id": 0})
    if not udoc:
        raise HTTPException(401, "Invalid email or password")
    if not verify_password(password, udoc.get("password_hash", "")):
        raise HTTPException(401, "Invalid email or password")
    user = User(**udoc)
    token = create_token(user.id)
    return AuthResponse(token=token, user=to_public(user))


@auth_router.get("/me", response_model=UserPublic)
async def me(user: User = Depends(get_current_user)):
    return to_public(user)


@auth_router.put("/me", response_model=UserPublic)
async def update_me(payload: Dict[str, Any], user: User = Depends(get_current_user)):
    updates: Dict[str, Any] = {}
    if "name" in payload:
        updates["name"] = (payload.get("name") or "").strip()
    if "password" in payload and payload["password"]:
        updates["password_hash"] = hash_password(payload["password"])
    if updates:
        updates["updated_at"] = now_iso()
        await real_db.users.update_one({"id": user.id}, {"$set": updates})
        user_doc = await real_db.users.find_one({"id": user.id}, {"_id": 0})
        user = User(**user_doc)
    return to_public(user)


# ---------------------------------------------------------------------------
# Password reset (Forgot Password) — 6-digit code via email
# ---------------------------------------------------------------------------
import secrets

RESET_CODE_TTL_MINUTES = 15
RESET_CODE_MAX_ATTEMPTS = 5


class ForgotPasswordRequest(BaseModel):
    email: str


class ResetPasswordRequest(BaseModel):
    email: str
    code: str
    new_password: str


def _generate_reset_code() -> str:
    # 6-digit numeric code
    return f"{secrets.randbelow(1_000_000):06d}"


@auth_router.post("/forgot-password")
async def forgot_password(payload: ForgotPasswordRequest, request: Request):
    """Send a 6-digit password reset code to the user's email.

    Always returns 200 with the same message regardless of whether the email
    exists in our system. This prevents email-enumeration attacks. The
    email is only sent when the user actually exists.
    """
    # Rate limit: 3 reset-code requests per IP per hour (anti email-bombing).
    _enforce_rate_limit(
        "auth.forgot",
        _client_ip(request),
        max_count=3,
        window_seconds=3600,
        message="Too many reset code requests. Please try again in an hour.",
    )
    email = (payload.email or "").strip().lower()
    generic_response = {
        "ok": True,
        "message": "If that email is registered, a 6-digit code has been sent.",
    }
    if not email or "@" not in email:
        return generic_response

    udoc = await real_db.users.find_one({"email": email}, {"_id": 0, "id": 1, "name": 1})
    if not udoc:
        # Silently pretend we sent it
        return generic_response

    code = _generate_reset_code()
    code_hash = hash_password(code)  # store hash, never plaintext
    expires_at = (datetime.now(timezone.utc) + timedelta(minutes=RESET_CODE_TTL_MINUTES)).isoformat()

    # Upsert reset record (only one active reset per user at a time)
    await real_db.password_resets.update_one(
        {"user_id": udoc["id"]},
        {
            "$set": {
                "user_id": udoc["id"],
                "email": email,
                "code_hash": code_hash,
                "expires_at": expires_at,
                "attempts": 0,
                "created_at": now_iso(),
            }
        },
        upsert=True,
    )

    # Fire-and-forget the email (never block the response on SMTP)
    try:
        send_password_reset_code(email, code, display_name=udoc.get("name") or "")
    except Exception as e:
        logging.error("Failed to send reset email to %s: %s", email, e)

    return generic_response


@auth_router.post("/reset-password", response_model=AuthResponse)
async def reset_password(payload: ResetPasswordRequest, request: Request):
    """Verify the 6-digit code and set a new password. On success, returns a
    fresh auth token so the user is logged in immediately.
    """
    # Rate limit: 5 reset attempts per IP per minute (anti code brute-force).
    _enforce_rate_limit(
        "auth.reset",
        _client_ip(request),
        max_count=5,
        window_seconds=60,
        message="Too many reset attempts. Please wait a minute and try again.",
    )
    email = (payload.email or "").strip().lower()
    code = (payload.code or "").strip()
    new_password = payload.new_password or ""

    if not email or not code or not new_password:
        raise HTTPException(400, "Email, code, and new password are required.")
    if len(new_password) < 6:
        raise HTTPException(400, "New password must be at least 6 characters.")

    udoc = await real_db.users.find_one({"email": email}, {"_id": 0})
    if not udoc:
        # Don't leak whether the email exists — generic failure
        raise HTTPException(400, "Invalid or expired code.")

    reset_doc = await real_db.password_resets.find_one({"user_id": udoc["id"]}, {"_id": 0})
    if not reset_doc:
        raise HTTPException(400, "Invalid or expired code.")

    # Check expiry
    try:
        expires_at = datetime.fromisoformat(reset_doc["expires_at"])
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
    except Exception:
        expires_at = None
    if not expires_at or datetime.now(timezone.utc) > expires_at:
        await real_db.password_resets.delete_one({"user_id": udoc["id"]})
        raise HTTPException(400, "Invalid or expired code.")

    # Check attempts
    if (reset_doc.get("attempts", 0) or 0) >= RESET_CODE_MAX_ATTEMPTS:
        await real_db.password_resets.delete_one({"user_id": udoc["id"]})
        raise HTTPException(
            429,
            "Too many incorrect attempts. Please request a new code.",
        )

    # Verify the code
    if not verify_password(code, reset_doc.get("code_hash", "")):
        await real_db.password_resets.update_one(
            {"user_id": udoc["id"]},
            {"$inc": {"attempts": 1}},
        )
        raise HTTPException(400, "Invalid or expired code.")

    # Success — update the password and burn the reset record
    await real_db.users.update_one(
        {"id": udoc["id"]},
        {"$set": {"password_hash": hash_password(new_password), "updated_at": now_iso()}},
    )
    await real_db.password_resets.delete_one({"user_id": udoc["id"]})

    user_doc = await real_db.users.find_one({"id": udoc["id"]}, {"_id": 0})
    user = User(**user_doc)
    token = create_token(user.id)
    return AuthResponse(token=token, user=to_public(user))


# ---------- DELETE ACCOUNT (irreversible) ----------
class DeleteAccountRequest(BaseModel):
    password: str


# Every collection that holds user-owned data — must wipe each on account delete.
USER_DATA_COLLECTIONS = [
    "tools",
    "dealers",
    "borrowers",
    "locations",
    "tags",
    "categories",
    "wishlist_items",
    "warranty_claims",
    "personal_profile",
    "checkout_history",
    "feedback",
    "password_resets",
    "saved_reports",
    "saved_report_presets",
    "report_presets",
    "maintenance_log",
    "maintenance_schedules",
    "tool_documents",
    "user_prefs",
    "preferences",
    "user_preferences",
]


@auth_router.delete("/account")
async def delete_account(
    payload: DeleteAccountRequest, user: User = Depends(get_current_user)
):
    """Permanently delete the authenticated user's account and ALL associated
    data across every collection. Requires the user's password as a
    confirmation. Returns 401 on wrong password, 200 on success.

    Idempotent across collections that may not exist (motor `delete_many`
    on a non-existent collection is a no-op).
    """
    udoc = await real_db.users.find_one({"id": user.id}, {"_id": 0})
    if not udoc:
        raise HTTPException(401, "Account not found")
    if not verify_password(payload.password or "", udoc.get("password_hash", "")):
        raise HTTPException(401, "Incorrect password")

    uid = user.id
    deleted = {"user_id": uid, "collections": {}, "total": 0}
    for coll in USER_DATA_COLLECTIONS:
        try:
            res = await real_db[coll].delete_many({"owner_id": uid})
            n = int(getattr(res, "deleted_count", 0) or 0)
            if n > 0:
                deleted["collections"][coll] = n
                deleted["total"] += n
        except Exception as e:
            logging.warning("Delete-account: skip collection %s (%s)", coll, e)

    # Also wipe any password reset rows tied to this user (might have been
    # missed if owner_id wasn't set on those rows).
    try:
        await real_db.password_resets.delete_many({"user_id": uid})
    except Exception:
        pass
    try:
        await real_db.password_resets.delete_many({"email": (udoc.get("email") or "").lower()})
    except Exception:
        pass

    # Finally remove the user record itself — by id AND by email (case-
    # insensitive), so the email is GUARANTEED freed for re-registration
    # even in the unlikely case of duplicate rows or case mismatches.
    user_email_lower = (udoc.get("email") or "").strip().lower()
    try:
        await real_db.users.delete_one({"id": uid})
    except Exception as e:
        logging.error("Delete-account: failed to delete user record by id: %s", e)
        raise HTTPException(500, "Could not delete account record")
    if user_email_lower:
        try:
            # Sweep any stragglers that match the email (case-insensitive).
            email_re = {"$regex": f"^{re.escape(user_email_lower)}$", "$options": "i"}
            extra = await real_db.users.delete_many({"email": email_re})
            if getattr(extra, "deleted_count", 0):
                logging.info(
                    "Delete-account: cleaned %d stray user row(s) for email %s",
                    extra.deleted_count,
                    user_email_lower,
                )
        except Exception as e:
            logging.warning("Delete-account: email sweep failed: %s", e)

    # Sanity check — log if any user with this email survived (should be 0)
    try:
        survivors = await real_db.users.count_documents(
            {"email": {"$regex": f"^{re.escape(user_email_lower)}$", "$options": "i"}}
            if user_email_lower
            else {"id": uid}
        )
        if survivors:
            logging.warning(
                "Delete-account: %d residual user row(s) remain for %s — email may still be locked.",
                survivors,
                user_email_lower or uid,
            )
    except Exception:
        pass

    logging.info("Account deleted: %s (%s) — %d records purged", uid, user_email_lower, deleted["total"])
    return {"ok": True, "deleted": deleted, "message": "Account permanently deleted."}


# ---------------------------------------------------------------------------
# AI RECEIPT SCANNER
# ---------------------------------------------------------------------------
class ReceiptScanRequest(BaseModel):
    image_base64: str  # raw base64 (no data: prefix needed; we strip it if present)


class ReceiptItem(BaseModel):
    name: Optional[str] = ""
    brand: Optional[str] = ""
    model: Optional[str] = ""
    serial_number: Optional[str] = ""
    cost: Optional[float] = 0.0
    quantity: Optional[int] = 1
    description: Optional[str] = ""


class ReceiptScanResponse(BaseModel):
    # Receipt-level fields (apply to ALL items on the receipt)
    dealer: Optional[str] = ""
    sold_by: Optional[str] = ""        # Sales rep / agent who sold (e.g. "Sold By: Wade Miller")
    purchase_date: Optional[str] = ""  # ISO YYYY-MM-DD (normalised)
    raw_text: Optional[str] = ""       # Full OCR transcription (so user can copy missing values)
    items: List[ReceiptItem] = []
    # Backward-compat top-level fields = mirror of items[0] when present
    name: Optional[str] = ""
    brand: Optional[str] = ""
    model: Optional[str] = ""
    serial_number: Optional[str] = ""
    cost: Optional[float] = 0.0
    quantity: Optional[int] = 1
    description: Optional[str] = ""
    raw: Optional[Dict[str, Any]] = None


def _normalize_date(s: str) -> str:
    """Best-effort normalise common receipt date formats to YYYY-MM-DD.
    Handles 'M/D/YYYY', 'MM/DD/YYYY', 'M-D-YYYY', 'YYYY/MM/DD', 'YYYY-MM-DD',
    'D/M/YYYY' (ambiguous — assumes US M/D when day<=12). Returns '' on failure.
    """
    if not s:
        return ""
    txt = str(s).strip()
    if not txt:
        return ""
    # Strip time-of-day if present
    txt = re.split(r"\s+", txt, maxsplit=1)[0]
    # Already ISO?
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", txt)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    # YYYY/MM/DD
    m = re.match(r"^(\d{4})/(\d{1,2})/(\d{1,2})$", txt)
    if m:
        y, mo, d = m.groups()
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    # M/D/YYYY or M-D-YYYY (US format — assume M first)
    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})$", txt)
    if m:
        a, b, c = m.groups()
        a_i, b_i, c_i = int(a), int(b), int(c)
        if c_i < 100:
            c_i += 2000 if c_i < 70 else 1900  # 2-digit year
        # Decide M/D vs D/M — if first part > 12 it must be a day
        if a_i > 12 and b_i <= 12:
            mo, d, y = b_i, a_i, c_i
        else:
            mo, d, y = a_i, b_i, c_i
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    return ""


@api_router.post("/ai/receipt-scan", response_model=ReceiptScanResponse)
async def ai_receipt_scan(payload: ReceiptScanRequest, user: User = Depends(get_current_user)):
    """Send the receipt image to GPT-4o vision and extract structured fields.
    Returns best-guess values that the user reviews/edits in the mapping UI
    before they are committed to a tool record.

    Supports multi-item receipts: the response includes a list of items so the
    frontend can prompt the user to pick which one to add.
    """
    # Rate limit: 30 AI receipt scans per user per hour (protect LLM budget).
    _enforce_rate_limit(
        "ai.receipt_scan",
        user.id,
        max_count=30,
        window_seconds=3600,
        message="You have used the AI receipt scanner a lot in the last hour. "
        "Please try again later.",
    )
    import base64
    import json
    import re
    import tempfile
    import os as _os

    raw_b64 = payload.image_base64 or ""
    if not raw_b64:
        raise HTTPException(status_code=400, detail="image_base64 is required")
    # Strip data URL prefix if present
    if raw_b64.startswith("data:"):
        comma = raw_b64.find(",")
        if comma > 0:
            raw_b64 = raw_b64[comma + 1 :]
    try:
        image_bytes = base64.b64decode(raw_b64, validate=False)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid base64 image")

    # Write to a temp file so emergentintegrations can pass it as a file path
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg")
    try:
        tmp.write(image_bytes)
        tmp.flush()
        tmp.close()

        from emergentintegrations.llm.chat import (
            LlmChat,
            UserMessage,
            ImageContent,
        )

        api_key = _os.environ.get("EMERGENT_LLM_KEY", "")
        if not api_key:
            raise HTTPException(
                status_code=500,
                detail="EMERGENT_LLM_KEY missing in backend environment",
            )

        system_prompt = (
            "You are a precise receipt-OCR extractor for a tool inventory app. "
            "Given an image of a receipt, packing slip, invoice, or product box, "
            "return ONLY a JSON object — no markdown, no commentary.\n\n"
            "Receipts often contain MULTIPLE distinct line items (different "
            "tools / parts purchased). Identify EVERY purchased line item. "
            "Skip subtotals, taxes, discounts, shipping, fees, totals, "
            "salesperson lines (those go in 'sold_by'), account numbers, "
            "transaction history rows, finance charges, payments — list ONLY "
            "actual products being purchased.\n\n"
            "Return EXACTLY this JSON shape:\n"
            "{\n"
            '  "dealer": "<store / seller / vendor / franchisee name — often a logo at top, e.g. \'Snap-on\', \'Cornwell\', \'Matco\'>",\n'
            '  "sold_by": "<the sales rep / agent who sold it — appears as \'Sold By:\', \'Salesperson:\', \'Rep:\', \'Sold by:\' on the receipt>",\n'
            '  "purchase_date": "YYYY-MM-DD (use ISO format; convert M/D/YYYY → YYYY-MM-DD)",\n'
            '  "raw_text": "<full OCR transcription of the receipt, line by line, as you read it>",\n'
            '  "items": [\n'
            "    {\n"
            '      "name": "<short product/tool name from the Description column>",\n'
            '      "brand": "<manufacturer / brand — often the same as the dealer for branded receipts>",\n'
            '      "model": "<model number / model name>",\n'
            '      "serial_number": "<model #, part #, item #, sku, catalog # — see note below>",\n'
            '      "cost": <number, no currency symbols>,\n'
            '      "quantity": <integer, default 1>,\n'
            '      "description": "<very brief 1-sentence description>"\n'
            "    }\n"
            "  ]\n"
            "}\n\n"
            "Important rules:\n"
            "- Receipts use various labels for the unique product identifier. "
            "Treat ANY of these as 'serial_number': 'Part #', 'Part Number', "
            "'Item #', 'Item Number', 'SKU', 'Catalog #', 'Catalog Number', "
            "'Product #', 'Product Code', 'Stock #'. Use the literal value "
            "shown next to the label.\n"
            "- 'cost' is the per-unit price (or extended/total if per-unit "
            "isn't shown). Strip currency symbols and commas. Always a number.\n"
            "- 'quantity' must be a positive integer (default 1).\n"
            "- 'sold_by' is the human salesperson/rep who sold it (e.g. "
            "'Wade Miller'). Receipts label this as 'Sold By:' or "
            "'Salesperson:'. Leave empty if not present.\n"
            "- 'purchase_date' MUST be ISO YYYY-MM-DD. Convert any other "
            "format (5/7/2025 → 2025-05-07).\n"
            "- Use empty strings or 0 if a value isn't present. Do NOT invent.\n"
            "- 'raw_text' is REQUIRED — include every line of text from the "
            "receipt so the user can copy any value the structured extraction "
            "missed.\n"
            "- Output ONLY the JSON object. No prose, no markdown fences."
        )

        chat = (
            LlmChat(
                api_key=api_key,
                session_id=f"receipt-scan-{uuid.uuid4()}",
                system_message=system_prompt,
            )
            .with_model("openai", "gpt-4o")
            .with_params(max_tokens=2000)
        )

        msg = UserMessage(
            text=(
                "Extract every line item from this receipt. Return ONLY the "
                "JSON object exactly as instructed."
            ),
            file_contents=[ImageContent(image_base64=raw_b64)],
        )
        response_text = await chat.send_message(msg)

        # Try parse JSON. Strip markdown fences if present, then regex-extract
        # the outermost JSON object.
        cleaned = (response_text or "").strip()
        # Strip ```json ... ``` fences
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s*```$", "", cleaned)
        m = re.search(r"\{.*\}", cleaned, re.DOTALL)
        json_text = m.group(0) if m else cleaned
        try:
            data = json.loads(json_text)
        except Exception:
            data = {}

        # Build items list — accept either {items:[...]} or a single-item shape
        raw_items = data.get("items")
        items_list: List[ReceiptItem] = []
        if isinstance(raw_items, list) and raw_items:
            for it in raw_items:
                if not isinstance(it, dict):
                    continue
                items_list.append(
                    ReceiptItem(
                        name=str(it.get("name") or "").strip(),
                        brand=str(it.get("brand") or "").strip(),
                        model=str(it.get("model") or "").strip(),
                        serial_number=str(it.get("serial_number") or "").strip(),
                        cost=_to_float(it.get("cost", 0)),
                        quantity=_to_int(it.get("quantity", 1), default=1),
                        description=str(it.get("description") or "").strip(),
                    )
                )
        # Legacy / single-item fallback — older prompts returned flat fields
        if not items_list and any(
            data.get(k) for k in ("name", "brand", "model", "serial_number", "cost")
        ):
            items_list.append(
                ReceiptItem(
                    name=str(data.get("name") or "").strip(),
                    brand=str(data.get("brand") or "").strip(),
                    model=str(data.get("model") or "").strip(),
                    serial_number=str(data.get("serial_number") or "").strip(),
                    cost=_to_float(data.get("cost", 0)),
                    quantity=_to_int(data.get("quantity", 1), default=1),
                    description=str(data.get("description") or "").strip(),
                )
            )

        first = items_list[0] if items_list else ReceiptItem()
        return ReceiptScanResponse(
            dealer=str(data.get("dealer") or "").strip(),
            sold_by=str(data.get("sold_by") or "").strip(),
            purchase_date=_normalize_date(str(data.get("purchase_date") or "").strip()),
            raw_text=str(data.get("raw_text") or "").strip(),
            items=items_list,
            # Mirror first item to top-level for backward compatibility
            name=first.name,
            brand=first.brand,
            model=first.model,
            serial_number=first.serial_number,
            cost=first.cost,
            quantity=first.quantity,
            description=first.description,
            raw=data if isinstance(data, dict) else None,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI receipt scan failed: {e}")
    finally:
        try:
            _os.unlink(tmp.name)
        except Exception:
            pass


@api_router.post("/ocr/receipt", response_model=ReceiptScanResponse)
async def ocr_receipt_alias(payload: ReceiptScanRequest, user: User = Depends(get_current_user)):
    """Alias for /api/ai/receipt-scan — kept for forward-compatibility with
    any older clients or tooling that hits the simpler `/ocr/receipt` path."""
    return await ai_receipt_scan(payload, user)




app.include_router(api_router)
app.include_router(auth_router)


# ---------------------------------------------------------------------------
# Subscription / Entitlement router (RevenueCat webhook + /subscription + /promo/redeem)
# ---------------------------------------------------------------------------
from subscriptions import make_router as _make_subscriptions_router  # noqa: E402

app.include_router(_make_subscriptions_router(real_db, get_current_user))

# ---------------------------------------------------------------------------
# Database backups (audit #17): admin endpoints + monthly scheduler.
# ---------------------------------------------------------------------------
from backups import (  # noqa: E402
    make_backup_router as _make_backup_router,
    start_backup_scheduler as _start_backup_scheduler,
)
from subscriptions import _require_admin as _require_admin_for_backups  # noqa: E402

app.include_router(
    _make_backup_router(
        lambda: real_db,
        get_current_user,
        _require_admin_for_backups,
    )
)


@app.on_event("startup")
async def _kick_off_backup_scheduler() -> None:
    """Idempotently start the monthly DB backup task on every boot."""
    _start_backup_scheduler(lambda: real_db)


# ---------------------------------------------------------------------------
# Setup-Guides viewer (/api/guides)
# Renders the markdown setup guides under /app/memory/ as a single styled
# HTML page so the user can bookmark it and reference it any time.
# ---------------------------------------------------------------------------
from fastapi.responses import HTMLResponse  # noqa: E402

_GUIDES_DIR = Path(__file__).parent.parent / "memory"
_GUIDE_FILES = [
    ("migration", "MIGRATION_GUIDE.md",              "MIGRATION — Take Everything Off This Container"),
    ("apple",      "setup_apple_subscriptions.md", "Apple App Store Connect — Subscription Setup"),
    ("google",     "setup_google_subscriptions.md", "Google Play Console — Subscription Setup"),
    ("revenuecat", "setup_revenuecat.md",           "RevenueCat — Setup"),
    ("privacy",    "PRIVACY_POLICY_TEMPLATE.md",    "Privacy Policy — Template"),
    ("terms",      "TERMS_OF_SERVICE_TEMPLATE.md",  "Terms of Service — Template"),
]


def _render_guides_html() -> str:
    try:
        import markdown as _md  # type: ignore
    except Exception:  # pragma: no cover
        _md = None

    sections = []
    toc = []
    for slug, fname, title in _GUIDE_FILES:
        fpath = _GUIDES_DIR / fname
        try:
            raw = fpath.read_text(encoding="utf-8")
        except FileNotFoundError:
            raw = f"_(file `{fname}` not found)_"
        if _md:
            html = _md.markdown(
                raw,
                extensions=["extra", "sane_lists", "toc", "tables", "fenced_code"],
            )
        else:
            # Fallback: <pre> if Python markdown isn't installed
            html = f"<pre>{raw}</pre>"
        sections.append(f'<section id="{slug}"><h1>{title}</h1>{html}</section>')
        toc.append(f'<a href="#{slug}">{title}</a>')

    css = """
    :root { color-scheme: light dark; }
    * { box-sizing: border-box; }
    body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
           line-height: 1.6; max-width: 880px; margin: 0 auto; padding: 24px;
           color: #1a1a1a; background: #fff; }
    @media (prefers-color-scheme: dark) {
        body { color: #e8e8e8; background: #111; }
        a { color: #4ea1ff; }
        .nav { background: #1c1c1e; border-color: #333; }
        code, pre { background: #1c1c1e !important; }
    }
    .nav { position: sticky; top: 0; background: #fff; padding: 12px 16px; margin: -24px -24px 24px;
           border-bottom: 1px solid #e5e5e5; display: flex; flex-wrap: wrap; gap: 6px 12px;
           font-size: 13px; z-index: 10; }
    .nav strong { width: 100%; font-size: 12px; color: #888; text-transform: uppercase; letter-spacing: .04em; }
    .nav a { color: #4ea1ff; text-decoration: none; padding: 4px 10px; border: 1px solid #4ea1ff; border-radius: 6px; }
    .nav a:hover { background: rgba(78,161,255,.1); }
    section { padding: 24px 0 48px; border-bottom: 2px solid #e5e5e5; }
    section:last-child { border-bottom: none; }
    h1 { font-size: 26px; padding-bottom: 8px; border-bottom: 2px solid #4ea1ff; }
    h2 { font-size: 20px; margin-top: 28px; padding-bottom: 4px; border-bottom: 1px solid #e5e5e5; }
    h3 { font-size: 16px; margin-top: 20px; }
    code { background: #f6f6f6; padding: 2px 6px; border-radius: 4px; font-size: 90%; }
    pre { background: #f6f6f6; padding: 12px; border-radius: 6px; overflow-x: auto; }
    pre code { background: transparent; padding: 0; }
    table { border-collapse: collapse; width: 100%; margin: 14px 0; font-size: 14px; }
    th, td { border: 1px solid #ccc; padding: 8px 10px; text-align: left; vertical-align: top; }
    th { background: rgba(127,127,127,.08); }
    blockquote { margin: 12px 0; padding: 8px 14px; border-left: 4px solid #4ea1ff;
                 background: rgba(78,161,255,.05); }
    hr { border: 0; border-top: 1px solid #e5e5e5; margin: 24px 0; }
    li input[type=checkbox] { margin-right: 6px; }
    """
    body = (
        '<div class="nav">'
        '<strong>Toolbox Vault — Setup Guides</strong>'
        + "".join(toc) +
        '</div>'
        + "".join(sections)
    )
    return (
        "<!DOCTYPE html>\n<html lang='en'>\n<head>"
        "<meta charset='utf-8'>"
        "<meta name='viewport' content='width=device-width, initial-scale=1'>"
        "<title>Toolbox Vault — Setup Guides</title>"
        f"<style>{css}</style>"
        "</head><body>"
        + body +
        "</body></html>"
    )


@app.get("/api/guides", response_class=HTMLResponse)
async def get_setup_guides():
    """Public bookmarkable page that renders all 5 setup guides as one HTML doc.

    No auth required — these guides contain no secrets. They are convenience
    documentation for the app owner.
    """
    return HTMLResponse(_render_guides_html())


# ---------------------------------------------------------------------------
# Migration download — token-protected endpoints to download the code +
# database tarballs created at /app/migration. Token lives in
# MIGRATION_DOWNLOAD_TOKEN env var. Used when the project owner wants to
# leave the platform and take their data with them.
# ---------------------------------------------------------------------------
from fastapi.responses import FileResponse  # noqa: E402

_MIG_DIR = Path(__file__).parent.parent / "migration"


@app.get("/api/migration/{filename}")
async def migration_download(filename: str, token: str = ""):
    if not token or token != os.environ.get("MIGRATION_DOWNLOAD_TOKEN", "").strip():
        raise HTTPException(403, "Invalid or missing token")
    # Only allow the two specific tarballs
    if filename not in ("mongo-dump.tar.gz", "toolbox-vault-code.tar.gz"):
        raise HTTPException(404, "File not available")
    fpath = _MIG_DIR / filename
    if not fpath.exists():
        raise HTTPException(404, "Archive not generated")
    return FileResponse(
        path=str(fpath),
        media_type="application/gzip",
        filename=filename,
    )


# ---------------------------------------------------------------------------
# Feedback endpoint — registered directly on app (api_router is already included above)
# ---------------------------------------------------------------------------
FEEDBACK_DEST_EMAIL = os.environ.get("GMAIL_FROM_ADDRESS", "MechanicVault@gmail.com")
FEEDBACK_RATE_WINDOW_SECONDS = 600  # 10 minutes
FEEDBACK_RATE_MAX = 5              # 5 submissions per IP per 10 min

# In-memory rate-limit bucket: {ip: [timestamp, ...]}
_feedback_rate_buckets: Dict[str, List[float]] = {}


class FeedbackRequest(BaseModel):
    name: str
    email: str
    subject: str
    message: str
    platform: Optional[str] = ""
    is_bug: bool = False
    is_feature: bool = False
    app_version: Optional[str] = ""
    # Honeypot — bots fill hidden fields; humans don't.
    website: Optional[str] = ""
    # Optional screenshot — user-attached PNG/JPEG image, base64-encoded.
    # Frontend caps at ~1 MB before sending to keep payload manageable.
    screenshot_base64: Optional[str] = None


def _feedback_rate_limit(ip: str) -> bool:
    """Return True if the IP is *under* the limit (allowed), False if rate-limited."""
    import time as _time
    now = _time.time()
    cutoff = now - FEEDBACK_RATE_WINDOW_SECONDS
    bucket = [t for t in _feedback_rate_buckets.get(ip, []) if t > cutoff]
    if len(bucket) >= FEEDBACK_RATE_MAX:
        _feedback_rate_buckets[ip] = bucket
        return False
    bucket.append(now)
    _feedback_rate_buckets[ip] = bucket
    return True


@app.post("/api/feedback")
async def submit_feedback(payload: FeedbackRequest, request: Request):
    """Receive feedback / bug report / feature request and email it to the
    operator (MechanicVault@gmail.com). Reply-To is set to the user's email
    so operator replies go straight to the user.
    """
    # Honeypot — silently drop bot traffic
    if (payload.website or "").strip():
        return {"ok": True, "message": "Thanks!"}

    # Basic validation
    name = (payload.name or "").strip()
    email_addr = (payload.email or "").strip()
    subject = (payload.subject or "").strip()
    message = (payload.message or "").strip()
    if not name:
        raise HTTPException(400, "Please provide your name.")
    if not email_addr or "@" not in email_addr or "." not in email_addr:
        raise HTTPException(400, "Please provide a valid email address.")
    if not subject:
        raise HTTPException(400, "Please provide a subject.")
    if not message:
        raise HTTPException(400, "Please provide a message.")
    if len(message) > 20000:
        raise HTTPException(400, "Message is too long.")

    # Rate-limit by IP. Behind K8s ingress, request.client.host is always the
    # ingress pod IP, so prefer X-Forwarded-For for the real client IP.
    xff = (request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
    client_ip = xff or (request.client.host if request.client else "unknown")
    if not _feedback_rate_limit(client_ip):
        raise HTTPException(
            429,
            "Too many messages from this device. Please try again in a few minutes.",
        )

    # Cap screenshot at 2 MB raw (~2.7 MB base64) to keep DB and email sane.
    screenshot_b64 = payload.screenshot_base64 or None
    if screenshot_b64 and len(screenshot_b64) > 3_500_000:
        screenshot_b64 = None  # silently drop oversized; user already submitted

    # Persist a record (so operator has a searchable log even if email fails)
    record = {
        "id": str(uuid.uuid4()),
        "name": name,
        "email": email_addr,
        "subject": subject,
        "message": message,
        "platform": (payload.platform or "").strip(),
        "is_bug": bool(payload.is_bug),
        "is_feature": bool(payload.is_feature),
        "app_version": (payload.app_version or "").strip(),
        "ip": client_ip,
        "has_screenshot": bool(screenshot_b64),
        "screenshot_base64": screenshot_b64,
        "created_at": now_iso(),
    }
    try:
        await real_db.feedback.insert_one(record)
    except Exception as e:
        logging.error("Failed to persist feedback record: %s", e)

    # Send the email (non-blocking-style: log errors but return success)
    sent = False
    try:
        sent = send_feedback_email(
            to_address=FEEDBACK_DEST_EMAIL,
            from_name=name,
            from_email=email_addr,
            subject=subject,
            message=message,
            is_bug=bool(payload.is_bug),
            is_feature=bool(payload.is_feature),
            platform=(payload.platform or ""),
            app_version=(payload.app_version or ""),
            screenshot_base64=screenshot_b64,
        )
    except Exception as e:
        logging.error("send_feedback_email raised: %s", e)

    if not sent:
        # Don't fail hard — the record is saved; operator can still see it.
        logging.warning("Feedback record %s saved but email send failed.", record["id"])

    return {"ok": True, "message": "Thanks — your message has been sent."}

# ---------------------------------------------------------------------------
# Unified report engine (HTML→PDF + CSV) — see /app/backend/reports.py
# ---------------------------------------------------------------------------
from reports import make_reports_router as _make_reports_router  # noqa: E402

_reports_api_router = APIRouter(prefix="/api")
_make_reports_router(_reports_api_router, lambda: db, get_current_user)
app.include_router(_reports_api_router)

# ---------------------------------------------------------------------------
# HTML → PDF rendering endpoint (uses xhtml2pdf, runs entirely server-side
# so reports work reliably regardless of browser quirks / CSP restrictions).
# ---------------------------------------------------------------------------
from fastapi import Body
from fastapi.responses import Response as FastAPIResponse
from io import BytesIO
import re as _re_html

# xhtml2pdf is intentionally imported lazily (heavy import, ~1s).
_pisa = None


def _get_pisa():
    global _pisa
    if _pisa is None:
        from xhtml2pdf import pisa as _pisa_mod  # noqa: WPS433
        _pisa = _pisa_mod
    return _pisa


def _sanitize_html_for_pdf(html: str) -> str:
    """xhtml2pdf is a strict, non-browser HTML parser. It chokes on
    a handful of modern features that browsers tolerate. Sanitise so
    the parser never errors on safe-to-ignore CSS / structures.

    Note: word-boundary lookbehind/lookahead are crucial — without them,
    `transform` would also strip the value of `text-transform`, leaving
    malformed CSS like `text- display: block;`.
    """
    # The list of property NAMES (full names only — no prefix overlap).
    # Note: page-break-* IS supported by xhtml2pdf and must NOT be stripped.
    bad_props = [
        "object-fit",
        "gap",
        "row-gap",
        "column-gap",
        "grid-template-columns",
        "grid-template-rows",
        "grid-template-areas",
        "grid-area",
        "grid-column",
        "grid-row",
        "grid-auto-flow",
        "tab-size",
        "will-change",
        "backdrop-filter",
        "box-shadow",
        "transform",
        "transition",
        "animation",
        "-webkit-print-color-adjust",
        "print-color-adjust",
        "flex",
        "flex-direction",
        "flex-wrap",
        "flex-flow",
        "flex-shrink",
        "flex-grow",
        "flex-basis",
        "justify-content",
        "justify-items",
        "justify-self",
        "align-items",
        "align-content",
        "align-self",
        "place-items",
        "place-content",
        "order",
        "filter",
    ]
    # `(?<![\w-])` — must NOT be preceded by a word char or '-'  (so `text-transform`
    # won't match `transform`). `(?=\s*:)` — must be followed by a colon.
    name_alt = "|".join(_re_html.escape(p) for p in bad_props)
    bad_decl = _re_html.compile(
        rf"(?<![\w-])(?:{name_alt})\s*:[^;}}]*;?",
        _re_html.IGNORECASE,
    )
    return bad_decl.sub("", html)


@app.post("/api/render-pdf")
async def render_pdf(
    payload: dict = Body(...),
    user=Depends(get_current_user),
):
    """Convert an HTML report payload into a PDF binary.

    Body: { "html": "<...>", "filename": "report.pdf" }
    Returns: application/pdf with Content-Disposition attachment.
    """
    # Rate limit: 20 PDF renders per user per hour (protect server CPU).
    _enforce_rate_limit(
        "pdf.render",
        user.id,
        max_count=20,
        window_seconds=3600,
        message="You have generated a lot of PDFs in the last hour. "
        "Please wait a bit before generating more.",
    )
    html = payload.get("html") or ""
    filename = payload.get("filename") or "report.pdf"
    if not html:
        raise HTTPException(400, "Missing 'html'")
    if not filename.lower().endswith(".pdf"):
        filename = f"{filename}.pdf"

    safe_html = _sanitize_html_for_pdf(html)
    pisa = _get_pisa()
    buf = BytesIO()
    try:
        result = pisa.CreatePDF(safe_html, dest=buf, encoding="utf-8")
    except Exception as exc:  # pragma: no cover
        raise HTTPException(500, f"PDF generation failed: {exc!s}") from exc

    if result.err:
        raise HTTPException(500, f"PDF generation failed (errors={result.err})")

    return FastAPIResponse(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


app.add_middleware(
    CORSMiddleware,
    # We authenticate with a JWT in the Authorization header (not cookies),
    # so we do not need credentialed CORS. Disabling credentials lets us
    # safely keep allow_origins=['*'] (Starlette refuses '*' + credentials=True).
    allow_credentials=False,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


# ---------------------------------------------------------------------------
# MongoDB indexes — ensure on startup. Idempotent: creating an existing index
# is a no-op. All indexes are non-unique (except where noted) and additive,
# so this cannot change query results — only speed them up.
# ---------------------------------------------------------------------------
INDEX_PLAN = [
    # collection,            keys,                                     options
    ("users",                [("email", 1)],                          {"unique": True, "name": "email_unique"}),
    ("users",                [("id", 1)],                              {"unique": True, "name": "id_unique"}),
    ("tools",                [("owner_id", 1)],                        {"name": "owner_id_idx"}),
    ("tools",                [("owner_id", 1), ("name", 1)],           {"name": "owner_name_idx"}),
    ("tools",                [("owner_id", 1), ("dealer_id", 1)],      {"name": "owner_dealer_idx"}),
    ("tools",                [("owner_id", 1), ("status", 1)],         {"name": "owner_status_idx"}),
    ("tools",                [("owner_id", 1), ("location_id", 1)],    {"name": "owner_location_idx"}),
    ("tools",                [("owner_id", 1), ("for_sale", 1)],       {"name": "owner_for_sale_idx"}),
    # Audit #16: hot-path compound indexes for filters/alerts that currently
    # do collection scans within owner_id partition. Each saves a partition
    # scan on lists that grow with tool count.
    ("tools",                [("owner_id", 1), ("is_sold", 1)],        {"name": "owner_is_sold_idx"}),
    ("tools",                [("owner_id", 1), ("lost_status.is_lost", 1)], {"name": "owner_lost_idx"}),
    ("tools",                [("owner_id", 1), ("is_sold", 1), ("lost_status.is_lost", 1)], {"name": "owner_active_idx"}),
    ("tools",                [("owner_id", 1), ("warranty_until", 1)], {"name": "owner_warranty_idx"}),
    ("tools",                [("owner_id", 1), ("created_at", 1)],     {"name": "owner_created_idx"}),
    ("tools",                [("owner_id", 1), ("brand", 1)],          {"name": "owner_brand_idx"}),
    ("tools",                [("owner_id", 1), ("checked_out", 1)],    {"name": "owner_checked_out_idx"}),
    ("tools",                [("id", 1)],                              {"unique": True, "name": "tool_id_unique"}),
    ("locations",            [("owner_id", 1)],                        {"name": "owner_id_idx"}),
    ("locations",            [("owner_id", 1), ("parent_id", 1)],      {"name": "owner_parent_idx"}),
    ("dealers",              [("owner_id", 1)],                        {"name": "owner_id_idx"}),
    ("dealers",              [("owner_id", 1), ("name", 1)],           {"name": "owner_name_idx"}),
    ("tags",                 [("owner_id", 1)],                        {"name": "owner_id_idx"}),
    ("categories",           [("owner_id", 1)],                        {"name": "owner_id_idx"}),
    ("borrowers",            [("owner_id", 1)],                        {"name": "owner_id_idx"}),
    ("borrowers",            [("owner_id", 1), ("name", 1)],           {"name": "owner_name_idx"}),
    ("checkouts",            [("owner_id", 1), ("tool_id", 1)],        {"name": "owner_tool_idx"}),
    ("checkouts",            [("owner_id", 1), ("borrower_id", 1)],    {"name": "owner_borrower_idx"}),
    ("checkouts",            [("owner_id", 1), ("returned_at", 1)],    {"name": "owner_returned_idx"}),
    ("wishlist_items",       [("owner_id", 1)],                        {"name": "owner_id_idx"}),
    ("transactions",         [("owner_id", 1), ("dealer_id", 1)],      {"name": "owner_dealer_idx"}),
    ("transactions",         [("owner_id", 1), ("date", -1)],          {"name": "owner_date_idx"}),
    ("warranty_claims",      [("owner_id", 1)],                        {"name": "owner_id_idx"}),
    ("warranty_claims",      [("owner_id", 1), ("status", 1)],         {"name": "owner_status_idx"}),
    ("warranty_claims",      [("owner_id", 1), ("dealer_id", 1)],      {"name": "owner_dealer_idx"}),
    ("warranty_claims",      [("tool_id", 1)],                         {"name": "tool_id_idx"}),
    ("maintenance_logs",     [("owner_id", 1), ("tool_id", 1)],        {"name": "owner_tool_idx"}),
    ("maintenance_logs",     [("owner_id", 1), ("scheduled_for", 1)],  {"name": "owner_scheduled_idx"}),
    ("activity_log",         [("owner_id", 1), ("at", -1)],            {"name": "owner_at_idx"}),
    ("activity_log",         [("owner_id", 1), ("tool_id", 1)],        {"name": "owner_tool_idx"}),
    ("subscriptions",        [("user_id", 1)],                         {"unique": True, "name": "user_id_unique"}),
    ("promo_codes",          [("code", 1)],                            {"unique": True, "name": "code_unique"}),
    ("password_reset_codes", [("email", 1)],                           {"name": "email_idx"}),
    # TTL: auto-delete expired reset codes 24h after they expire
    ("password_reset_codes", [("expires_at", 1)],                      {"name": "expires_ttl", "expireAfterSeconds": 86400}),
    ("feedback",             [("created_at", -1)],                     {"name": "created_at_idx"}),
]


@app.on_event("startup")
async def ensure_mongo_indexes():
    """Create / refresh indexes on startup. Idempotent.

    We swallow IndexOptionsConflict (raised when an index with the same name
    already exists with slightly different options) and log it — that's fine
    in dev, and we don't want to crash the API on boot for that.
    """
    created = 0
    skipped = 0
    errors: list[str] = []
    for coll_name, keys, options in INDEX_PLAN:
        try:
            await real_db[coll_name].create_index(keys, **options)
            created += 1
        except Exception as e:  # pymongo.errors.OperationFailure most often
            msg = str(e)
            # Index already exists with different options — non-fatal in dev.
            if "IndexOptionsConflict" in msg or "already exists with a different name" in msg or "already exists" in msg.lower():
                skipped += 1
                continue
            errors.append(f"{coll_name}.{options.get('name','?')}: {msg[:140]}")
    if errors:
        logger.warning("Mongo index init: %d ok, %d skipped, %d errors -> %s",
                       created, skipped, len(errors), "; ".join(errors[:3]))
    else:
        logger.info("Mongo index init: %d created/verified, %d skipped (already existed)",
                    created, skipped)
