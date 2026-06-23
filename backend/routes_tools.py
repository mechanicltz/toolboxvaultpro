"""Tool routes — CRUD, CSV import/export, list/filter/search.

The largest route group, extracted from server.py (god-file refactor B3).
Registered on the shared api_router via register_tools_routes(); all
dependencies come from core/models/helpers/routes_taxonomy so this module
never imports server (no cycle). The CSV import/export request models
(ImportRow / ImportPayload / export bodies) are defined inline below exactly
as they were in server.py.

NOTE: we intentionally do NOT use `from __future__ import annotations` here —
the inline Pydantic request-body models (ImportPayload, ExportPayload) are
defined locally inside register_tools_routes(), and FastAPI must see the real
class objects (not lazy string annotations) to treat them as request bodies.
"""

import io
import csv
import logging
from datetime import datetime, timezone, timedelta
from typing import List, Dict, Any, Optional

from pydantic import BaseModel, Field
from fastapi import APIRouter, HTTPException, Depends

from core import db, real_db, get_current_user, current_user_id_var
from auth import User
from helpers import build_tool_query, _validate_photo_payload
from routes_taxonomy import _ensure_brand_saved
import media
from models import now_iso, ToolCreate, ToolUpdate, Tool, Document, RepairInfo, Category, Location, Dealer, Tag, WarrantyClaim

logger = logging.getLogger("routes_tools")


def register_tools_routes(api_router: APIRouter) -> None:
    # ---------- Tools ----------

    # Backward-compat shim. The frontend writes model_numbers[] / serial_numbers[]
    # for new builds, but older app builds still send legacy {model, serial_number,
    # set_serials, is_set} fields. This helper folds either shape into a normalized
    # dict that we can persist consistently. Always writes BOTH the new arrays and
    # the legacy single-value mirrors so older app installs keep rendering data.
    def _resolve_model_serial_arrays(
        updates: Dict[str, Any],
        existing: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Mutates `updates` in place to keep model_numbers/serial_numbers in sync
        with the legacy model / serial_number / set_serials fields."""
        existing = existing or {}

        def _clean(arr: Any) -> List[str]:
            if not isinstance(arr, list):
                return []
            return [str(x).strip() for x in arr if x is not None and str(x).strip()]

        has_new_models = "model_numbers" in updates and updates["model_numbers"] is not None
        has_new_serials = "serial_numbers" in updates and updates["serial_numbers"] is not None
        has_legacy_set = "set_serials" in updates and updates["set_serials"] is not None
        has_legacy_serial = "serial_number" in updates
        has_legacy_model = "model" in updates

        # ----- Resolve model_numbers -----
        if has_new_models:
            mns = _clean(updates["model_numbers"])
        elif has_legacy_set or has_legacy_serial or has_legacy_model:
            # Older app sent legacy fields → derive
            candidates: List[str] = []
            if has_legacy_set:
                candidates.extend(_clean(updates.get("set_serials")))
            if has_legacy_serial and updates.get("serial_number"):
                candidates.append(str(updates["serial_number"]).strip())
            if has_legacy_model and updates.get("model"):
                candidates.append(str(updates["model"]).strip())
            # Dedupe preserving order
            seen = set()
            mns = []
            for v in candidates:
                if v and v not in seen:
                    seen.add(v)
                    mns.append(v)
        else:
            mns = None  # not provided in this update at all

        # ----- Resolve serial_numbers -----
        if has_new_serials:
            sns = _clean(updates["serial_numbers"])
        else:
            sns = None  # not provided in this update at all

        # Persist resolved arrays + legacy mirrors so old app builds still render
        if mns is not None:
            updates["model_numbers"] = mns
            # Legacy mirrors derived from model_numbers
            updates["serial_number"] = mns[0] if mns else ""
            updates["set_serials"] = mns
            updates["is_set"] = len(mns) > 1
            # Clear legacy `model` since we no longer accept it separately
            if "model" not in updates:
                updates["model"] = ""

        if sns is not None:
            updates["serial_numbers"] = sns


    @api_router.post("/tools", response_model=Tool)
    async def create_tool(payload: ToolCreate, user: User = Depends(get_current_user)):
        _validate_photo_payload(payload.photos)
        # Free-tier 15-item limit. Pro / lifetime users always pass.
        from subscriptions import enforce_tool_limit  # local import to avoid cycles
        await enforce_tool_limit(real_db, user.id)
        payload_dict = payload.dict()
        # Normalize legacy/new model & serial fields so both shapes survive.
        _resolve_model_serial_arrays(payload_dict)
        # Strip None values so Tool() applies its defaults instead of crashing
        payload_dict = {k: v for k, v in payload_dict.items() if v is not None}
        tool = Tool(**payload_dict)

        # Denormalize the *_name fields from their *_id counterparts so the
        # tool description card has the right names from the very first
        # render. Without this, freshly-created tools showed empty
        # location/dealer/category names until the next edit (bug noticed
        # while fixing the rename-cascade issue, 2026-05-23).
        if tool.location_id and not tool.location_name:
            loc = await db.locations.find_one(
                {"id": tool.location_id}, {"_id": 0, "name": 1}
            )
            tool.location_name = (loc or {}).get("name", "") or ""
        if tool.dealer_id and not tool.dealer_name:
            dl = await db.dealers.find_one(
                {"id": tool.dealer_id}, {"_id": 0, "name": 1}
            )
            tool.dealer_name = (dl or {}).get("name", "") or ""
        if tool.category_id and not tool.category_name:
            cat = await db.categories.find_one(
                {"id": tool.category_id}, {"_id": 0, "name": 1}
            )
            tool.category_name = (cat or {}).get("name", "") or ""

        tool.photos = (await media.offload_list(user.id, tool.photos)) or []
        tool.receipts = (await media.offload_list(user.id, getattr(tool, "receipts", None))) or []
        tool.documents = (await media.offload_list(user.id, getattr(tool, "documents", None))) or []
        # Offload inside-item photos (bundle sub-items) to GridFS too.
        if tool.inside_items:
            for s in tool.inside_items:
                if getattr(s, "photo", None):
                    s.photo = await media.offload_value(user.id, s.photo)
        await db.tools.insert_one(tool.dict())
        # If created already broken, also create a warranty claim mirror with broken_photo
        if tool.needs_repair:
            ri = (tool.repair_info or RepairInfo()).dict() if hasattr(tool.repair_info, "dict") else (tool.repair_info or {})
            if isinstance(ri, dict):
                claim = WarrantyClaim(
                    tool_id=tool.id,
                    tool_name=tool.name,
                    tool_photo=(tool.photos or [None])[0] if tool.photos else None,
                    broken_photo=ri.get("broken_photo") or "",
                    dealer_id=tool.dealer_id,
                    dealer_name=tool.dealer_name or "",
                    repair_company=ri.get("company_notified") or "",
                    contact=ri.get("contact") or "",
                    notified_at=ri.get("notified_at") or "",
                    expected_completion=ri.get("expected_completion") or "",
                    claim_status="broken",
                    notes=ri.get("notes") or "",
                    repair_cost=float(ri.get("repair_cost") or 0),
                )
                await db.warranty_claims.insert_one(claim.dict())
        # Persist any new brand string to the brands collection so future
        # tools see it as a typeahead suggestion (per user 2026-05-27).
        await _ensure_brand_saved(tool.brand)
        return tool


    # ---------------------------------------------------------------------------
    # CSV Import / Export
    # ---------------------------------------------------------------------------

    # Logical field set the import wizard maps to. Keep this list authoritative.
    _IMPORT_FIELDS = [
        {"id": "name", "label": "Name *", "required": True},
        {"id": "brand", "label": "Brand"},
        {"id": "model", "label": "Model"},
        {"id": "serial_number", "label": "Model number"},
        {"id": "quantity", "label": "Quantity"},
        {"id": "cost", "label": "Cost (per unit)"},
        {"id": "msrp_price", "label": "MSRP (per unit, optional)"},
        {"id": "description", "label": "Description / Notes"},
        {"id": "category", "label": "Category (by name)"},
        {"id": "location", "label": "Location (by name)"},
        {"id": "dealer", "label": "Dealer (by name)"},
        {"id": "condition", "label": "Condition"},
        {"id": "purchase_date", "label": "Purchase date (YYYY-MM-DD)"},
        {"id": "warranty_expiry", "label": "Warranty expiry (YYYY-MM-DD)"},
        {"id": "tags", "label": "Tags (comma-separated)"},
    ]


    @api_router.get("/tools/import-fields")
    async def tools_import_fields(user: User = Depends(get_current_user)):
        return {"fields": _IMPORT_FIELDS}


    # Export field registry — id, CSV header label, and how to read the value
    # from a tool dict given pre-resolved {cats, locs, dlrs, tags} lookup maps.
    _EXPORT_FIELDS: List[Dict[str, Any]] = [
        {"id": "name", "label": "Name"},
        {"id": "brand", "label": "Brand"},
        {"id": "model", "label": "Model"},
        {"id": "serial_number", "label": "Model number"},
        {"id": "serial_numbers", "label": "Serial number(s)"},
        {"id": "quantity", "label": "Quantity"},
        {"id": "cost", "label": "Cost"},
        {"id": "msrp_price", "label": "MSRP"},
        {"id": "category", "label": "Category"},
        {"id": "location", "label": "Location"},
        {"id": "dealer", "label": "Dealer"},
        {"id": "tags", "label": "Tags"},
        {"id": "condition", "label": "Condition"},
        {"id": "purchase_date", "label": "Purchase date"},
        {"id": "warranty_expiry", "label": "Warranty expiry"},
        {"id": "description", "label": "Description"},
        {"id": "is_consumable", "label": "Is consumable"},
        {"id": "is_set", "label": "Is set"},
        {"id": "set_serials", "label": "Set serials"},
    ]
    _EXPORT_FIELD_IDS = [f["id"] for f in _EXPORT_FIELDS]


    def _build_export_value(field_id: str, t: Dict[str, Any], lookups: Dict[str, Dict[str, str]]) -> Any:
        if field_id == "name":
            return t.get("name") or ""
        if field_id == "brand":
            return t.get("brand") or ""
        if field_id == "model":
            return t.get("model") or ""
        if field_id == "serial_number":
            # Legacy "Model number" export. Prefer the new model_numbers[] array
            # (semicolon-joined) so users see all model numbers in one cell;
            # fall back to the single legacy field for tools not yet migrated.
            mns = t.get("model_numbers") or []
            if mns:
                return "; ".join([str(x) for x in mns if x])
            return t.get("serial_number") or ""
        if field_id == "serial_numbers":
            return "; ".join([str(x) for x in (t.get("serial_numbers") or []) if x])
        if field_id == "quantity":
            return t.get("quantity") or 1
        if field_id == "cost":
            return t.get("cost") or 0
        if field_id == "msrp_price":
            return t.get("msrp_price") or 0
        if field_id == "category":
            return lookups["cats"].get(t.get("category_id") or "", "")
        if field_id == "location":
            return lookups["locs"].get(t.get("location_id") or "", "")
        if field_id == "dealer":
            return lookups["dlrs"].get(t.get("dealer_id") or "", "")
        if field_id == "tags":
            return ", ".join(
                sorted([
                    lookups["tags"].get(tid, "")
                    for tid in (t.get("tag_ids") or [])
                    if lookups["tags"].get(tid)
                ])
            )
        if field_id == "condition":
            return t.get("condition") or ""
        if field_id == "purchase_date":
            return t.get("purchase_date") or ""
        if field_id == "warranty_expiry":
            return t.get("warranty_expiry") or ""
        if field_id == "description":
            return t.get("description") or ""
        if field_id == "is_consumable":
            return "yes" if t.get("is_consumable") else ""
        if field_id == "is_set":
            return "yes" if t.get("is_set") else ""
        if field_id == "set_serials":
            return "; ".join(t.get("set_serials") or [])
        return ""


    @api_router.get("/tools/export-fields")
    async def tools_export_fields(user: User = Depends(get_current_user)):
        """List of fields the user can choose from when exporting."""
        return {"fields": _EXPORT_FIELDS}


    class ExportPayload(BaseModel):
        fields: Optional[List[str]] = None  # subset of _EXPORT_FIELD_IDS; None/empty = all
        format: Optional[str] = "csv"        # "csv" | "xlsx"


    @api_router.post("/tools/export-csv")
    async def tools_export_csv_post(
        payload: ExportPayload,
        user: User = Depends(get_current_user),
    ):
        """Field-customisable export. POST body:
           {fields: ["name","brand",...], format: "csv" | "xlsx"}
        Falls through to all fields when `fields` is empty/missing."""
        requested = payload.fields or []
        if requested:
            chosen = [fid for fid in requested if fid in _EXPORT_FIELD_IDS]
        else:
            chosen = list(_EXPORT_FIELD_IDS)
        if not chosen:
            chosen = list(_EXPORT_FIELD_IDS)
        fmt = (payload.format or "csv").lower().strip()
        if fmt not in ("csv", "xlsx"):
            fmt = "csv"

        return await _do_export(chosen, fmt)


    @api_router.get("/tools/export-csv")
    async def tools_export_csv(user: User = Depends(get_current_user)):
        """Backwards-compat GET — exports all fields as CSV."""
        return await _do_export(list(_EXPORT_FIELD_IDS), "csv")


    async def _do_export(chosen_field_ids: List[str], fmt: str) -> Dict[str, Any]:
        """Shared implementation used by POST and legacy GET. `fmt` in ("csv","xlsx")."""
        tools = await db.tools.find({}, {"_id": 0}).sort("name", 1).to_list(20000)
        cat_ids = {t.get("category_id") for t in tools if t.get("category_id")}
        cats = {
            c["id"]: c.get("name") or ""
            for c in await db.categories.find(
                {"id": {"$in": list(cat_ids)}}, {"_id": 0, "id": 1, "name": 1}
            ).to_list(5000)
        } if cat_ids else {}
        loc_ids = {t.get("location_id") for t in tools if t.get("location_id")}
        locs = {
            lc["id"]: lc.get("name") or ""
            for lc in await db.locations.find(
                {"id": {"$in": list(loc_ids)}}, {"_id": 0, "id": 1, "name": 1}
            ).to_list(5000)
        } if loc_ids else {}
        dlr_ids = {t.get("dealer_id") for t in tools if t.get("dealer_id")}
        dlrs = {
            d["id"]: d.get("name") or ""
            for d in await db.dealers.find(
                {"id": {"$in": list(dlr_ids)}}, {"_id": 0, "id": 1, "name": 1}
            ).to_list(5000)
        } if dlr_ids else {}
        all_tag_ids: List[str] = []
        for t in tools:
            all_tag_ids.extend(t.get("tag_ids") or [])
        uniq_tag_ids = list(set(all_tag_ids))
        tags = {
            tg["id"]: tg.get("name") or ""
            for tg in await db.tags.find(
                {"id": {"$in": uniq_tag_ids}}, {"_id": 0, "id": 1, "name": 1}
            ).to_list(5000)
        } if uniq_tag_ids else {}

        lookups = {"cats": cats, "locs": locs, "dlrs": dlrs, "tags": tags}
        label_for = {f["id"]: f["label"] for f in _EXPORT_FIELDS}
        headers = [label_for[fid] for fid in chosen_field_ids]
        rows_data = [
            [_build_export_value(fid, t, lookups) for fid in chosen_field_ids]
            for t in tools
        ]

        today = datetime.utcnow().strftime("%Y-%m-%d")
        import base64 as _b64

        if fmt == "xlsx":
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except Exception as e:
                raise HTTPException(500, f"openpyxl not available: {e}")
            import io as _io
            wb = Workbook()
            ws = wb.active
            ws.title = "Tools"
            ws.append(headers)
            header_font = Font(bold=True, color="000000")
            header_fill = PatternFill("solid", fgColor="FFB300")
            for col_i in range(1, len(headers) + 1):
                c = ws.cell(row=1, column=col_i)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(vertical="center")
            for row in rows_data:
                ws.append(row)
            # Freeze header row + auto-size columns based on content
            ws.freeze_panes = "A2"
            for col_i, header in enumerate(headers, start=1):
                max_len = len(str(header))
                for row in rows_data:
                    v = row[col_i - 1]
                    vlen = len(str(v)) if v is not None else 0
                    if vlen > max_len:
                        max_len = vlen
                # Cap width so sheet stays readable
                ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = min(
                    max(12, max_len + 2), 50
                )
            buf = _io.BytesIO()
            wb.save(buf)
            raw = buf.getvalue()
            return {
                "filename": f"toolbox-vault-export-{today}.xlsx",
                "base64": _b64.b64encode(raw).decode("ascii"),
                "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "rows": len(tools),
                "fields": chosen_field_ids,
                "format": "xlsx",
            }

        # Default: CSV
        import csv as _csv
        import io as _io
        buf = _io.StringIO()
        w = _csv.writer(buf)
        w.writerow(headers)
        for row in rows_data:
            w.writerow(row)
        raw = buf.getvalue().encode("utf-8")
        return {
            "filename": f"toolbox-vault-export-{today}.csv",
            "base64": _b64.b64encode(raw).decode("ascii"),
            "mime": "text/csv",
            "rows": len(tools),
            "fields": chosen_field_ids,
            "format": "csv",
        }


    class ImportRow(BaseModel):
        name: Optional[str] = ""
        brand: Optional[str] = ""
        model: Optional[str] = ""
        serial_number: Optional[str] = ""
        # NOTE: quantity / cost accept Any so we can tolerate values from
        # third-party CSVs like "13,500.00", "$1,200", "1.0", "1 ea", "" etc.
        # The raw value is sanitised inside tools_import() via _to_int / _to_float.
        quantity: Optional[Any] = 1
        cost: Optional[Any] = 0.0
        msrp_price: Optional[Any] = 0.0
        description: Optional[str] = ""
        category: Optional[str] = ""        # name (case-insensitive lookup; auto-create if missing)
        location: Optional[str] = ""        # name match (existing only)
        dealer: Optional[str] = ""          # name match (existing only)
        condition: Optional[str] = ""
        purchase_date: Optional[str] = ""
        warranty_expiry: Optional[str] = ""
        tags: Optional[str] = ""            # comma-separated names; auto-create if missing


    class ImportPayload(BaseModel):
        rows: List[ImportRow]
        create_missing_categories: bool = True
        create_missing_tags: bool = True
        create_missing_locations: bool = True
        create_missing_dealers: bool = True


    def _norm(s: Optional[str]) -> str:
        return (s or "").strip() if isinstance(s, str) else (str(s).strip() if s is not None else "")


    def _to_float(v: Any) -> float:
        """Tolerant float parser — handles strings with currency symbols,
        thousand separators, percent signs, blanks, etc.
        Returns 0.0 if the value cannot be coerced."""
        if v is None or v == "":
            return 0.0
        if isinstance(v, (int, float)):
            try:
                return float(v)
            except Exception:
                return 0.0
        s = str(v).strip()
        if not s:
            return 0.0
        # Strip everything except digits, dot, minus, comma — then drop commas.
        # Handles "$13,500.00", "13.500,00 €" (best-effort), "1,234", "1.0"
        cleaned = "".join(ch for ch in s if ch.isdigit() or ch in ".-,")
        # If both ',' and '.' present, assume ',' is thousand-sep (US/UK format)
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(",", "")
        elif "," in cleaned and "." not in cleaned:
            # Could be European decimal (e.g. "13,50") OR a thousands-sep ("13,500").
            # Heuristic: if there are exactly 3 digits after the comma and no other
            # comma, treat as thousands (e.g. 13,500 → 13500). Otherwise treat
            # the comma as a decimal point.
            parts = cleaned.split(",")
            if len(parts) == 2 and len(parts[1]) == 3:
                cleaned = cleaned.replace(",", "")
            else:
                cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned) if cleaned not in ("", "-", ".") else 0.0
        except Exception:
            return 0.0


    def _to_int(v: Any, default: int = 1) -> int:
        """Tolerant int parser. Falls back to `default` on failure (and clamps to >=1)."""
        f = _to_float(v)
        try:
            n = int(f)
        except Exception:
            n = default
        return max(1, n) if default >= 1 else n


    def _norm_lower(s: Optional[str]) -> str:
        return (s or "").strip().lower()


    @api_router.post("/tools/import")
    async def tools_import(payload: ImportPayload, user: User = Depends(get_current_user)):
        """Bulk-create tools from a normalised list of rows. The frontend is
        responsible for column mapping; this endpoint just creates tools and
        resolves FK names → ids (with optional auto-create for categories/tags).
        """
        # Free-tier 15-item limit applied once for the whole batch.
        from subscriptions import enforce_tool_limit  # local import to avoid cycles
        rows_count = len(payload.rows or [])
        if rows_count > 0:
            await enforce_tool_limit(real_db, user.id, additional=rows_count)
        # Pre-load all the lookup collections once.
        cats = await db.categories.find({}, {"_id": 0}).to_list(5000)
        cats_by_name = {(_norm_lower(c.get("name"))): c for c in cats if c.get("name")}
        tags = await db.tags.find({}, {"_id": 0}).to_list(5000)
        tags_by_name = {(_norm_lower(t.get("name"))): t for t in tags if t.get("name")}
        locs = await db.locations.find({}, {"_id": 0}).to_list(5000)
        locs_by_name = {(_norm_lower(lc.get("name"))): lc for lc in locs if lc.get("name")}
        dlrs = await db.dealers.find({}, {"_id": 0}).to_list(5000)
        dlrs_by_name = {(_norm_lower(d.get("name"))): d for d in dlrs if d.get("name")}

        created: List[Dict[str, Any]] = []
        errors: List[Dict[str, Any]] = []
        auto_created = {
            "categories": [],  # list of {id, name}
            "tags": [],
            "locations": [],
            "dealers": [],
        }

        for idx, raw in enumerate(payload.rows):
            try:
                name = _norm(raw.name)
                if not name:
                    raise ValueError("Name is required")

                # Category — auto-create if missing
                category_id = None
                category_name = ""
                cname = _norm(raw.category)
                if cname:
                    key = cname.lower()
                    if key in cats_by_name:
                        c = cats_by_name[key]
                        category_id = c.get("id")
                        category_name = c.get("name") or cname
                    elif payload.create_missing_categories:
                        new_cat = Category(name=cname)
                        await db.categories.insert_one(new_cat.dict())
                        cats_by_name[key] = new_cat.dict()
                        category_id = new_cat.id
                        category_name = new_cat.name
                        auto_created["categories"].append({"id": new_cat.id, "name": new_cat.name})

                # Location — match existing (case-insensitive) or auto-create
                location_id = None
                location_name = ""
                lname = _norm(raw.location)
                if lname:
                    key = lname.lower()
                    if key in locs_by_name:
                        lc = locs_by_name[key]
                        location_id = lc.get("id")
                        location_name = lc.get("name") or lname
                    elif payload.create_missing_locations:
                        new_loc = Location(name=lname)
                        await db.locations.insert_one(new_loc.dict())
                        locs_by_name[key] = new_loc.dict()
                        location_id = new_loc.id
                        location_name = new_loc.name
                        auto_created["locations"].append({"id": new_loc.id, "name": new_loc.name})

                # Dealer — match existing (case-insensitive) or auto-create
                dealer_id = None
                dealer_name = ""
                dname = _norm(raw.dealer)
                if dname:
                    key = dname.lower()
                    if key in dlrs_by_name:
                        d = dlrs_by_name[key]
                        dealer_id = d.get("id")
                        dealer_name = d.get("name") or dname
                    elif payload.create_missing_dealers:
                        new_dlr = Dealer(name=dname)
                        await db.dealers.insert_one(new_dlr.dict())
                        dlrs_by_name[key] = new_dlr.dict()
                        dealer_id = new_dlr.id
                        dealer_name = new_dlr.name
                        auto_created["dealers"].append({"id": new_dlr.id, "name": new_dlr.name})

                # Tags — comma-separated; auto-create if missing
                tag_ids: List[str] = []
                tag_names: List[str] = []
                if raw.tags:
                    for piece in (raw.tags or "").split(","):
                        tname = _norm(piece)
                        if not tname:
                            continue
                        key = tname.lower()
                        if key in tags_by_name:
                            tg = tags_by_name[key]
                            tag_ids.append(tg.get("id"))
                            tag_names.append(tg.get("name") or tname)
                        elif payload.create_missing_tags:
                            new_tag = Tag(name=tname)
                            await db.tags.insert_one(new_tag.dict())
                            tags_by_name[key] = new_tag.dict()
                            tag_ids.append(new_tag.id)
                            tag_names.append(new_tag.name)
                            auto_created["tags"].append({"id": new_tag.id, "name": new_tag.name})

                qty = _to_int(raw.quantity, default=1)
                cost = _to_float(raw.cost)
                msrp_price = _to_float(raw.msrp_price)

                # Build model_numbers[] from legacy import fields (deduped).
                _mn_cands = []
                if _norm(raw.serial_number):
                    _mn_cands.append(_norm(raw.serial_number))
                if _norm(raw.model):
                    _mn_cands.append(_norm(raw.model))
                _mn_seen = set()
                _model_numbers: List[str] = []
                for _v in _mn_cands:
                    if _v and _v not in _mn_seen:
                        _mn_seen.add(_v)
                        _model_numbers.append(_v)

                tool = Tool(
                    name=name,
                    brand=_norm(raw.brand),
                    model=_norm(raw.model),
                    serial_number=_norm(raw.serial_number),
                    model_numbers=_model_numbers,
                    serial_numbers=[],
                    quantity=qty,
                    cost=cost,
                    msrp_price=msrp_price,
                    description=_norm(raw.description),
                    category_id=category_id,
                    category_name=category_name,
                    location_id=location_id,
                    location_name=location_name,
                    dealer_id=dealer_id,
                    dealer_name=dealer_name,
                    tag_ids=tag_ids,
                    tag_names=tag_names,
                    condition=_norm(raw.condition),
                    purchase_date=_norm(raw.purchase_date) or None,
                    warranty_expiry=_norm(raw.warranty_expiry) or None,
                )
                await db.tools.insert_one(tool.dict())
                created.append({"id": tool.id, "name": tool.name})
            except Exception as e:
                errors.append({"row": idx + 1, "name": _norm(raw.name), "error": str(e)})

        return {
            "created": len(created),
            "errors": errors,
            "ids": [c["id"] for c in created],
            "auto_created": auto_created,
        }


    # ---------------------------------------------------------------------------
    # Photo size cap — prevent oversized base64 payloads from bloating the DB.
    # Each photo gets its own ~5MB cap (well above what camera+compress yields)
    # and the total per-tool photo payload is capped at 25MB so a single tool
    # can have multiple photos but never balloon the document.
    # ---------------------------------------------------------------------------


    @api_router.get("/tools", response_model=List[Tool])
    async def list_tools(
        search: Optional[str] = None,
        location_id: Optional[str] = None,
        tag_id: Optional[str] = None,
        category_id: Optional[str] = None,
        dealer_id: Optional[str] = None,
        checked_out: Optional[bool] = None,
        is_consumable: Optional[bool] = None,
        needs_repair: Optional[bool] = None,
        for_sale: Optional[bool] = None,
        is_sold: Optional[bool] = None,
        is_bundle: Optional[bool] = None,
        expansion_of: Optional[str] = None,
    ):
        """List tools — returns a slim payload for fast list rendering.

        To keep the inventory list snappy on phones, this endpoint:
          - Returns ONLY the first photo (the cover), not all photos
          - Strips `documents` (heavy base64 PDFs/images attached to the tool)
          - Strips `receipts` (heavy base64 receipt images)
        The full set of photos / documents / receipts is only loaded when the
        user opens a tool's detail page (GET /tools/{id}).
        """
        query = build_tool_query(
            search, location_id, tag_id, category_id, dealer_id,
            checked_out, is_consumable, needs_repair, for_sale, is_sold,
            is_bundle, expansion_of,
        )
        items = await db.tools.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
        out: List[Tool] = []
        for i in items:
            # Slim payload: keep only first photo, drop documents & receipts
            photos = i.get("photos") or []
            i["photos"] = [media.thumb_url(photos[0])] if photos else []
            i["documents"] = []
            i["receipts"] = []
            # Slim inside items: keep name/model/cost + a thumb of the photo.
            ins = i.get("inside_items") or []
            for s in ins:
                if isinstance(s, dict) and s.get("photo"):
                    s["photo"] = media.thumb_url(s["photo"])
            out.append(Tool(**i))
        return out


    @api_router.get("/tools/{tool_id}", response_model=Tool)
    async def get_tool(tool_id: str):
        doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Tool not found")
        return Tool(**doc)


    @api_router.put("/tools/{tool_id}", response_model=Tool)
    async def update_tool(tool_id: str, payload: ToolUpdate):
        if payload.photos is not None:
            _validate_photo_payload(payload.photos)
        doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Tool not found")

        # Use exclude_unset so the client can EXPLICITLY clear a field by
        # sending null (e.g., `{"location_id": null}` to detach the tool from
        # its location). The previous "drop if v is None" rule made the
        # endpoint physically unable to clear a foreign-key field — values
        # could only be reassigned, never removed.
        updates = payload.dict(exclude_unset=True)
        updates["updated_at"] = now_iso()

        # Offload any incoming base64 photos to GridFS (-> /api/files URLs).
        _owner = doc.get("owner_id") or ""
        for _f in ("photos", "receipts", "documents"):
            if _f in updates:
                updates[_f] = (await media.offload_list(_owner, updates[_f])) or []
        # Offload inside-item photos when the bundle's sub-items are edited.
        if "inside_items" in updates and updates["inside_items"]:
            for s in updates["inside_items"]:
                if isinstance(s, dict) and s.get("photo"):
                    s["photo"] = await media.offload_value(_owner, s["photo"])

        # Keep legacy model/serial fields and new model_numbers/serial_numbers
        # arrays in sync — whichever shape the client sent, both get persisted.
        _resolve_model_serial_arrays(updates, doc)

        # ---------------------------------------------------------------
        # Keep denormalized *_name fields in sync with their *_id fields.
        # The frontend only sends *_id when re-assigning location/dealer/
        # category, but the tool model also stores a cached *_name field
        # (used by list views to avoid an extra join). If we update the id
        # without refreshing the name, the description card keeps showing
        # the OLD name forever — exactly the "renamed to test2 but card
        # still shows test" bug reported 2026-05-23.
        # ---------------------------------------------------------------
        if "location_id" in updates:
            if updates["location_id"]:
                loc = await db.locations.find_one(
                    {"id": updates["location_id"]}, {"_id": 0, "name": 1}
                )
                updates["location_name"] = (loc or {}).get("name", "") or ""
            else:
                updates["location_name"] = ""

        if "dealer_id" in updates:
            if updates["dealer_id"]:
                dl = await db.dealers.find_one(
                    {"id": updates["dealer_id"]}, {"_id": 0, "name": 1}
                )
                updates["dealer_name"] = (dl or {}).get("name", "") or ""
            else:
                updates["dealer_name"] = ""

        if "category_id" in updates:
            if updates["category_id"]:
                cat = await db.categories.find_one(
                    {"id": updates["category_id"]}, {"_id": 0, "name": 1}
                )
                updates["category_name"] = (cat or {}).get("name", "") or ""
            else:
                updates["category_name"] = ""

        # When the caller marks the tool as Repaired (needs_repair: false, repair_info: null),
        # Pydantic's None values get filtered out above. Restore the null so the tool's
        # repair_info (including broken_photo) is actually cleared — otherwise the next claim
        # would inherit the previous claim's photo and notes.
        if payload.needs_repair is False and doc.get("needs_repair"):
            updates["repair_info"] = None

        # Auto-checkin when a tool is being newly flagged as broken / needing repair
        becomes_broken = (
            updates.get("needs_repair") is True
            and not doc.get("needs_repair")
            and doc.get("is_checked_out")
        )
        if becomes_broken:
            record = doc.get("current_checkout") or {}
            if record:
                record = dict(record)
                record["checked_in_at"] = now_iso()
                note_extra = " [auto check-in: marked for repair]"
                record["notes"] = (record.get("notes") or "") + note_extra
                history = doc.get("checkout_history") or []
                history.append(record)
                updates["is_checked_out"] = False
                updates["current_checkout"] = None
                updates["checkout_history"] = history

        await db.tools.update_one({"id": tool_id}, {"$set": updates})
        new_doc = await db.tools.find_one({"id": tool_id}, {"_id": 0})

        # Sync warranty claim record:
        # Newly broken → create a new open claim (if no open claim already exists for this tool)
        just_flagged = updates.get("needs_repair") is True and not doc.get("needs_repair")
        just_unflagged = updates.get("needs_repair") is False and doc.get("needs_repair")
        if just_flagged:
            existing_open = await db.warranty_claims.find_one(
                {"tool_id": tool_id, "claim_status": {"$nin": ["completed", "rejected"]}},
                {"_id": 0, "id": 1},
            )
            if not existing_open:
                ri = (new_doc.get("repair_info") or {})
                _bundle_model = ""
                if new_doc.get("is_bundle"):
                    _mns = new_doc.get("model_numbers") or []
                    _bundle_model = (_mns[0] if _mns else "") or new_doc.get("model") or ""
                claim = WarrantyClaim(
                    tool_id=tool_id,
                    tool_name=new_doc.get("name", ""),
                    tool_photo=(new_doc.get("photos") or [None])[0],
                    inside_item_id=ri.get("inside_item_id") or None,
                    inside_item_name=ri.get("inside_item_name") or "",
                    inside_item_model=ri.get("inside_item_model") or "",
                    bundle_model=_bundle_model,
                    broken_photo=ri.get("broken_photo") or "",
                    dealer_id=new_doc.get("dealer_id"),
                    dealer_name=new_doc.get("dealer_name") or "",
                    repair_company=ri.get("company_notified") or "",
                    contact=ri.get("contact") or "",
                    notified_at=ri.get("notified_at") or "",
                    expected_completion=ri.get("expected_completion") or "",
                    claim_status="broken",
                    notes=ri.get("notes") or "",
                    repair_cost=float(ri.get("repair_cost") or 0),
                )
                await db.warranty_claims.insert_one(claim.dict())
        elif "repair_info" in updates and new_doc.get("needs_repair"):
            # Repair info edited while still broken → keep open claim in sync
            ri = updates.get("repair_info") or {}
            await db.warranty_claims.update_many(
                {"tool_id": tool_id, "claim_status": {"$nin": ["completed", "rejected"]}},
                {
                    "$set": {
                        "repair_company": ri.get("company_notified") or "",
                        "contact": ri.get("contact") or "",
                        "notified_at": ri.get("notified_at") or "",
                        "expected_completion": ri.get("expected_completion") or "",
                        "notes": ri.get("notes") or "",
                        "broken_photo": ri.get("broken_photo") or "",
                        "repair_cost": float(ri.get("repair_cost") or 0),
                        "updated_at": now_iso(),
                    }
                },
            )
        if just_unflagged:
            # User hit "Mark Repaired" — close any still-open claim as completed
            await db.warranty_claims.update_many(
                {"tool_id": tool_id, "claim_status": {"$nin": ["completed", "rejected"]}},
                {
                    "$set": {
                        "claim_status": "completed",
                        "completed_at": now_iso(),
                        "updated_at": now_iso(),
                    }
                },
            )

        # Persist any new brand string to the brands collection so future
        # tools see it as a typeahead suggestion (per user 2026-05-27).
        await _ensure_brand_saved(new_doc.get("brand"))
        return Tool(**new_doc)


    @api_router.delete("/tools/{tool_id}")
    async def delete_tool(tool_id: str):
        _doc = await db.tools.find_one({"id": tool_id}, {"_id": 0, "photos": 1, "receipts": 1, "documents": 1})
        res = await db.tools.delete_one({"id": tool_id})
        if res.deleted_count == 0:
            raise HTTPException(404, "Tool not found")
        if _doc:
            await media.delete_values(_doc.get("photos"))
            await media.delete_values(_doc.get("receipts"))
            await media.delete_values(_doc.get("documents"))
        # Cascade: also remove any warranty claims that referenced this tool —
        # otherwise the dealer-claims summary keeps counting orphaned claims
        # but the detail screen can't resolve them back to a tool.
        await db.warranty_claims.delete_many({"tool_id": tool_id})
        return {"ok": True}


    # Bundles / Sets -> routes_bundles.py (god-file refactor B3).
    from routes_bundles import register_bundle_routes  # noqa: E402
    register_bundle_routes(api_router)
