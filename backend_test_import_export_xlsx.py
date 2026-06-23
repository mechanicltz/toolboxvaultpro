"""
Import/Export Database — XLSX format support + CSV legacy + Import tests.
Run: python3 /app/backend_test_import_export_xlsx.py
"""
import base64
import io
import sys
import requests

BASE = "https://toolbox-vault-v3.preview.emergentagent.com/api"
EMAIL = "subtest@example.com"
PASSWORD = "password123"

PASS = 0
FAIL = 0


def log(ok, msg):
    global PASS, FAIL
    if ok:
        PASS += 1
        print(f"  PASS  {msg}")
    else:
        FAIL += 1
        print(f"  FAIL  {msg}")


def login():
    r = requests.post(f"{BASE}/auth/login", json={"email": EMAIL, "password": PASSWORD}, timeout=30)
    r.raise_for_status()
    return r.json()["token"]


def main():
    token = login()
    H = {"Authorization": f"Bearer {token}"}

    # ---------------- 1. /tools/export-fields ----------------
    print("\n[1] GET /api/tools/export-fields")
    r = requests.get(f"{BASE}/tools/export-fields", headers=H, timeout=30)
    log(r.status_code == 200, f"status=200 (got {r.status_code})")
    body = r.json()
    fields = body.get("fields", [])
    ids = [f.get("id") for f in fields]
    expected = ["name","brand","model","serial_number","quantity","cost","category",
                "location","dealer","tags","condition","purchase_date","warranty_expiry",
                "description","is_consumable","is_set","set_serials"]
    log(len(fields) == 17, f"17 fields (got {len(fields)})")
    for fid in expected:
        log(fid in ids, f"contains field id '{fid}'")
    # each field has id + label
    log(all("id" in f and "label" in f for f in fields), "each field has id + label keys")

    # Create a tool first so export has data
    print("\n[Setup] Create a pre-existing tool for export")
    r = requests.post(f"{BASE}/tools", headers=H, json={"name": "Pre-existing Tool", "cost": 5}, timeout=30)
    log(r.status_code == 200, f"POST /tools setup status=200 (got {r.status_code})")
    setup_tool_id = r.json().get("id") if r.status_code == 200 else None

    try:
        # ---------------- 2a. CSV export ----------------
        print("\n[2a] POST /api/tools/export-csv  (fields=['name','brand','cost'], format='csv')")
        r = requests.post(f"{BASE}/tools/export-csv", headers=H,
                          json={"fields": ["name", "brand", "cost"], "format": "csv"}, timeout=30)
        log(r.status_code == 200, f"status=200 (got {r.status_code})")
        b = r.json()
        log(b.get("format") == "csv", f"response.format == 'csv' (got {b.get('format')})")
        fname = b.get("filename", "")
        log(fname.endswith(".csv"), f"filename ends with .csv (got {fname})")
        log(b.get("mime") == "text/csv", f"mime=='text/csv' (got {b.get('mime')})")
        log(b.get("fields") == ["name", "brand", "cost"], f"fields round-trip (got {b.get('fields')})")
        log(isinstance(b.get("rows"), int), f"rows is int (got {type(b.get('rows')).__name__}={b.get('rows')})")
        raw = base64.b64decode(b.get("base64", ""))
        txt = raw.decode("utf-8")
        first_line = txt.splitlines()[0] if txt else ""
        log(first_line == "Name,Brand,Cost", f"CSV header row == 'Name,Brand,Cost' (got {first_line!r})")

        # ---------------- 2b. XLSX export ----------------
        print("\n[2b] POST /api/tools/export-csv  (fields=['name','brand','cost'], format='xlsx')")
        r = requests.post(f"{BASE}/tools/export-csv", headers=H,
                          json={"fields": ["name", "brand", "cost"], "format": "xlsx"}, timeout=30)
        log(r.status_code == 200, f"status=200 (got {r.status_code})")
        b = r.json()
        log(b.get("format") == "xlsx", f"response.format == 'xlsx' (got {b.get('format')})")
        fname = b.get("filename", "")
        log(fname.endswith(".xlsx"), f"filename ends with .xlsx (got {fname})")
        log(b.get("mime") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"mime=xlsx (got {b.get('mime')})")
        log(b.get("fields") == ["name", "brand", "cost"], f"fields round-trip (got {b.get('fields')})")
        log(isinstance(b.get("rows"), int), "rows is int")
        raw = base64.b64decode(b.get("base64", ""))
        log(raw[:4] == b"PK\x03\x04", f"base64 decodes to bytes starting with PK zip signature (got {raw[:4]!r})")

        # Verify XLSX can be opened with openpyxl
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw))
            ws = wb.active
            row1 = [c.value for c in ws[1]]
            log(row1 == ["Name", "Brand", "Cost"],
                f"openpyxl row 1 == ['Name','Brand','Cost'] (got {row1!r})")
        except Exception as e:
            log(False, f"openpyxl load_workbook failed: {e}")

        # ---------------- 2c. empty fields with xlsx -> all 17 + xlsx ----------------
        print("\n[2c] POST /api/tools/export-csv  (fields=[], format='xlsx')")
        r = requests.post(f"{BASE}/tools/export-csv", headers=H,
                          json={"fields": [], "format": "xlsx"}, timeout=30)
        log(r.status_code == 200, f"status=200 (got {r.status_code})")
        b = r.json()
        log(b.get("format") == "xlsx", f"format=='xlsx' (got {b.get('format')})")
        log(len(b.get("fields", [])) == 17, f"fields fallback length=17 (got {len(b.get('fields', []))})")
        fname = b.get("filename", "")
        log(fname.endswith(".xlsx"), f"filename ends with .xlsx (got {fname})")
        raw = base64.b64decode(b.get("base64", ""))
        log(raw[:4] == b"PK\x03\x04", "valid xlsx zip signature")
        try:
            import openpyxl
            openpyxl.load_workbook(io.BytesIO(raw))
            log(True, "openpyxl can open the fallback xlsx")
        except Exception as e:
            log(False, f"openpyxl open failed: {e}")

        # ---------------- 2d. garbage format -> csv fallback ----------------
        print("\n[2d] POST /api/tools/export-csv  (fields=['name'], format='garbage')")
        r = requests.post(f"{BASE}/tools/export-csv", headers=H,
                          json={"fields": ["name"], "format": "garbage"}, timeout=30)
        log(r.status_code == 200, f"status=200 (got {r.status_code})")
        b = r.json()
        log(b.get("format") == "csv", f"garbage -> csv fallback (got format={b.get('format')})")
        fname = b.get("filename", "")
        log(fname.endswith(".csv"), f"filename ends with .csv (got {fname})")

        # ---------------- 2e. no-auth -> 401 ----------------
        print("\n[2e] POST /api/tools/export-csv without Authorization header")
        r = requests.post(f"{BASE}/tools/export-csv",
                          json={"fields": ["name"], "format": "csv"}, timeout=30)
        log(r.status_code == 401, f"status=401 (got {r.status_code})")

    finally:
        # cleanup setup tool
        if setup_tool_id:
            requests.delete(f"{BASE}/tools/{setup_tool_id}", headers=H, timeout=30)

    # ---------------- 3. /tools/import ----------------
    print("\n[3] POST /api/tools/import")
    # pre-check existing state for the test category/tags
    pre_cat_names = set()
    pre_tag_names = set()
    r = requests.get(f"{BASE}/categories", headers=H, timeout=30)
    if r.status_code == 200:
        pre_cat_names = {c.get("name") for c in r.json()}
    r = requests.get(f"{BASE}/tags", headers=H, timeout=30)
    if r.status_code == 200:
        pre_tag_names = {t.get("name") for t in r.json()}

    test_cat = "TestCategoryIE"
    cat_existed = test_cat in pre_cat_names
    tag_red_existed = "red" in pre_tag_names
    tag_blue_existed = "blue" in pre_tag_names

    payload = {
        "rows": [
            {"name": "IE Test Widget", "brand": "Acme", "cost": "12.50"},
            {"name": "IE Test Screwdriver", "brand": "Acme", "category": test_cat, "tags": "red,blue"},
        ],
        "create_missing_categories": True,
        "create_missing_tags": True,
    }
    r = requests.post(f"{BASE}/tools/import", headers=H, json=payload, timeout=30)
    log(r.status_code == 200, f"status=200 (got {r.status_code})")
    body = r.json()
    created = body.get("created", 0)
    log(created >= 2, f"created >= 2 (got {created})")
    imported_ids = body.get("ids", [])

    # Verify by GET /tools and find by name
    r = requests.get(f"{BASE}/tools", headers=H, timeout=30)
    tools = r.json() if r.status_code == 200 else []
    widget = next((t for t in tools if t.get("name") == "IE Test Widget"), None)
    screw = next((t for t in tools if t.get("name") == "IE Test Screwdriver"), None)
    log(widget is not None, "IE Test Widget found via GET /tools")
    log(screw is not None, "IE Test Screwdriver found via GET /tools")
    if widget:
        log(widget.get("brand") == "Acme", f"widget.brand=='Acme' (got {widget.get('brand')})")
        log(abs(float(widget.get("cost") or 0) - 12.50) < 0.01,
            f"widget.cost==12.50 (got {widget.get('cost')})")
    if screw:
        log(screw.get("category_name") == test_cat,
            f"screwdriver.category_name=='{test_cat}' (got {screw.get('category_name')})")
        tag_names = screw.get("tag_names") or []
        log("red" in tag_names and "blue" in tag_names,
            f"screwdriver.tag_names has 'red' and 'blue' (got {tag_names})")

    # Verify categories endpoint has the auto-created category
    r = requests.get(f"{BASE}/categories", headers=H, timeout=30)
    cat_names = {c.get("name") for c in r.json()} if r.status_code == 200 else set()
    log(test_cat in cat_names, f"TestCategoryIE exists in GET /categories")
    cat_id = None
    if r.status_code == 200:
        for c in r.json():
            if c.get("name") == test_cat:
                cat_id = c.get("id")
                break

    # Verify tags endpoint has red + blue
    r = requests.get(f"{BASE}/tags", headers=H, timeout=30)
    tag_list = r.json() if r.status_code == 200 else []
    tag_names_all = {t.get("name") for t in tag_list}
    log("red" in tag_names_all, "tag 'red' exists in GET /tags")
    log("blue" in tag_names_all, "tag 'blue' exists in GET /tags")
    red_id = next((t.get("id") for t in tag_list if t.get("name") == "red"), None)
    blue_id = next((t.get("id") for t in tag_list if t.get("name") == "blue"), None)

    # Edge case: empty name must go to errors[]
    print("\n[3b] POST /api/tools/import  (empty name)")
    r = requests.post(f"{BASE}/tools/import", headers=H,
                      json={"rows": [{"name": "", "brand": "Foo"}]}, timeout=30)
    log(r.status_code == 200, f"status=200 (got {r.status_code})")
    b = r.json()
    errors = b.get("errors", [])
    log(b.get("created", -1) == 0, f"created==0 for empty-name row (got {b.get('created')})")
    log(len(errors) >= 1, f"errors[] has >=1 entry (got {len(errors)})")
    if errors:
        err_msg = (errors[0].get("error") or "").lower()
        log("name" in err_msg and "required" in err_msg,
            f"error message mentions 'Name is required' (got {errors[0].get('error')!r})")

    # ---------------- Cleanup ----------------
    print("\n[Cleanup]")
    # delete the 2 imported tools
    if widget:
        r = requests.delete(f"{BASE}/tools/{widget['id']}", headers=H, timeout=30)
        log(r.status_code == 200, f"DELETE widget (got {r.status_code})")
    if screw:
        r = requests.delete(f"{BASE}/tools/{screw['id']}", headers=H, timeout=30)
        log(r.status_code == 200, f"DELETE screwdriver (got {r.status_code})")
    # delete test category if it didn't exist before
    if not cat_existed and cat_id:
        r = requests.delete(f"{BASE}/categories/{cat_id}", headers=H, timeout=30)
        log(r.status_code == 200, f"DELETE test category (got {r.status_code})")
    # delete tags if they didn't exist before
    if not tag_red_existed and red_id:
        r = requests.delete(f"{BASE}/tags/{red_id}", headers=H, timeout=30)
        log(r.status_code == 200, f"DELETE tag 'red' (got {r.status_code})")
    if not tag_blue_existed and blue_id:
        r = requests.delete(f"{BASE}/tags/{blue_id}", headers=H, timeout=30)
        log(r.status_code == 200, f"DELETE tag 'blue' (got {r.status_code})")

    print(f"\n==== RESULT:  PASS={PASS}  FAIL={FAIL} ====")
    sys.exit(0 if FAIL == 0 else 1)


if __name__ == "__main__":
    main()
